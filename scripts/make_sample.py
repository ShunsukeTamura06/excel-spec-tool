"""デモ用サンプル `.xlsx` (小売店の月次運営ブック) を生成する.

xlblueprint が抽出 / 解析 / 注釈する対象としてリアリスティックに
なるよう、以下を意図的に仕込む:

- シート 8 枚, シート間参照あり (依存グラフが面白くなる)
- Excel テーブル / 名前付き範囲 / 入力規則 (リストドロップダウン) /
  条件付き書式 / 結合セル / グラフ
- INDIRECT / OFFSET (動的数式 — risk_analyzer の検出対象)
- VLOOKUP / SUMIFS / COUNTIFS の連鎖

VBA はこの xlsx には含まれない. 別途 `scripts/inject_vba.ps1` を
Windows + Excel 環境で実行することで .bas を取り込み xlsm として保存する。

実行:
    uv run python scripts/make_sample.py

出力:
    frontend/public/samples/retail_monthly_ops.xlsx
"""

from __future__ import annotations

import datetime as dt
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================================
# データ定義
# ============================================================================

# fmt: off
PRODUCTS: list[tuple[str, str, str, int, int]] = [
    # (code, name, category, unit_price, stock)
    ("P001", "白米 5kg",              "穀物",     1980, 120),
    ("P002", "玄米 5kg",              "穀物",     2480,  60),
    ("P003", "もち米 2kg",            "穀物",     1280,  40),
    ("P004", "雑穀ミックス 1kg",      "穀物",     1580,  90),
    ("P005", "押し麦 800g",           "穀物",      780, 150),
    ("P006", "オートミール 500g",     "穀物",      680, 200),
    ("P007", "黒米 500g",             "穀物",     1180,  25),
    ("P008", "赤米 500g",             "穀物",     1280,  30),
    ("P009", "ハト麦 300g",           "穀物",      980,  45),
    ("P010", "古代米セット",          "穀物",     2880,  15),
    ("D001", "緑茶 200g",             "飲料",      980,  80),
    ("D002", "ほうじ茶 150g",         "飲料",      880,  70),
    ("D003", "玄米茶 200g",           "飲料",     1080,  55),
    ("D004", "ルイボスティー 100g",   "飲料",     1480,  30),
    ("D005", "麦茶パック 30個",       "飲料",      580, 100),
    ("S001", "海苔詰合せ",            "乾物",     1980,  18),
    ("S002", "鰹節 削り 80g",         "乾物",     1280,  35),
    ("S003", "煮干し 150g",           "乾物",      980,  42),
    ("C001", "ごま油 200ml",          "調味料",    780,  60),
    ("C002", "本みりん 500ml",        "調味料",    980,  48),
]
# fmt: on

CATEGORIES: list[str] = ["穀物", "飲料", "乾物", "調味料"]
SCENARIOS: list[str] = ["楽観", "標準", "悲観"]
STAFF: list[str] = ["田中", "鈴木", "佐藤", "高橋", "渡辺"]


# ============================================================================
# スタイル
# ============================================================================

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")  # indigo-900
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(size=14, bold=True, color="312E81")  # indigo-950
SUBTITLE_FONT = Font(size=11, bold=True, color="4338CA")  # indigo-700
NOTE_FONT = Font(size=9, color="6B7280", italic=True)
KPI_VALUE_FONT = Font(size=18, bold=True, color="1E3A8A")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _set_header(ws, row: int, values: list[str]) -> None:
    for col_idx, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col_idx, value=v)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============================================================================
# シートビルダー
# ============================================================================


def build_readme_sheet(wb: Workbook) -> None:
    """1. README — 使い方説明."""
    ws = wb.active
    ws.title = "README"

    ws["A1"] = "小売店の月次運営ブック"
    ws["A1"].font = Font(size=18, bold=True, color="1E3A8A")
    ws.merge_cells("A1:F1")

    ws["A3"] = "このブックは xlblueprint のデモ用サンプルです。"
    ws["A3"].font = NOTE_FONT

    ws["A5"] = "シート構成"
    ws["A5"].font = SUBTITLE_FONT

    sheets_doc = [
        ("設定",            "税率・為替・期間・シナリオを管理。Worksheet_Change で再計算"),
        ("商品マスタ",      "20 商品の在庫・単価・カテゴリ。Excel テーブル化"),
        ("日次売上",        "30 日 × 商品の売上ログ (ピボットの元データ)"),
        ("シナリオ計算",    "シナリオに応じた予測。INDIRECT / OFFSET を使用"),
        ("ダッシュボード",  "KPI と棒グラフ・折れ線。月次更新ボタン (VBA)"),
        ("PivotByCategory", "カテゴリ別売上集計 (SUMIFS 擬似ピボット)"),
        ("月次レポート",    "印刷向け整形出力"),
    ]
    for i, (name, desc) in enumerate(sheets_doc, start=7):
        ws.cell(row=i, column=1, value=name).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)

    ws["A16"] = "改修例 (チャットで試せる質問)"
    ws["A16"].font = SUBTITLE_FONT
    examples = [
        "・「税率を 10% から 12% に変えたとき、どこに波及しますか?」",
        "・「カテゴリ『穀物』を 2 つに分割したい。どこを直す必要がありますか?」",
        "・「日次売上を 60 日分に拡張したらシナリオ計算は追従しますか?」",
        "・「INDIRECT を使っている箇所と、その実行時の参照先を一覧で」",
    ]
    for i, line in enumerate(examples, start=17):
        ws.cell(row=i, column=1, value=line)

    _set_col_widths(ws, {"A": 22, "B": 70})


def build_settings_sheet(wb: Workbook) -> None:
    """2. 設定 — 入力規則 (ドロップダウン) と名前付き範囲の集約.

    Worksheet_Change イベントマクロが「シナリオ」セル変更を監視する想定。"""
    ws = wb.create_sheet("設定")

    ws["A1"] = "ブック設定"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    # ----- パラメータ -----
    rows = [
        ("項目",       "値",            "備考"),
        ("税率",       0.10,            "消費税率"),
        ("為替 (USD)", 150,             "USD → JPY (月平均)"),
        ("集計期間",   "標準",          "30日 / 標準 / 60日"),
        ("シナリオ",   "標準",          "楽観 / 標準 / 悲観 (シナリオ計算が参照)"),
        ("最終更新",   "(未更新)",      "Workbook_Open で自動更新"),
    ]
    for i, (label, value, note) in enumerate(rows, start=3):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=value)
        c3 = ws.cell(row=i, column=3, value=note)
        if i == 3:
            for c in (c1, c2, c3):
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
        else:
            c1.font = Font(bold=True)
            if isinstance(value, float):
                c2.number_format = "0.0%"
            elif isinstance(value, int):
                c2.number_format = "#,##0"

    # ----- 入力規則 (ドロップダウン) -----
    # 集計期間
    period_dv = DataValidation(
        type="list",
        formula1='"30日,標準,60日"',
        allow_blank=False,
        showErrorMessage=True,
    )
    period_dv.add("B6")
    ws.add_data_validation(period_dv)

    # シナリオ
    scenario_dv = DataValidation(
        type="list",
        formula1='"楽観,標準,悲観"',
        allow_blank=False,
        showErrorMessage=True,
        prompt="シナリオ計算シートが INDIRECT でこの値を参照します",
        promptTitle="シナリオ切替",
        showInputMessage=True,
    )
    scenario_dv.add("B7")
    ws.add_data_validation(scenario_dv)

    # ----- シナリオ係数表 (シナリオ計算 が INDIRECT で読みに来る) -----
    ws["E3"] = "シナリオ別係数"
    ws["E3"].font = SUBTITLE_FONT
    _set_header(ws, 4, [""] * 0)  # nothing
    ws["E4"] = "シナリオ"
    ws["E4"].fill = HEADER_FILL
    ws["E4"].font = HEADER_FONT
    ws["F4"] = "売上係数"
    ws["F4"].fill = HEADER_FILL
    ws["F4"].font = HEADER_FONT
    for i, (sc, factor) in enumerate(zip(SCENARIOS, [1.15, 1.00, 0.85], strict=True), start=5):
        ws.cell(row=i, column=5, value=sc).font = Font(bold=True)
        ws.cell(row=i, column=6, value=factor).number_format = "0.00"

    ws["E9"] = "(シナリオ計算!D2 が INDIRECT(\"設定!F\"&MATCH(...)) で引く)"
    ws["E9"].font = NOTE_FONT

    _set_col_widths(ws, {"A": 14, "B": 16, "C": 30, "D": 2, "E": 14, "F": 12})


def build_master_sheet(wb: Workbook) -> None:
    """3. 商品マスタ — Excel テーブル + 入力規則 + 条件付き書式."""
    ws = wb.create_sheet("商品マスタ")

    ws["A1"] = "商品マスタ"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    headers = ["商品コード", "商品名", "カテゴリ", "単価", "在庫数"]
    _set_header(ws, 3, headers)

    for i, (code, name, cat, price, stock) in enumerate(PRODUCTS, start=4):
        ws.cell(row=i, column=1, value=code)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=cat)
        ws.cell(row=i, column=4, value=price).number_format = "#,##0"
        ws.cell(row=i, column=5, value=stock).number_format = "#,##0"

    last_row = 3 + len(PRODUCTS)

    # Excel テーブル化 (T_Products)
    table = Table(displayName="T_Products", ref=f"A3:E{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # カテゴリのドロップダウン
    cat_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CATEGORIES)}"',
        allow_blank=False,
    )
    cat_dv.add(f"C4:C{last_row}")
    ws.add_data_validation(cat_dv)

    # 在庫アラート: 30 未満は赤
    ws.conditional_formatting.add(
        f"E4:E{last_row}",
        CellIsRule(
            operator="lessThan",
            formula=["30"],
            fill=PatternFill("solid", fgColor="FCA5A5"),
        ),
    )
    # 単価ヒートマップ
    ws.conditional_formatting.add(
        f"D4:D{last_row}",
        ColorScaleRule(
            start_type="min",
            start_color="ECFCCB",
            mid_type="percentile",
            mid_value=50,
            mid_color="FDE68A",
            end_type="max",
            end_color="FCA5A5",
        ),
    )

    _set_col_widths(ws, {"A": 12, "B": 28, "C": 12, "D": 12, "E": 12})


def build_sales_sheet(wb: Workbook) -> None:
    """4. 日次売上 — 30 日 × 商品の長い表 (ピボット用元データ)."""
    ws = wb.create_sheet("日次売上")

    ws["A1"] = "日次売上ログ"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    _set_header(ws, 3, ["日付", "商品コード", "カテゴリ", "数量", "売上金額", "担当者"])

    rnd = random.Random(20260528)  # 再現性
    start = dt.date(2026, 5, 1)
    row = 4
    for day_offset in range(30):
        d = start + dt.timedelta(days=day_offset)
        # 1 日あたり 5〜10 件の売上
        n_sales = rnd.randint(5, 10)
        for _ in range(n_sales):
            code, name, cat, price, _ = rnd.choice(PRODUCTS)
            qty = rnd.randint(1, 12)
            amount = qty * price
            ws.cell(row=row, column=1, value=d).number_format = "yyyy/mm/dd"
            ws.cell(row=row, column=2, value=code)
            ws.cell(row=row, column=3, value=cat)
            ws.cell(row=row, column=4, value=qty).number_format = "#,##0"
            ws.cell(row=row, column=5, value=amount).number_format = "#,##0"
            ws.cell(row=row, column=6, value=rnd.choice(STAFF))
            row += 1

    last_row = row - 1
    # Excel テーブル化 (T_Sales) — ピボット元データ
    table = Table(displayName="T_Sales", ref=f"A3:F{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium3",
        showRowStripes=True,
    )
    ws.add_table(table)

    _set_col_widths(ws, {"A": 12, "B": 12, "C": 12, "D": 10, "E": 14, "F": 12})


def build_scenario_calc_sheet(wb: Workbook) -> None:
    """5. シナリオ計算 — INDIRECT / OFFSET を意図的に含む.

    risk_analyzer がこのシートを「未解析リスク」として high で拾うことを期待。
    """
    ws = wb.create_sheet("シナリオ計算")

    ws["A1"] = "シナリオ別売上予測"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    # --- シナリオ係数を INDIRECT で取る ---
    ws["A2"] = "選択中シナリオ:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = "=設定!B7"
    ws["B2"].font = SUBTITLE_FONT
    ws["C2"] = "係数:"
    ws["C2"].font = Font(bold=True)
    # INDIRECT で動的に係数表を参照 (risk: dynamic_formula)
    ws["D2"] = '=INDIRECT("設定!F"&(MATCH(B2,設定!E5:E7,0)+4))'
    ws["D2"].number_format = "0.00"
    ws["D2"].font = SUBTITLE_FONT

    ws["A3"] = "↑ INDIRECT を使っているため、設定!F の構造が変わると壊れる"
    ws["A3"].font = NOTE_FONT

    # --- 商品ごとの予測表 ---
    _set_header(ws, 5, ["商品コード", "商品名", "単価", "標準月販予測", "シナリオ調整後", "在庫月数"])

    last_master_row = 3 + len(PRODUCTS)
    for i, _ in enumerate(PRODUCTS, start=6):
        master_row = i - 2  # 商品マスタ の対応行
        ws.cell(row=i, column=1, value=f"=商品マスタ!A{master_row}")
        ws.cell(row=i, column=2, value=f"=VLOOKUP(A{i},商品マスタ!$A$4:$E${last_master_row},2,FALSE)")
        ws.cell(row=i, column=3, value=f"=VLOOKUP(A{i},商品マスタ!$A$4:$E${last_master_row},4,FALSE)").number_format = "#,##0"
        # 標準月販予測 = 日次売上の SUMIFS (商品コード一致)
        ws.cell(
            row=i,
            column=4,
            value=f"=SUMIFS(日次売上!D:D,日次売上!B:B,A{i})",
        ).number_format = "#,##0"
        # シナリオ調整後 = 標準 × 係数
        ws.cell(row=i, column=5, value=f"=D{i}*$D$2").number_format = "#,##0"
        # 在庫月数 = 在庫数 / OFFSET で過去 7 日合計を使う動的範囲 (risk: dynamic_formula)
        # OFFSET(基準, 行offset, 列offset, 高さ, 幅) で「日次売上 D 列の最後 7 行」を仮想的に取る
        ws.cell(
            row=i,
            column=6,
            value=(
                f"=IF(D{i}=0,0,"
                f"VLOOKUP(A{i},商品マスタ!$A$4:$E${last_master_row},5,FALSE)/"
                f"(SUM(OFFSET(日次売上!D4,0,0,COUNTA(日次売上!D:D)-1,1))/30))"
            ),
        ).number_format = "0.0"

    last_row = 5 + len(PRODUCTS)

    # 条件付き書式: 在庫月数 6 以上を赤
    ws.conditional_formatting.add(
        f"F6:F{last_row}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["6"],
            fill=PatternFill("solid", fgColor="FCA5A5"),
        ),
    )

    _set_col_widths(ws, {"A": 12, "B": 24, "C": 10, "D": 14, "E": 16, "F": 12})


def build_pivot_by_category_sheet(wb: Workbook) -> None:
    """6. PivotByCategory — SUMIFS 擬似ピボット (本物のピボットは別途 Excel で)."""
    ws = wb.create_sheet("PivotByCategory")

    ws["A1"] = "カテゴリ別月次集計"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = "(日次売上を SUMIFS でカテゴリ集計。Excel のピボットテーブルでも代替可)"
    ws["A2"].font = NOTE_FONT

    _set_header(ws, 4, ["カテゴリ", "売上金額", "数量", "客単価 (= 売上 / 数量)"])

    for i, cat in enumerate(CATEGORIES, start=5):
        ws.cell(row=i, column=1, value=cat).font = Font(bold=True)
        ws.cell(row=i, column=2, value=f'=SUMIFS(日次売上!E:E,日次売上!C:C,A{i})').number_format = "#,##0"
        ws.cell(row=i, column=3, value=f'=SUMIFS(日次売上!D:D,日次売上!C:C,A{i})').number_format = "#,##0"
        ws.cell(row=i, column=4, value=f"=IF(C{i}=0,0,B{i}/C{i})").number_format = "#,##0"

    # 合計行
    total_row = 5 + len(CATEGORIES)
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    for col_idx in (2, 3):
        c = ws.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({get_column_letter(col_idx)}5:{get_column_letter(col_idx)}{total_row - 1})",
        )
        c.number_format = "#,##0"
        c.font = Font(bold=True)

    _set_col_widths(ws, {"A": 14, "B": 14, "C": 12, "D": 18})


def build_dashboard_sheet(wb: Workbook) -> None:
    """7. ダッシュボード — KPI + グラフ + ボタン (ボタンは inject_vba.ps1 が追加)."""
    ws = wb.create_sheet("ダッシュボード")

    ws["A1"] = "月次ダッシュボード"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    # --- KPI カード ---
    kpi_defs = [
        ("総売上",       "=SUM(日次売上!E:E)",                       "#,##0"),
        ("販売数量",     "=SUM(日次売上!D:D)",                       "#,##0"),
        ("客単価",       "=IFERROR(SUM(日次売上!E:E)/SUM(日次売上!D:D),0)", "#,##0"),
        ("在庫過剰商品", "=COUNTIF(シナリオ計算!F6:F100,\">=6\")",       "0"),
    ]
    for i, (label, formula, fmt) in enumerate(kpi_defs):
        col = 1 + i * 2
        ws.cell(row=3, column=col, value=label).font = SUBTITLE_FONT
        c = ws.cell(row=4, column=col, value=formula)
        c.number_format = fmt
        c.font = KPI_VALUE_FONT

    # --- カテゴリ別売上 (PivotByCategory を参照) ---
    ws["A7"] = "カテゴリ別売上"
    ws["A7"].font = SUBTITLE_FONT
    _set_header(ws, 8, ["カテゴリ", "売上"])
    for i, cat in enumerate(CATEGORIES, start=9):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=f"=PivotByCategory!B{i - 4}").number_format = "#,##0"

    # 棒グラフ
    bar = BarChart()
    bar.type = "col"
    bar.style = 11
    bar.title = "カテゴリ別売上"
    bar.y_axis.title = "売上金額"
    bar.x_axis.title = "カテゴリ"
    data = Reference(ws, min_col=2, min_row=8, max_row=8 + len(CATEGORIES), max_col=2)
    cats = Reference(ws, min_col=1, min_row=9, max_row=8 + len(CATEGORIES))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.dataLabels = DataLabelList(showVal=True)
    bar.height = 8
    bar.width = 16
    ws.add_chart(bar, "D7")

    # --- 担当者別売上 (折れ線) ---
    ws["A16"] = "担当者別売上"
    ws["A16"].font = SUBTITLE_FONT
    _set_header(ws, 17, ["担当者", "売上"])
    for i, person in enumerate(STAFF, start=18):
        ws.cell(row=i, column=1, value=person)
        ws.cell(
            row=i,
            column=2,
            value=f'=SUMIFS(日次売上!E:E,日次売上!F:F,A{i})',
        ).number_format = "#,##0"

    line = LineChart()
    line.title = "担当者別売上"
    line.style = 12
    data2 = Reference(ws, min_col=2, min_row=17, max_row=17 + len(STAFF), max_col=2)
    cats2 = Reference(ws, min_col=1, min_row=18, max_row=17 + len(STAFF))
    line.add_data(data2, titles_from_data=True)
    line.set_categories(cats2)
    line.height = 8
    line.width = 16
    ws.add_chart(line, "D16")

    # --- ボタン用プレースホルダ (inject_vba.ps1 でフォームコントロール追加) ---
    ws["A26"] = "[ボタン: 月次更新] ← inject_vba.ps1 でフォームコントロールが追加され、MonthlyUpdate にリンクします"
    ws["A26"].font = NOTE_FONT

    _set_col_widths(ws, {"A": 16, "B": 14, "C": 16, "D": 14, "E": 16, "F": 14})


def build_report_sheet(wb: Workbook) -> None:
    """8. 月次レポート — 印刷向け整形 (結合セル多用 + 他シート参照)."""
    ws = wb.create_sheet("月次レポート")

    ws["A1"] = "月次運営レポート"
    ws["A1"].font = Font(size=20, bold=True, color="1E3A8A")
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "=\"対象期間: \"&設定!B6"
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:F2")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPI を 3 列で配置
    ws["A4"] = "総売上"
    ws["A4"].font = SUBTITLE_FONT
    ws.merge_cells("A4:B4")
    ws["A5"] = "=ダッシュボード!A4"
    ws["A5"].font = KPI_VALUE_FONT
    ws["A5"].number_format = "#,##0"
    ws.merge_cells("A5:B5")

    ws["C4"] = "販売数量"
    ws["C4"].font = SUBTITLE_FONT
    ws.merge_cells("C4:D4")
    ws["C5"] = "=ダッシュボード!C4"
    ws["C5"].font = KPI_VALUE_FONT
    ws["C5"].number_format = "#,##0"
    ws.merge_cells("C5:D5")

    ws["E4"] = "客単価"
    ws["E4"].font = SUBTITLE_FONT
    ws.merge_cells("E4:F4")
    ws["E5"] = "=ダッシュボード!E4"
    ws["E5"].font = KPI_VALUE_FONT
    ws["E5"].number_format = "#,##0"
    ws.merge_cells("E5:F5")

    # カテゴリ別表
    ws["A8"] = "カテゴリ別内訳"
    ws["A8"].font = SUBTITLE_FONT
    ws.merge_cells("A8:F8")
    _set_header(ws, 9, ["カテゴリ", "売上金額", "数量", "客単価", "構成比", ""])
    for i, cat in enumerate(CATEGORIES, start=10):
        src_row = i - 5  # PivotByCategory の対応行
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=f"=PivotByCategory!B{src_row}").number_format = "#,##0"
        ws.cell(row=i, column=3, value=f"=PivotByCategory!C{src_row}").number_format = "#,##0"
        ws.cell(row=i, column=4, value=f"=PivotByCategory!D{src_row}").number_format = "#,##0"
        ws.cell(row=i, column=5, value=f"=IFERROR(B{i}/PivotByCategory!B{src_row + len(CATEGORIES)},0)").number_format = "0.0%"

    ws["A16"] = "備考"
    ws["A16"].font = SUBTITLE_FONT
    ws["A17"] = '=" シナリオ: "&設定!B7&" / 税率: "&TEXT(設定!B4,"0.0%")'
    ws.merge_cells("A17:F17")

    _set_col_widths(ws, {"A": 14, "B": 14, "C": 12, "D": 12, "E": 12, "F": 4})


# ============================================================================
# 名前付き範囲
# ============================================================================


def add_named_ranges(wb: Workbook) -> None:
    last_master_row = 3 + len(PRODUCTS)
    wb.defined_names["商品マスタ"] = DefinedName(
        name="商品マスタ",
        attr_text=f"商品マスタ!$A$4:$E${last_master_row}",
    )
    wb.defined_names["税率"] = DefinedName(
        name="税率",
        attr_text="設定!$B$4",
    )
    wb.defined_names["シナリオ"] = DefinedName(
        name="シナリオ",
        attr_text="設定!$B$7",
    )


# ============================================================================
# エントリポイント
# ============================================================================


def main() -> Path:
    wb = Workbook()
    build_readme_sheet(wb)
    build_settings_sheet(wb)
    build_master_sheet(wb)
    build_sales_sheet(wb)
    build_scenario_calc_sheet(wb)
    build_pivot_by_category_sheet(wb)
    build_dashboard_sheet(wb)
    build_report_sheet(wb)
    add_named_ranges(wb)

    out = (
        Path(__file__).resolve().parents[1]
        / "frontend"
        / "public"
        / "samples"
        / "retail_monthly_ops.xlsx"
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


if __name__ == "__main__":
    p = main()
    print(f"wrote: {p} ({p.stat().st_size:,} bytes)")
