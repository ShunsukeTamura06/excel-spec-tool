"""frontend_streamlit.api_client のテスト (LEGACY).

httpx の ASGITransport を使って、Frontend クライアントを実 Backend スタックに
直結させる。HTTP の往復は経由しないが、FastAPI のルーティング・依存性・
シリアライゼーションは本物が走る。
"""

from collections.abc import Iterator
from pathlib import Path

import httpx
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook as OpyWorkbook

from backend.dependencies import get_llm_client, get_storage
from backend.llm_client import MockLLMClient
from backend.main import create_app
from backend.storage import Storage
from frontend_streamlit.api_client import BackendClient, BackendError


@pytest.fixture
def backend_storage(tmp_path: Path) -> Storage:
    return Storage(tmp_path / "jobs")


@pytest.fixture
def api(backend_storage: Storage) -> Iterator[BackendClient]:
    """FastAPI TestClient を内部 http_client として直結させた BackendClient.

    TestClient は httpx.Client を継承しており、ASGI アプリに対し同期的に
    リクエストできるため、Frontend の BackendClient (sync) からそのまま使える。
    """
    app = create_app()
    app.dependency_overrides[get_storage] = lambda: backend_storage
    app.dependency_overrides[get_llm_client] = lambda: MockLLMClient()
    test_client = TestClient(app)
    client = BackendClient(base_url="http://testserver", http_client=test_client)
    try:
        yield client
    finally:
        client.close()
        test_client.close()
        app.dependency_overrides.clear()


@pytest.fixture
def xlsx_bytes(tmp_path: Path) -> bytes:
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Calc"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    out = tmp_path / "x.xlsx"
    wb.save(out)
    return out.read_bytes()


# ---------- 正常系 (end-to-end フロー) ----------


class TestEndToEnd:
    def test_full_flow(self, api: BackendClient, xlsx_bytes: bytes) -> None:
        # health
        assert api.health() is True

        # extract
        job_id = api.extract("sample.xlsx", xlsx_bytes)
        assert job_id

        # list
        jobs = api.list_jobs()
        assert any(j["job_id"] == job_id for j in jobs)

        # analyze
        result = api.analyze(job_id)
        assert result == {"status": "ok"}

        # spec
        spec = api.get_spec(job_id)
        assert "spec_md" in spec
        assert "# 設計書" in spec["spec_md"]
        assert spec["meta"]["filename"] == "sample.xlsx"

        # references (=SUM(A1:A2) があるので "A1:A2" が逆引きヒット)
        refs = api.get_references(job_id, "A1:A2")
        assert len(refs) >= 1
        assert refs[0]["from"]  # alias
        assert refs[0]["kind"] == "formula"

        # references (該当なし)
        assert api.get_references(job_id, "NoSuch!Z9") == []

        # chat
        chat_result = api.chat(job_id, "How do I update A3?")
        assert "reply" in chat_result
        assert "[mock" in chat_result["reply"]
        assert len(chat_result["history"]) == 2

        # history
        history = api.get_chat_history(job_id)
        assert len(history) == 2

        # delete
        assert api.delete_job(job_id) is True
        assert all(j["job_id"] != job_id for j in api.list_jobs())


# ---------- エラー応答 ----------


class TestErrors:
    def test_invalid_job_id_raises_backend_error_400(self, api: BackendClient) -> None:
        with pytest.raises(BackendError) as exc:
            api.get_spec("not-a-uuid")
        assert exc.value.status_code == 400

    def test_missing_job_raises_404(self, api: BackendClient) -> None:
        import uuid

        with pytest.raises(BackendError) as exc:
            api.get_spec(str(uuid.uuid4()))
        assert exc.value.status_code == 404

    def test_corrupt_file_raises_422(self, api: BackendClient) -> None:
        with pytest.raises(BackendError) as exc:
            api.extract("bad.xlsx", b"not real xlsx")
        assert exc.value.status_code == 422

    def test_empty_file_raises_400(self, api: BackendClient) -> None:
        with pytest.raises(BackendError) as exc:
            api.extract("a.xlsx", b"")
        assert exc.value.status_code == 400

    def test_chat_invalid_id_raises_400(self, api: BackendClient) -> None:
        with pytest.raises(BackendError) as exc:
            api.chat("not-uuid", "hi")
        assert exc.value.status_code == 400


# ---------- BackendClient 単体 ----------


class TestBackendClientConfig:
    def test_default_base_url(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.delenv("BACKEND_URL", raising=False)
        c = BackendClient()
        assert c.base_url == "http://localhost:8000"
        c.close()

    def test_env_overrides_default(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("BACKEND_URL", "http://internal.example/api")
        c = BackendClient()
        assert c.base_url == "http://internal.example/api"
        c.close()

    def test_explicit_base_url_wins_over_env(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("BACKEND_URL", "http://env-url")
        c = BackendClient(base_url="http://explicit")
        assert c.base_url == "http://explicit"
        c.close()

    def test_trailing_slash_stripped(self) -> None:
        c = BackendClient(base_url="http://x/")
        assert c.base_url == "http://x"
        c.close()

    def test_url_path_normalization(self) -> None:
        c = BackendClient(base_url="http://x")
        assert c._url("/foo") == "http://x/foo"
        assert c._url("foo") == "http://x/foo"
        c.close()

    def test_context_manager_closes_owned_client(self) -> None:
        with BackendClient(base_url="http://x") as c:
            assert c.base_url == "http://x"
        # close 後は内部 client が閉じられている (例外を直接観測するのは難しいが
        # is_closed 属性を見れる)
        assert c._client.is_closed

    def test_does_not_close_external_client(self) -> None:
        external = httpx.Client()
        c = BackendClient(base_url="http://x", http_client=external)
        c.close()
        # 外部から渡された client は閉じない
        assert not external.is_closed
        external.close()


class TestBackendErrorParsing:
    def test_health_returns_false_on_failure(self) -> None:
        # 存在しない URL なので health は False
        c = BackendClient(base_url="http://127.0.0.1:1")
        assert c.health() is False
        c.close()
