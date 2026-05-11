"""core.extractors.cells のテスト."""

import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook as OpyWorkbook

from core.exceptions import ExtractionError
from core.extractors.cells import extract_cells_to_sqlite


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """ヘッダ+データの小さな表を持つ xlsx を生成."""
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Portfolio"
    # ヘッダ行 (1行目)
    ws["A1"] = "銘柄コード"
    ws["B1"] = "銘柄名"
    ws["C1"] = "保有口数"
    ws["D1"] = "現在値"
    ws["E1"] = "評価損益"
    # データ行
    ws["A2"] = "ABC"
    ws["B2"] = "株式会社A"
    ws["C2"] = 100
    ws["D2"] = 1500
    ws["E2"] = "=C2*D2"
    ws["A3"] = "DEF"
    ws["B3"] = "株式会社D"
    ws["C3"] = 200
    ws["D3"] = 800
    ws["E3"] = "=C3*D3"

    # マージセルを足す
    ws.merge_cells("A5:E5")
    ws["A5"] = "備考欄"

    # もう1つシート
    ws2 = wb.create_sheet("Settings")
    ws2["A1"] = "税率"
    ws2["B1"] = 0.1

    out = tmp_path / "sample.xlsx"
    wb.save(out)
    return out


# ---------- 正常系 ----------


class TestExtractCells:
    def test_creates_db_file(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        n = extract_cells_to_sqlite(sample_xlsx, db)
        assert db.is_file()
        assert n > 0

    def test_inserts_all_non_empty_cells(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute("SELECT COUNT(*) FROM cells")
            count = cur.fetchone()[0]
            # ヘッダ5 + データ2行×5 + Settings 2 + 備考欄 1 = 18
            assert count >= 16
        finally:
            conn.close()

    def test_string_cells_stored_as_text(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute(
                "SELECT value FROM cells WHERE sheet=? AND coord=?", ("Portfolio", "A1")
            )
            row = cur.fetchone()
            assert row[0] == "銘柄コード"
        finally:
            conn.close()

    def test_number_cells_stored_as_string(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute(
                "SELECT value, data_type FROM cells WHERE sheet=? AND coord=?",
                ("Portfolio", "C2"),
            )
            row = cur.fetchone()
            assert row[0] == "100"
            assert row[1] == "n"
        finally:
            conn.close()

    def test_formula_cells_have_formula_text(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute(
                "SELECT formula, data_type FROM cells WHERE sheet=? AND coord=?",
                ("Portfolio", "E2"),
            )
            row = cur.fetchone()
            assert row[0] == "=C2*D2"
            assert row[1] == "f"
        finally:
            conn.close()

    def test_merged_ranges_stored(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute("SELECT range FROM merged_ranges WHERE sheet=?", ("Portfolio",))
            rows = [r[0] for r in cur.fetchall()]
            assert "A5:E5" in rows
        finally:
            conn.close()

    def test_overwrites_existing_db(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        # 既存ファイルに別データを入れておく
        db.write_bytes(b"garbage")
        # それでも正常に上書きできる
        n = extract_cells_to_sqlite(sample_xlsx, db)
        assert n > 0

    def test_multiple_sheets(self, sample_xlsx: Path, tmp_path: Path) -> None:
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute("SELECT DISTINCT sheet FROM cells ORDER BY sheet")
            sheets = [r[0] for r in cur.fetchall()]
            assert sheets == ["Portfolio", "Settings"]
        finally:
            conn.close()

    def test_querying_a_row(self, sample_xlsx: Path, tmp_path: Path) -> None:
        # SQLite から1行をまとめて引けること
        db = tmp_path / "cells.db"
        extract_cells_to_sqlite(sample_xlsx, db)
        conn = sqlite3.connect(db)
        try:
            cur = conn.execute(
                "SELECT col, value FROM cells WHERE sheet=? AND row=? ORDER BY col",
                ("Portfolio", 1),
            )
            row = cur.fetchall()
            values = [r[1] for r in row]
            assert "銘柄コード" in values
            assert "現在値" in values
        finally:
            conn.close()


# ---------- エラー系 ----------


class TestExtractCellsErrors:
    def test_file_not_found(self, tmp_path: Path) -> None:
        with pytest.raises(ExtractionError):
            extract_cells_to_sqlite(tmp_path / "no_such.xlsx", tmp_path / "x.db")

    def test_xls_raises(self, tmp_path: Path) -> None:
        fake = tmp_path / "legacy.xls"
        fake.write_bytes(b"not real")
        with pytest.raises(ExtractionError):
            extract_cells_to_sqlite(fake, tmp_path / "x.db")

    def test_corrupt_xlsx(self, tmp_path: Path) -> None:
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not real xlsx")
        with pytest.raises(ExtractionError):
            extract_cells_to_sqlite(bad, tmp_path / "x.db")
