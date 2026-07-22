"""core.workbook_diff のテスト."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from core.exceptions import DiffError
from core.models import (
    AnalysisRisk,
    CellDiff,
    ChartObject,
    ChartSeries,
    ConditionalFormat,
    DataValidation,
    NamedRange,
    NamedRangeDiff,
    PivotTableInfo,
    Reference,
    ReferenceIndex,
    SheetInfo,
    VbaModule,
    Workbook,
)
from core.workbook_diff import (
    _diff_cells,
    _diff_charts,
    _diff_conditional_formats,
    _diff_data_validations,
    _diff_pivot_tables,
    _diff_vba_modules,
    _same_formula_with_cache_only_change,
    build_blast_radius,
    diff_named_ranges,
    diff_workbooks,
)


def _write_xlsx(path: Path, cells: dict[str, object], sheet_name: str = "Sheet") -> None:
    """簡易xlsx生成ヘルパー. cells は {"A1": 値または数式文字列, ...}."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    for coord, value in cells.items():
        ws[coord] = value
    wb.save(path)


class TestCellDiff:
    def test_added_cell(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        after = tmp_path / "after.xlsx"
        _write_xlsx(before, {})
        _write_xlsx(after, {"A1": "x"})

        diffs = _diff_cells(before, after)
        assert len(diffs) == 1
        assert diffs[0].sheet == "Sheet"
        assert diffs[0].coord == "A1"
        assert diffs[0].change_type == "added"
        assert diffs[0].after_value == "x"

    def test_removed_cell(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        after = tmp_path / "after.xlsx"
        _write_xlsx(before, {"A1": "x"})
        _write_xlsx(after, {})

        diffs = _diff_cells(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "removed"
        assert diffs[0].before_value == "x"

    def test_modified_formula(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        after = tmp_path / "after.xlsx"
        _write_xlsx(before, {"A1": 1, "B1": "=A1+1"})
        _write_xlsx(after, {"A1": 1, "B1": "=A1+2"})

        diffs = _diff_cells(before, after)
        assert len(diffs) == 1
        assert diffs[0].coord == "B1"
        assert diffs[0].change_type == "modified"
        assert diffs[0].before_formula == "=A1+1"
        assert diffs[0].after_formula == "=A1+2"

    def test_modified_value_only(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        after = tmp_path / "after.xlsx"
        _write_xlsx(before, {"A1": 1})
        _write_xlsx(after, {"A1": 2})

        diffs = _diff_cells(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"
        assert diffs[0].before_value == "1"
        assert diffs[0].after_value == "2"
        assert diffs[0].before_formula is None
        assert diffs[0].after_formula is None

    def test_modified_number_format_only(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        after = tmp_path / "after.xlsx"
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        assert ws1 is not None
        ws1["A1"] = 1
        ws1["A1"].number_format = "0.00"
        wb1.save(before)

        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        assert ws2 is not None
        ws2["A1"] = 1
        ws2["A1"].number_format = "0.0%"
        wb2.save(after)

        diffs = _diff_cells(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"
        assert diffs[0].before_number_format == "0.00"
        assert diffs[0].after_number_format == "0.0%"

    def test_unchanged_cell_not_reported(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        after = tmp_path / "after.xlsx"
        _write_xlsx(before, {"A1": 1, "B1": "hello"})
        _write_xlsx(after, {"A1": 1, "B1": "hello"})

        assert _diff_cells(before, after) == []

    def test_same_formula_cache_update_is_not_reported(self) -> None:
        """OfficeCLIが再計算キャッシュだけを保存しても構造変更に数えない."""

        assert _same_formula_with_cache_only_change(
            (None, "=SUM(A1:A3)", "#,##0"),
            ("6", "=SUM(A1:A3)", "#,##0"),
        )
        assert not _same_formula_with_cache_only_change(
            (None, "=SUM(A1:A3)", "#,##0"),
            ("6", "=SUM(A1:A4)", "#,##0"),
        )


class TestNamedRangeDiff:
    def _wb(self, named_ranges: list[NamedRange]) -> Workbook:
        return Workbook(
            filename="x.xlsx",
            sheets=[SheetInfo(name="Sheet1", rows=1, cols=1, named_ranges=named_ranges)],
        )

    def test_added(self) -> None:
        before = self._wb([])
        after = self._wb([NamedRange(name="税率", refers_to="設定!$B$4")])
        diffs = diff_named_ranges(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "added"
        assert diffs[0].after_refers_to == "設定!$B$4"

    def test_removed(self) -> None:
        before = self._wb([NamedRange(name="税率", refers_to="設定!$B$4")])
        after = self._wb([])
        diffs = diff_named_ranges(before, after)
        assert diffs[0].change_type == "removed"

    def test_modified(self) -> None:
        before = self._wb([NamedRange(name="税率", refers_to="設定!$B$4")])
        after = self._wb([NamedRange(name="税率", refers_to="設定!$B$5")])
        diffs = diff_named_ranges(before, after)
        assert diffs[0].change_type == "modified"
        assert diffs[0].before_refers_to == "設定!$B$4"
        assert diffs[0].after_refers_to == "設定!$B$5"

    def test_unchanged_not_reported(self) -> None:
        wb1 = self._wb([NamedRange(name="税率", refers_to="設定!$B$4")])
        wb2 = self._wb([NamedRange(name="税率", refers_to="設定!$B$4")])
        assert diff_named_ranges(wb1, wb2) == []


class TestConditionalFormatDiff:
    def _wb(self, cfs: list[ConditionalFormat]) -> Workbook:
        return Workbook(
            filename="x.xlsx",
            sheets=[SheetInfo(name="Sheet1", rows=1, cols=1, conditional_formats=cfs)],
        )

    def test_rule_changed_on_same_range(self) -> None:
        before = self._wb([ConditionalFormat(range="A1:A10", rule="cellIs greaterThan 0")])
        after = self._wb([ConditionalFormat(range="A1:A10", rule="cellIs greaterThan 10")])
        diffs = _diff_conditional_formats(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"

    def test_added(self) -> None:
        before = self._wb([])
        after = self._wb([ConditionalFormat(range="A1:A10", rule="cellIs greaterThan 0")])
        diffs = _diff_conditional_formats(before, after)
        assert diffs[0].change_type == "added"

    def test_removed(self) -> None:
        before = self._wb([ConditionalFormat(range="A1:A10", rule="cellIs greaterThan 0")])
        after = self._wb([])
        diffs = _diff_conditional_formats(before, after)
        assert diffs[0].change_type == "removed"


class TestDataValidationDiff:
    def _wb(self, dvs: list[DataValidation]) -> Workbook:
        return Workbook(
            filename="x.xlsx",
            sheets=[SheetInfo(name="Sheet1", rows=1, cols=1, data_validations=dvs)],
        )

    def test_type_changed_on_same_range(self) -> None:
        before = self._wb([DataValidation(range="A2:A100", type="list", formula='"a,b,c"')])
        after = self._wb([DataValidation(range="A2:A100", type="whole", formula="1")])
        diffs = _diff_data_validations(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"

    def test_added(self) -> None:
        before = self._wb([])
        after = self._wb([DataValidation(range="A2:A100", type="list", formula='"a,b,c"')])
        diffs = _diff_data_validations(before, after)
        assert diffs[0].change_type == "added"

    def test_removed(self) -> None:
        before = self._wb([DataValidation(range="A2:A100", type="list", formula='"a,b,c"')])
        after = self._wb([])
        diffs = _diff_data_validations(before, after)
        assert diffs[0].change_type == "removed"


class TestChartDiff:
    def _wb(self, charts: list[ChartObject]) -> Workbook:
        return Workbook(
            filename="x.xlsx",
            sheets=[SheetInfo(name="Sheet1", rows=1, cols=1, charts=charts)],
        )

    def test_series_ref_changed(self) -> None:
        before = self._wb(
            [ChartObject(name="Chart1", anchor="E2", series=[ChartSeries(values_ref="A1:A10")])]
        )
        after = self._wb(
            [ChartObject(name="Chart1", anchor="E2", series=[ChartSeries(values_ref="A1:A20")])]
        )
        diffs = _diff_charts(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"
        assert diffs[0].key == "Chart1"

    def test_anonymous_charts_disambiguated_by_anchor(self) -> None:
        before = self._wb([ChartObject(name="", anchor="E2"), ChartObject(name="", anchor="E20")])
        after = self._wb(
            [
                ChartObject(name="", anchor="E2"),
                ChartObject(name="", anchor="E20", title="変更後"),
            ]
        )
        diffs = _diff_charts(before, after)
        assert len(diffs) == 1
        assert diffs[0].key == "E20"

    def test_added(self) -> None:
        before = self._wb([])
        after = self._wb([ChartObject(name="Chart1", anchor="E2")])
        diffs = _diff_charts(before, after)
        assert diffs[0].change_type == "added"

    def test_removed(self) -> None:
        before = self._wb([ChartObject(name="Chart1", anchor="E2")])
        after = self._wb([])
        diffs = _diff_charts(before, after)
        assert diffs[0].change_type == "removed"


class TestPivotTableDiff:
    def _wb(self, pts: list[PivotTableInfo]) -> Workbook:
        return Workbook(
            filename="x.xlsx",
            sheets=[SheetInfo(name="Sheet1", rows=1, cols=1, pivot_tables=pts)],
        )

    def test_source_ref_changed(self) -> None:
        before = self._wb([PivotTableInfo(name="Pivot1", source_ref="Data!A1:D100")])
        after = self._wb([PivotTableInfo(name="Pivot1", source_ref="Data!A1:D200")])
        diffs = _diff_pivot_tables(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"

    def test_added(self) -> None:
        before = self._wb([])
        after = self._wb([PivotTableInfo(name="Pivot1")])
        diffs = _diff_pivot_tables(before, after)
        assert diffs[0].change_type == "added"

    def test_removed(self) -> None:
        before = self._wb([PivotTableInfo(name="Pivot1")])
        after = self._wb([])
        diffs = _diff_pivot_tables(before, after)
        assert diffs[0].change_type == "removed"


class TestVbaModuleDiff:
    def _wb(self, modules: list[VbaModule]) -> Workbook:
        return Workbook(filename="x.xlsm", vba_modules=modules)

    def test_code_changed(self) -> None:
        before = self._wb([VbaModule(name="Module1", type="Module", code="Sub Foo()\nEnd Sub")])
        after = self._wb(
            [VbaModule(name="Module1", type="Module", code="Sub Foo()\nMsgBox 1\nEnd Sub")]
        )
        diffs = _diff_vba_modules(before, after)
        assert len(diffs) == 1
        assert diffs[0].change_type == "modified"

    def test_added_module(self) -> None:
        before = self._wb([])
        after = self._wb([VbaModule(name="Module1", type="Module", code="")])
        diffs = _diff_vba_modules(before, after)
        assert diffs[0].change_type == "added"

    def test_removed_module(self) -> None:
        before = self._wb([VbaModule(name="Module1", type="Module", code="")])
        after = self._wb([])
        diffs = _diff_vba_modules(before, after)
        assert diffs[0].change_type == "removed"

    def test_unchanged_module_not_reported(self) -> None:
        before = self._wb([VbaModule(name="Module1", type="Module", code="Sub Foo()\nEnd Sub")])
        after = self._wb([VbaModule(name="Module1", type="Module", code="Sub Foo()\nEnd Sub")])
        assert _diff_vba_modules(before, after) == []


class TestBlastRadius:
    def test_removed_cell_still_referenced(self) -> None:
        idx = ReferenceIndex(
            refs={"Input!A:A": [Reference(kind="formula", from_="Calc!H2", to="Input!A:A")]}
        )
        cells = [CellDiff(sheet="Input", coord="A1", change_type="removed", before_value="1")]
        entries = build_blast_radius(cells, [], idx)
        assert len(entries) == 1
        assert entries[0].location == "Input!A1"
        assert entries[0].referenced_by[0].from_ == "Calc!H2"

    def test_modified_cell_with_no_references_not_reported(self) -> None:
        idx = ReferenceIndex(refs={})
        cells = [CellDiff(sheet="Input", coord="A1", change_type="modified")]
        assert build_blast_radius(cells, [], idx) == []

    def test_added_cell_excluded_even_if_matching_ref_exists(self) -> None:
        idx = ReferenceIndex(
            refs={"Input!A1": [Reference(kind="formula", from_="Calc!H2", to="Input!A1")]}
        )
        cells = [CellDiff(sheet="Input", coord="A1", change_type="added")]
        assert build_blast_radius(cells, [], idx) == []

    def test_removed_named_range_still_referenced(self) -> None:
        idx = ReferenceIndex(
            refs={"設定!B4": [Reference(kind="formula", from_="Calc!C1", to="設定!B4")]}
        )
        named_ranges = [
            NamedRangeDiff(name="税率", change_type="removed", before_refers_to="設定!$B$4")
        ]
        entries = build_blast_radius([], named_ranges, idx)
        assert len(entries) == 1
        assert entries[0].location == "設定!$B$4"
        assert entries[0].referenced_by[0].from_ == "Calc!C1"


class TestNoChange:
    def test_identical_workbooks_produce_empty_diff(self, tmp_path: Path) -> None:
        before = tmp_path / "a.xlsx"
        after = tmp_path / "b.xlsx"
        _write_xlsx(before, {"A1": 1, "B1": "=A1+1"})
        _write_xlsx(after, {"A1": 1, "B1": "=A1+1"})

        wb = Workbook(filename="a.xlsx", sheets=[SheetInfo(name="Sheet", rows=1, cols=2)])
        diff = diff_workbooks(before, after, wb, wb, ReferenceIndex())
        assert diff.is_empty()
        assert diff.blast_radius == []


class TestNoiseAvoidance:
    """spikes/xlsx_diff_noise/probe.py で検証済みの仮説をCI常設テストに昇格."""

    def test_resave_without_edits_produces_no_cell_diff(self, tmp_path: Path) -> None:
        baseline = tmp_path / "baseline.xlsx"
        _write_xlsx(baseline, {"A1": 1, "B1": "=A1+1", "C1": "hello"})

        resaved = tmp_path / "resaved.xlsx"
        wb = openpyxl.load_workbook(baseline)
        wb.save(resaved)
        wb.close()

        assert _diff_cells(baseline, resaved) == []

    def test_single_formula_edit_isolates_exact_cell(self, tmp_path: Path) -> None:
        baseline = tmp_path / "baseline.xlsx"
        _write_xlsx(baseline, {"A1": 1, "B1": "=A1+1", "C1": "hello"})

        edited = tmp_path / "edited.xlsx"
        wb = openpyxl.load_workbook(baseline)
        wb["Sheet"]["B1"] = "=A1+2"
        wb.save(edited)
        wb.close()

        diffs = _diff_cells(baseline, edited)
        assert len(diffs) == 1
        assert diffs[0].sheet == "Sheet"
        assert diffs[0].coord == "B1"
        assert diffs[0].change_type == "modified"


class TestExistingRisksPassthrough:
    def test_before_analysis_risks_carried_into_diff(self, tmp_path: Path) -> None:
        before = tmp_path / "a.xlsx"
        after = tmp_path / "b.xlsx"
        _write_xlsx(before, {})
        _write_xlsx(after, {})

        risk = AnalysisRisk(
            category="dynamic_formula",
            severity="medium",
            location="Sheet!A1",
            evidence="INDIRECT(...)",
            description="動的参照",
            recommendation="確認してください",
        )
        wb_before = Workbook(filename="a.xlsx", analysis_risks=[risk])
        wb_after = Workbook(filename="b.xlsx")

        diff = diff_workbooks(before, after, wb_before, wb_after, ReferenceIndex())
        assert diff.existing_risks == [risk]

    def test_new_risk_introduced_by_change_is_also_surfaced(self, tmp_path: Path) -> None:
        """変更(例: VBA置換)で新たに生じたリスクは、before側になくても診断に残る.

        変更前は無かったリスクが変更後に増えても before_wb.analysis_risks だけを
        再掲すると見逃す。verify_expected_diff の needs_review 判定にも影響するため、
        before/after 双方のリスクを統合しておく必要がある。
        """
        before = tmp_path / "a.xlsx"
        after = tmp_path / "b.xlsx"
        _write_xlsx(before, {})
        _write_xlsx(after, {})

        new_risk = AnalysisRisk(
            category="dynamic_vba",
            severity="high",
            location="Module1.UpdateReport",
            evidence="Range(cellRef).Value = 1",
            description="置換後コードに動的参照が追加された",
            recommendation="手動確認してください",
        )
        wb_before = Workbook(filename="a.xlsx")
        wb_after = Workbook(filename="b.xlsx", analysis_risks=[new_risk])

        diff = diff_workbooks(before, after, wb_before, wb_after, ReferenceIndex())

        assert diff.existing_risks == [new_risk]

    def test_duplicate_risks_between_before_and_after_are_not_repeated(
        self, tmp_path: Path
    ) -> None:
        before = tmp_path / "a.xlsx"
        after = tmp_path / "b.xlsx"
        _write_xlsx(before, {})
        _write_xlsx(after, {})

        risk = AnalysisRisk(
            category="external_dependency",
            severity="low",
            location="Sheet!A1",
            evidence="外部接続",
            description="外部接続あり",
            recommendation="確認してください",
        )
        wb_before = Workbook(filename="a.xlsx", analysis_risks=[risk])
        wb_after = Workbook(filename="b.xlsx", analysis_risks=[risk])

        diff = diff_workbooks(before, after, wb_before, wb_after, ReferenceIndex())

        assert diff.existing_risks == [risk]


class TestErrors:
    def test_missing_before_file_raises_diff_error(self, tmp_path: Path) -> None:
        after = tmp_path / "after.xlsx"
        _write_xlsx(after, {})
        with pytest.raises(DiffError):
            _diff_cells(tmp_path / "missing.xlsx", after)

    def test_xls_input_raises_diff_error(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xls"
        before.write_bytes(b"not a real xls")
        after = tmp_path / "after.xlsx"
        _write_xlsx(after, {})
        with pytest.raises(DiffError):
            _diff_cells(before, after)
