"""core.extractors.workbook のテスト.

xlsx は openpyxl で動的に作って tmp_path に書き出し、それを読み戻して検証する。
バイナリの xlsx fixture をコミットせずに済ませる。
"""

import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook as OpyWorkbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName

from core.exceptions import ExtractionError
from core.extractors.workbook import (
    _extract_formula_refs,
    _extract_sheet_name,
    extract_workbook,
)

# ---------- 内部ヘルパーの単体テスト ----------


class TestExtractSheetName:
    def test_simple(self) -> None:
        assert _extract_sheet_name("Sheet1!$A$1") == "Sheet1"

    def test_quoted_with_space(self) -> None:
        assert _extract_sheet_name("'My Sheet'!A1:B2") == "My Sheet"

    def test_no_sheet(self) -> None:
        assert _extract_sheet_name("$A$1") is None

    def test_empty(self) -> None:
        assert _extract_sheet_name("") is None


class TestExtractFormulaRefs:
    def test_simple_sum(self) -> None:
        refs = _extract_formula_refs("=SUM(A1:A10)")
        assert "A1:A10" in refs

    def test_cross_sheet(self) -> None:
        refs = _extract_formula_refs("=Sheet2!A1")
        assert "Sheet2!A1" in refs

    def test_sumif_three_refs(self) -> None:
        refs = _extract_formula_refs("=SUMIF(Input!A:A, A2, Input!E:E)")
        assert "Input!A:A" in refs
        assert "A2" in refs
        assert "Input!E:E" in refs

    def test_no_leading_equals(self) -> None:
        # 先頭の "=" が無くても正しく動くこと (補完される)
        refs = _extract_formula_refs("SUM(B1:B5)")
        assert "B1:B5" in refs

    def test_invalid_formula_returns_empty(self) -> None:
        # Tokenizer が壊れる入力でも例外を漏らさないこと
        # 完全に空でもクラッシュしない
        assert _extract_formula_refs("") == []


# ---------- extract_workbook の統合テスト ----------


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    """A1=10, A2=20, A3=数式, B1=数式 の xlsx を生成."""
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Calc"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    ws["B1"] = "=A3*2"
    out = tmp_path / "simple.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def cross_sheet_xlsx(tmp_path: Path) -> Path:
    """Sheet1.A1=10, Sheet2.B1=`=Sheet1!A1` の xlsx を生成."""
    wb = OpyWorkbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Input"
    ws1["A1"] = 10

    ws2 = wb.create_sheet("Calc")
    ws2["B1"] = "=Input!A1"
    ws2["B2"] = "=SUMIF(Input!A:A, B1, Input!A:A)"

    out = tmp_path / "cross.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def named_range_xlsx(tmp_path: Path) -> Path:
    """名前付き範囲を持つ xlsx."""
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = 1
    wb.defined_names["TaxRate"] = DefinedName(name="TaxRate", attr_text="Data!$A$1")

    out = tmp_path / "named.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def conditional_format_xlsx(tmp_path: Path) -> Path:
    """条件付き書式を含む xlsx."""
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")
    rule = CellIsRule(operator="greaterThan", formula=["100"], fill=fill)
    ws.conditional_formatting.add("A1:A10", rule)

    out = tmp_path / "cf.xlsx"
    wb.save(out)
    return out


class TestExtractFormulas:
    def test_simple(self, simple_xlsx: Path) -> None:
        wb = extract_workbook(simple_xlsx)
        assert wb.filename == "simple.xlsx"
        assert len(wb.sheets) == 1
        sheet = wb.sheets[0]
        assert sheet.name == "Calc"

        coords = {f.coord for f in sheet.formulas}
        assert coords == {"Calc!A3", "Calc!B1"}

        a3 = next(f for f in sheet.formulas if f.coord == "Calc!A3")
        assert a3.formula == "=SUM(A1:A2)"
        assert "A1:A2" in a3.refs

        b1 = next(f for f in sheet.formulas if f.coord == "Calc!B1")
        assert b1.formula == "=A3*2"
        assert "A3" in b1.refs

    def test_cross_sheet_refs(self, cross_sheet_xlsx: Path) -> None:
        wb = extract_workbook(cross_sheet_xlsx)
        assert {s.name for s in wb.sheets} == {"Input", "Calc"}

        calc = next(s for s in wb.sheets if s.name == "Calc")

        b1 = next(f for f in calc.formulas if f.coord == "Calc!B1")
        assert "Input!A1" in b1.refs

        b2 = next(f for f in calc.formulas if f.coord == "Calc!B2")
        # SUMIF の3つの引数すべてが拾えていること
        assert "Input!A:A" in b2.refs
        assert "B1" in b2.refs


class TestNamedRanges:
    def test_named_range_attached_to_correct_sheet(self, named_range_xlsx: Path) -> None:
        wb = extract_workbook(named_range_xlsx)
        data_sheet = next(s for s in wb.sheets if s.name == "Data")
        assert len(data_sheet.named_ranges) == 1
        nr = data_sheet.named_ranges[0]
        assert nr.name == "TaxRate"
        assert "Data" in nr.refers_to


class TestConditionalFormats:
    def test_extracted(self, conditional_format_xlsx: Path) -> None:
        wb = extract_workbook(conditional_format_xlsx)
        sheet = wb.sheets[0]
        assert len(sheet.conditional_formats) >= 1
        cf = sheet.conditional_formats[0]
        assert "A1:A10" in cf.range


class TestXlsHandling:
    """`.xls` (旧バイナリ) が来たら空シートで返すこと."""

    def test_xls_returns_empty_sheets(self, tmp_path: Path) -> None:
        # 中身のないファイルでも .xls 拡張子なら早期return される
        fake_xls = tmp_path / "legacy.xls"
        fake_xls.write_bytes(b"not a real xls")
        wb = extract_workbook(fake_xls)
        assert wb.filename == "legacy.xls"
        assert wb.sheets == []
        assert wb.vba_modules == []


class TestErrors:
    def test_missing_file(self, tmp_path: Path) -> None:
        with pytest.raises(ExtractionError):
            extract_workbook(tmp_path / "no_such_file.xlsx")

    def test_corrupt_xlsx(self, tmp_path: Path) -> None:
        bad = tmp_path / "corrupt.xlsx"
        bad.write_bytes(b"definitely not xlsx")
        with pytest.raises(ExtractionError):
            extract_workbook(bad)


class TestEmptyWorkbook:
    def test_no_formulas_no_named_ranges(self, empty_xlsx: Path) -> None:
        wb = extract_workbook(empty_xlsx)
        assert len(wb.sheets) == 1
        sheet = wb.sheets[0]
        assert sheet.formulas == []
        assert sheet.named_ranges == []
        assert sheet.conditional_formats == []


# ---------- 新規: Excel テーブル / マージ / プレビュー ----------


@pytest.fixture
def table_xlsx(tmp_path: Path) -> Path:
    """Excel テーブル (ListObject) を含む xlsx."""
    from openpyxl.worksheet.table import Table

    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Portfolio"
    ws["A1"] = "銘柄コード"
    ws["B1"] = "銘柄名"
    ws["A2"] = "ABC"
    ws["B2"] = "株式会社A"
    ws["A3"] = "DEF"
    ws["B3"] = "株式会社D"
    tbl = Table(displayName="PortfolioTable", ref="A1:B3")
    ws.add_table(tbl)

    out = tmp_path / "tbl.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def merged_xlsx(tmp_path: Path) -> Path:
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    ws["A1"] = "タイトル"
    ws.merge_cells("A1:C1")
    out = tmp_path / "merged.xlsx"
    wb.save(out)
    return out


class TestExcelTables:
    def test_table_extracted(self, table_xlsx: Path) -> None:
        wb = extract_workbook(table_xlsx)
        sheet = wb.sheets[0]
        assert len(sheet.tables) == 1
        t = sheet.tables[0]
        assert t.name == "PortfolioTable"
        assert t.ref == "A1:B3"
        assert t.header_row_count == 1

    def test_no_table_returns_empty(self, simple_xlsx: Path) -> None:
        wb = extract_workbook(simple_xlsx)
        assert wb.sheets[0].tables == []


class TestMergedRanges:
    def test_extracted(self, merged_xlsx: Path) -> None:
        wb = extract_workbook(merged_xlsx)
        sheet = wb.sheets[0]
        assert "A1:C1" in sheet.merged_ranges


class TestPreview:
    def test_preview_includes_first_rows(self, table_xlsx: Path) -> None:
        wb = extract_workbook(table_xlsx)
        sheet = wb.sheets[0]
        assert sheet.preview_rows
        first_row = sheet.preview_rows[0]
        assert "銘柄コード" in first_row
        assert "銘柄名" in first_row

    def test_preview_origin_recorded(self, table_xlsx: Path) -> None:
        wb = extract_workbook(table_xlsx)
        assert wb.sheets[0].preview_origin.startswith("A1:")

    def test_preview_truncated_at_limits(self, tmp_path: Path) -> None:
        # 30行 × 30列の xlsx で preview が 20×20 に絞られること
        wb = OpyWorkbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Big"
        for r in range(1, 31):
            for c in range(1, 31):
                ws.cell(row=r, column=c, value=f"r{r}c{c}")
        out = tmp_path / "big.xlsx"
        wb.save(out)

        from core.extractors.workbook import PREVIEW_MAX_COLS, PREVIEW_MAX_ROWS

        wb_extracted = extract_workbook(out)
        preview = wb_extracted.sheets[0].preview_rows
        assert len(preview) == PREVIEW_MAX_ROWS
        assert all(len(row) == PREVIEW_MAX_COLS for row in preview)

    def test_empty_workbook_has_preview_at_least_one_cell(self, empty_xlsx: Path) -> None:
        wb = extract_workbook(empty_xlsx)
        sheet = wb.sheets[0]
        # empty_xlsx の A1 に "hello" が入っているので最低限1セル
        assert sheet.preview_rows
        assert sheet.preview_rows[0][0] == "hello"


# ---------- データ検証 (入力規則) ----------


class TestDataValidations:
    def test_extracts_list_validation(self, tmp_path: Path) -> None:
        """セルにドロップダウンリストが設定されていれば抽出される."""
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = OpyWorkbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Form"
        dv = DataValidation(
            type="list",
            formula1='"Apple,Banana,Cherry"',
            allow_blank=True,
            prompt="果物を選択",
        )
        dv.add("A2:A10")
        ws.add_data_validation(dv)
        out = tmp_path / "dv_list.xlsx"
        wb.save(out)

        result = extract_workbook(out)
        sheet = next(s for s in result.sheets if s.name == "Form")
        assert len(sheet.data_validations) == 1
        ext = sheet.data_validations[0]
        assert ext.type == "list"
        assert "Apple" in ext.formula
        assert ext.range == "A2:A10"
        assert ext.prompt == "果物を選択"

    def test_extracts_numeric_validation_with_operator(self, tmp_path: Path) -> None:
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = OpyWorkbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Nums"
        dv = DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="100",
            allow_blank=False,
        )
        dv.add("B2:B5")
        ws.add_data_validation(dv)
        out = tmp_path / "dv_num.xlsx"
        wb.save(out)

        result = extract_workbook(out)
        sheet = next(s for s in result.sheets if s.name == "Nums")
        assert any(
            d.type == "whole" and d.operator == "between" and d.range == "B2:B5"
            for d in sheet.data_validations
        )

    def test_no_validations_when_none(self, tmp_path: Path) -> None:
        wb = OpyWorkbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Empty"
        out = tmp_path / "no_dv.xlsx"
        wb.save(out)
        result = extract_workbook(out)
        assert result.sheets[0].data_validations == []


# ---------- フォームコントロール (VML 抽出) ----------


class TestFormControlsParser:
    def test_parses_button_with_macro_from_vml(self) -> None:
        """合成 VML から FmlaMacro / ObjectType が拾えること."""
        from core.extractors.workbook import _parse_vml_form_controls

        vml_str = """<?xml version="1.0" encoding="UTF-8"?>
<xml xmlns:v="urn:schemas-microsoft-com:vml"
     xmlns:x="urn:schemas-microsoft-com:office:excel"
     xmlns:o="urn:schemas-microsoft-com:office:office">
  <v:shape id="_x0000_s1025" o:spid="_x0000_s1025" type="#_x0000_t201">
    <v:textbox>
      <div><font>更新</font></div>
    </v:textbox>
    <x:ClientData ObjectType="Button">
      <x:Anchor>2,12,3,12,5,0,5,0</x:Anchor>
      <x:FmlaMacro>Module1.UpdateDaily</x:FmlaMacro>
    </x:ClientData>
  </v:shape>
  <v:shape id="_x0000_s1026" o:spid="_x0000_s1026">
    <x:ClientData ObjectType="Checkbox">
      <x:FmlaMacro>Module1.ToggleFilter</x:FmlaMacro>
    </x:ClientData>
  </v:shape>
</xml>
"""
        vml = vml_str.encode("utf-8")
        controls = _parse_vml_form_controls(vml)
        assert len(controls) == 2
        btn = controls[0]
        assert btn.kind == "button"
        assert btn.macro == "Module1.UpdateDaily"
        assert "更新" in btn.text
        # アンカー: FromCol=2, FromRow=3 → C4
        assert btn.anchor == "C4"

        cb = controls[1]
        assert cb.kind == "checkbox"
        assert cb.macro == "Module1.ToggleFilter"

    def test_skips_shapes_without_macro_or_text(self) -> None:
        from core.extractors.workbook import _parse_vml_form_controls

        vml = b"""<?xml version="1.0" encoding="UTF-8"?>
<xml xmlns:v="urn:schemas-microsoft-com:vml"
     xmlns:x="urn:schemas-microsoft-com:office:excel">
  <v:shape>
    <x:ClientData ObjectType="Button"></x:ClientData>
  </v:shape>
</xml>
"""
        controls = _parse_vml_form_controls(vml)
        # macro も text も無いので空
        assert controls == []

    def test_malformed_vml_returns_empty(self) -> None:
        from core.extractors.workbook import _parse_vml_form_controls

        assert _parse_vml_form_controls(b"<not xml") == []

    def test_xlsm_without_buttons_returns_empty_map(self, tmp_path: Path) -> None:
        """ボタン無しの xlsm でも例外を出さず空マップを返す."""
        from core.extractors.workbook import _extract_form_controls

        wb = OpyWorkbook()
        ws = wb.active
        assert ws is not None
        ws.title = "S"
        out = tmp_path / "no_buttons.xlsm"
        wb.save(out)
        result = _extract_form_controls(out, ["S"])
        # キーは作られるが空配列
        assert result == {"S": []}

    def test_non_xlsm_extension_returns_empty(self, tmp_path: Path) -> None:
        """.xlsx (マクロなし) では VML 解析自体をスキップ."""
        from core.extractors.workbook import _extract_form_controls

        wb = OpyWorkbook()
        ws = wb.active
        assert ws is not None
        ws.title = "S"
        out = tmp_path / "plain.xlsx"
        wb.save(out)
        result = _extract_form_controls(out, ["S"])
        assert result == {"S": []}

    def test_extracts_controls_from_all_vml_relationships(self, tmp_path: Path) -> None:
        """1シートに複数VMLがある場合、後続VMLのボタンも抽出する."""
        from core.extractors.workbook import _extract_form_controls

        workbook_xml = """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Problem" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
"""
        workbook_rels = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
                Target="worksheets/sheet1.xml"/>
</Relationships>
"""
        sheet_rels = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1"
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"
                Target="../drawings/vmlDrawing1.vml"/>
  <Relationship Id="rId2"
                Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"
                Target="/xl/drawings/vmlDrawing2.vml"/>
</Relationships>
"""
        empty_vml = """<?xml version="1.0" encoding="UTF-8"?>
<xml xmlns:v="urn:schemas-microsoft-com:vml"
     xmlns:x="urn:schemas-microsoft-com:office:excel">
  <v:shape>
    <x:ClientData ObjectType="Button"></x:ClientData>
  </v:shape>
</xml>
"""
        button_vml = """<?xml version="1.0" encoding="UTF-8"?>
<xml xmlns:v="urn:schemas-microsoft-com:vml"
     xmlns:x="urn:schemas-microsoft-com:office:excel"
     xmlns:o="urn:schemas-microsoft-com:office:office">
  <v:shape id="_x0000_s2048" o:spid="_x0000_s2048">
    <v:textbox><div>実行</div></v:textbox>
    <x:ClientData ObjectType="Button">
      <x:Anchor>1,0,2,0,3,0,4,0</x:Anchor>
      <x:FmlaMacro>Module1.Run</x:FmlaMacro>
    </x:ClientData>
  </v:shape>
</xml>
"""
        xlsm = tmp_path / "multi_vml.xlsm"
        with zipfile.ZipFile(xlsm, "w") as zf:
            zf.writestr("xl/workbook.xml", workbook_xml)
            zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            zf.writestr("xl/worksheets/_rels/sheet1.xml.rels", sheet_rels)
            zf.writestr("xl/drawings/vmlDrawing1.vml", empty_vml)
            zf.writestr("xl/drawings/vmlDrawing2.vml", button_vml)

        result = _extract_form_controls(xlsm, ["Problem"])

        assert len(result["Problem"]) == 1
        assert result["Problem"][0].kind == "button"
        assert result["Problem"][0].text == "実行"
        assert result["Problem"][0].macro == "Module1.Run"
