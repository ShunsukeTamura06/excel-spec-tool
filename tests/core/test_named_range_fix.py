"""core.named_range_fix のテスト."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from core.exceptions import NamedRangeFixError
from core.models import NamedRange, Reference, ReferenceIndex, SheetInfo, Workbook
from core.named_range_fix import apply_named_range_fix, propose_named_range_fix


def _wb(named_ranges: list[NamedRange]) -> Workbook:
    return Workbook(
        filename="x.xlsx",
        sheets=[SheetInfo(name="Sheet1", rows=1, cols=1, named_ranges=named_ranges)],
    )


class TestProposeNamedRangeFix:
    def test_found_name_returns_diff_with_one_named_range_change(self) -> None:
        before = _wb([NamedRange(name="TaxRate", refers_to="Data!$A$1")])
        diff = propose_named_range_fix(before, ReferenceIndex(), "TaxRate", "Data!$B$1")

        assert len(diff.named_ranges) == 1
        nr = diff.named_ranges[0]
        assert nr.name == "TaxRate"
        assert nr.change_type == "modified"
        assert nr.before_refers_to == "Data!$A$1"
        assert nr.after_refers_to == "Data!$B$1"
        assert diff.cells == []
        assert diff.conditional_formats == []

    def test_unknown_name_raises(self) -> None:
        before = _wb([NamedRange(name="TaxRate", refers_to="Data!$A$1")])
        with pytest.raises(NamedRangeFixError):
            propose_named_range_fix(before, ReferenceIndex(), "NoSuchName", "Data!$B$1")

    def test_blast_radius_populated_when_referenced(self) -> None:
        before = _wb([NamedRange(name="TaxRate", refers_to="Data!$A$1")])
        idx = ReferenceIndex(
            refs={"Data!A1": [Reference(kind="formula", from_="Calc!C1", to="Data!A1")]}
        )
        diff = propose_named_range_fix(before, idx, "TaxRate", "Data!$B$1")

        assert len(diff.blast_radius) == 1
        assert diff.blast_radius[0].referenced_by[0].from_ == "Calc!C1"

    def test_does_not_mutate_input_before_wb(self) -> None:
        before = _wb([NamedRange(name="TaxRate", refers_to="Data!$A$1")])
        propose_named_range_fix(before, ReferenceIndex(), "TaxRate", "Data!$B$1")

        assert before.sheets[0].named_ranges[0].refers_to == "Data!$A$1"


class TestApplyNamedRangeFix:
    def _make_xlsx(self, path: Path, name: str = "TaxRate", refers_to: str = "Data!$A$1") -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws["A1"] = 0.1
        wb.defined_names[name] = DefinedName(name=name, attr_text=refers_to)
        wb.save(path)

    def test_writes_new_refers_to(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        self._make_xlsx(before)

        apply_named_range_fix(before, "TaxRate", "Data!$B$1", out)

        wb = openpyxl.load_workbook(out)
        assert wb.defined_names["TaxRate"].value == "Data!$B$1"

    def test_unknown_name_raises(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        self._make_xlsx(before)

        with pytest.raises(NamedRangeFixError):
            apply_named_range_fix(before, "NoSuchName", "Data!$B$1", out)

    def test_out_path_independent_of_in_path(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        self._make_xlsx(before)

        apply_named_range_fix(before, "TaxRate", "Data!$B$1", out)

        before_wb = openpyxl.load_workbook(before)
        assert before_wb.defined_names["TaxRate"].value == "Data!$A$1"
