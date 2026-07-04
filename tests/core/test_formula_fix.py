"""core.formula_fix のテスト."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from core.exceptions import FormulaFixError
from core.formula_fix import (
    apply_fixed_ref_replace,
    apply_range_expansion,
    propose_fixed_ref_replace,
    propose_range_expansion,
)
from core.models import CellFormula, Reference, ReferenceIndex, SheetInfo, Workbook


def _wb(formulas_by_sheet: dict[str, list[CellFormula]]) -> Workbook:
    return Workbook(
        filename="x.xlsx",
        sheets=[
            SheetInfo(name=name, rows=10, cols=10, formulas=formulas)
            for name, formulas in formulas_by_sheet.items()
        ],
    )


class TestProposeFixedRefReplace:
    def test_replaces_qualified_and_unqualified_tokens(self) -> None:
        before = _wb(
            {
                "Calc": [
                    CellFormula(coord="Calc!C1", formula="=Data!$B$5*2"),
                ],
                "Data": [
                    CellFormula(coord="Data!C2", formula="=SUM($B$5,B6)"),
                ],
            }
        )
        diff = propose_fixed_ref_replace(before, ReferenceIndex(), "Data!$B$5", "Data!$B$6")

        assert len(diff.cells) == 2
        by_sheet = {c.sheet: c for c in diff.cells}
        assert by_sheet["Calc"].after_formula == "=Data!$B$6*2"
        # 元がシート修飾なし・同一シートへの置換なら修飾なし表記を保つ
        assert by_sheet["Data"].after_formula == "=SUM($B$6,B6)"
        assert by_sheet["Data"].coord == "C2"
        assert all(c.change_type == "modified" for c in diff.cells)

    def test_does_not_touch_string_literals_or_partial_matches(self) -> None:
        before = _wb(
            {
                "Calc": [
                    CellFormula(coord="Calc!C1", formula='=IF(A10>0,"A1",A1)'),
                ],
            }
        )
        diff = propose_fixed_ref_replace(before, ReferenceIndex(), "Calc!A1", "Calc!B1")

        assert len(diff.cells) == 1
        # 文字列リテラル "A1" と部分一致 A10 は置換されない
        assert diff.cells[0].after_formula == '=IF(A10>0,"A1",B1)'

    def test_dollar_signs_ignored_for_matching(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=A1+$A$1")]})
        diff = propose_fixed_ref_replace(before, ReferenceIndex(), "Calc!$A$1", "Calc!$A$2")

        assert diff.cells[0].after_formula == "=$A$2+$A$2"

    def test_cross_sheet_replacement_keeps_qualifier(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=B5")]})
        diff = propose_fixed_ref_replace(before, ReferenceIndex(), "Calc!B5", "Data!B5")

        assert diff.cells[0].after_formula == "=Data!B5"

    def test_sheet_name_with_space_is_quoted(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=B5")]})
        diff = propose_fixed_ref_replace(before, ReferenceIndex(), "Calc!B5", "My Data!B5")

        assert diff.cells[0].after_formula == "='My Data'!B5"

    def test_no_matching_formula_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=A1")]})
        with pytest.raises(FormulaFixError):
            propose_fixed_ref_replace(before, ReferenceIndex(), "Calc!Z99", "Calc!Z100")

    def test_unqualified_old_ref_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=A1")]})
        with pytest.raises(FormulaFixError):
            propose_fixed_ref_replace(before, ReferenceIndex(), "A1", "B1")

    def test_blast_radius_populated_when_changed_cell_is_referenced(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=A1")]})
        idx = ReferenceIndex(
            refs={"Calc!C1": [Reference(kind="formula", from_="Report!D1", to="Calc!C1")]}
        )
        diff = propose_fixed_ref_replace(before, idx, "Calc!A1", "Calc!B1")

        assert len(diff.blast_radius) == 1
        assert diff.blast_radius[0].referenced_by[0].from_ == "Report!D1"

    def test_does_not_mutate_input_before_wb(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=A1")]})
        propose_fixed_ref_replace(before, ReferenceIndex(), "Calc!A1", "Calc!B1")

        assert before.sheets[0].formulas[0].formula == "=A1"


class TestProposeRangeExpansion:
    def test_expands_range_in_formulas(self) -> None:
        before = _wb(
            {
                "Calc": [
                    CellFormula(coord="Calc!C1", formula="=SUM(Data!$A$1:$A$100)"),
                    CellFormula(coord="Calc!C2", formula="=AVERAGE(Data!A1:A100)"),
                ],
            }
        )
        diff = propose_range_expansion(
            before, ReferenceIndex(), "Data!$A$1:$A$100", "Data!$A$1:$A$200"
        )

        assert len(diff.cells) == 2
        assert diff.cells[0].after_formula == "=SUM(Data!$A$1:$A$200)"
        assert diff.cells[1].after_formula == "=AVERAGE(Data!$A$1:$A$200)"

    def test_shrinking_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=SUM(Data!A1:A100)")]})
        with pytest.raises(FormulaFixError):
            propose_range_expansion(before, ReferenceIndex(), "Data!A1:A100", "Data!A1:A50")

    def test_disjoint_range_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=SUM(Data!A1:A100)")]})
        with pytest.raises(FormulaFixError):
            propose_range_expansion(before, ReferenceIndex(), "Data!A1:A100", "Data!B1:B200")

    def test_identical_range_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=SUM(Data!A1:A100)")]})
        with pytest.raises(FormulaFixError):
            propose_range_expansion(before, ReferenceIndex(), "Data!A1:A100", "Data!$A$1:$A$100")

    def test_cross_sheet_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=SUM(Data!A1:A100)")]})
        with pytest.raises(FormulaFixError):
            propose_range_expansion(before, ReferenceIndex(), "Data!A1:A100", "Other!A1:A200")

    def test_whole_column_raises(self) -> None:
        before = _wb({"Calc": [CellFormula(coord="Calc!C1", formula="=SUM(Data!A:A)")]})
        with pytest.raises(FormulaFixError):
            propose_range_expansion(before, ReferenceIndex(), "Data!A:A", "Data!A:B")


class TestApplyFixedRefReplace:
    def _make_xlsx(self, path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws["B5"] = 10
        ws["B6"] = 20
        ws["C1"] = "=$B$5*2"
        ws["C2"] = '=IF(B50>0,"B5",B5)'
        wb.save(path)

    def test_replaces_only_matching_tokens(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        self._make_xlsx(before)

        replaced = apply_fixed_ref_replace(before, "Data!$B$5", "Data!$B$6", out)

        assert replaced == 2
        wb = openpyxl.load_workbook(out)
        ws = wb["Data"]
        assert ws["C1"].value == "=$B$6*2"
        # 文字列リテラル "B5" と部分一致 B50 は温存される
        assert ws["C2"].value == '=IF(B50>0,"B5",$B$6)'

    def test_no_match_raises(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        self._make_xlsx(before)

        with pytest.raises(FormulaFixError):
            apply_fixed_ref_replace(before, "Data!Z99", "Data!Z100", out)

    def test_source_file_unchanged(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        self._make_xlsx(before)

        apply_fixed_ref_replace(before, "Data!$B$5", "Data!$B$6", out)

        wb = openpyxl.load_workbook(before)
        assert wb["Data"]["C1"].value == "=$B$5*2"


class TestApplyRangeExpansion:
    def test_expands_range(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws["C1"] = "=SUM(A1:A100)"
        wb.save(before)

        replaced = apply_range_expansion(before, "Data!A1:A100", "Data!A1:A200", out)

        assert replaced == 1
        out_wb = openpyxl.load_workbook(out)
        assert out_wb["Data"]["C1"].value == "=SUM(A1:A200)"

    def test_invalid_expansion_raises_before_writing(self, tmp_path: Path) -> None:
        before = tmp_path / "before.xlsx"
        out = tmp_path / "out.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Data"
        ws["C1"] = "=SUM(A1:A100)"
        wb.save(before)

        with pytest.raises(FormulaFixError):
            apply_range_expansion(before, "Data!A1:A100", "Data!A1:A50", out)
        assert not out.exists()
