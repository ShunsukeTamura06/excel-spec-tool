"""変更計画とopenpyxlプロバイダー境界のテスト."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from core.exceptions import UnsupportedMutationError
from core.extractors.workbook import extract_workbook
from core.models import CellFormula, ReferenceIndex, SheetInfo, Workbook
from core.mutation import (
    CellTextBatchOperation,
    CellTextEdit,
    FixedRefReplaceOperation,
    MutationPlan,
    NamedRangeSetOperation,
    OpenPyxlMutationProvider,
    RangeExpansionOperation,
    build_safe_change_plan,
    propose_mutation,
)
from core.reference_index import build_reference_index


def _named_range_workbook(path: Path) -> None:
    """名前定義を持つxlsxを作成する."""

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    sheet["A1"] = 1
    workbook.defined_names["TaxRate"] = DefinedName(name="TaxRate", attr_text="Data!$A$1")
    workbook.save(path)
    workbook.close()


def test_propose_and_apply_share_one_plan(tmp_path: Path) -> None:
    """同じ変更計画を期待差分生成と実適用の両方に使える."""

    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _named_range_workbook(source)
    workbook = extract_workbook(source)
    index = build_reference_index(workbook)
    plan = MutationPlan(
        source_job_id="00000000-0000-4000-8000-000000000000",
        operation=NamedRangeSetOperation(
            name="TaxRate",
            new_refers_to="Data!$A$2",
        ),
    )

    expected = propose_mutation(plan, workbook, index)
    result = OpenPyxlMutationProvider().apply(plan, source, output)

    assert expected.named_ranges[0].after_refers_to == "Data!$A$2"
    assert result.provider == "openpyxl"
    assert result.changed_count == 1
    assert source.read_bytes() != output.read_bytes()
    changed = openpyxl.load_workbook(output)
    assert changed.defined_names["TaxRate"].attr_text == "Data!$A$2"
    changed.close()


def test_openpyxl_capability_is_explicit() -> None:
    """対応形式と操作をcapabilityとして機械取得できる."""

    capability = OpenPyxlMutationProvider().capability()

    assert capability.available
    assert capability.supported_extensions == [".xlsx", ".xlsm"]
    assert set(capability.supported_operations) == {
        "named_range_set",
        "fixed_ref_replace",
        "range_expansion",
    }


def test_cell_text_batch_proposes_added_cells() -> None:
    """空セルへの固定テキスト追加をプロバイダー非依存の期待差分にする."""

    workbook = Workbook(filename="tool.xlsx", sheets=[SheetInfo(name="Output", rows=2, cols=2)])
    plan = MutationPlan(
        source_job_id="00000000-0000-4000-8000-000000000000",
        requested_provider="officecli",
        operation=CellTextBatchOperation(
            edits=[
                CellTextEdit(sheet="Output", coord="C3", value="説明"),
                CellTextEdit(sheet="Output", coord="C4", value="商品数を示します"),
            ]
        ),
    )

    expected = propose_mutation(plan, workbook, ReferenceIndex())

    assert [(cell.coord, cell.after_value) for cell in expected.cells] == [
        ("C3", "説明"),
        ("C4", "商品数を示します"),
    ]
    safe_plan = build_safe_change_plan(plan, workbook, ReferenceIndex())
    assert safe_plan.title == "説明テキストを追加する"
    assert safe_plan.expected_change_count == 2


def test_openpyxl_rejects_unsupported_extension(tmp_path: Path) -> None:
    """対応外形式を曖昧に処理せず明示的に拒否する."""

    source = tmp_path / "source.xls"
    source.write_bytes(b"not-an-xls")
    plan = MutationPlan(
        source_job_id="00000000-0000-4000-8000-000000000000",
        operation=NamedRangeSetOperation(name="TaxRate", new_refers_to="Data!$A$2"),
    )

    with pytest.raises(UnsupportedMutationError, match="does not support"):
        OpenPyxlMutationProvider().apply(plan, source, tmp_path / "output.xls")


def test_safe_change_plan_describes_range_expansion_without_writing() -> None:
    """範囲拡張を期待差分・確認事項付きの一般ユーザー向け計画にする."""
    workbook = Workbook(
        filename="tool.xlsx",
        sheets=[
            SheetInfo(
                name="集計",
                rows=1,
                cols=1,
                formulas=[
                    CellFormula(
                        coord="集計!A1",
                        formula="=SUM(入力!A1:A10)",
                        refs=["入力!A1:A10"],
                    )
                ],
            ),
            SheetInfo(name="入力", rows=10, cols=1),
        ],
    )
    plan = MutationPlan(
        source_job_id="00000000-0000-4000-8000-000000000000",
        operation=RangeExpansionOperation(
            old_ref="入力!A1:A10",
            new_ref="入力!A1:A20",
        ),
    )

    safe_plan = build_safe_change_plan(plan, workbook, ReferenceIndex())

    assert safe_plan.automation == "supported"
    assert safe_plan.can_apply
    assert safe_plan.expected_change_count == 1
    assert safe_plan.affected_locations == ["集計!A1"]
    assert safe_plan.expected_diff.cells[0].after_formula == "=SUM('入力'!A1:A20)"
    assert "再計算値" in safe_plan.verification_scope


def test_safe_change_plan_requires_review_when_existing_risks_exist() -> None:
    """既存リスクを持つ変更は自動対応可能でも追加確認扱いにする."""
    from core.models import AnalysisRisk

    workbook = Workbook(
        filename="tool.xlsx",
        sheets=[
            SheetInfo(
                name="集計",
                rows=1,
                cols=1,
                formulas=[
                    CellFormula(
                        coord="集計!A1",
                        formula="=入力!A1",
                        refs=["入力!A1"],
                    )
                ],
            ),
            SheetInfo(name="入力", rows=1, cols=2),
        ],
        analysis_risks=[
            AnalysisRisk(
                category="runtime_state",
                severity="medium",
                location="Module1",
                evidence="ActiveSheet",
                description="実行時の選択シートに依存します。",
                recommendation="Excelで確認してください。",
            )
        ],
    )
    plan = MutationPlan(
        source_job_id="00000000-0000-4000-8000-000000000000",
        operation=FixedRefReplaceOperation(old_ref="入力!A1", new_ref="入力!B1"),
    )

    safe_plan = build_safe_change_plan(plan, workbook, ReferenceIndex())

    assert safe_plan.automation == "needs_review"
    assert safe_plan.warnings == ["実行時の選択シートに依存します。"]
