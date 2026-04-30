"""backend テスト用 fixture.

`storage` を tmp_path に向け、`llm_client` をモックに固定した TestClient を提供する。
"""

from collections.abc import Iterator
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from backend.dependencies import get_llm_client, get_storage
from backend.llm_client import MockLLMClient
from backend.main import create_app
from backend.storage import Storage


@pytest.fixture
def backend_storage(tmp_path: Path) -> Storage:
    return Storage(tmp_path / "jobs")


@pytest.fixture
def backend_llm() -> MockLLMClient:
    return MockLLMClient()


@pytest.fixture
def client(
    backend_storage: Storage,
    backend_llm: MockLLMClient,
) -> Iterator[TestClient]:
    app = create_app()
    app.dependency_overrides[get_storage] = lambda: backend_storage
    app.dependency_overrides[get_llm_client] = lambda: backend_llm
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture
def sample_xlsx_bytes(tmp_path: Path) -> bytes:
    """テスト用の小さな xlsx をバイト列で生成."""
    from openpyxl import Workbook as OpyWorkbook

    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Calc"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    out = tmp_path / "sample.xlsx"
    wb.save(out)
    return out.read_bytes()
