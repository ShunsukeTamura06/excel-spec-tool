"""Event loop ブロッキング回帰テスト.

重い同期処理 (extract / analyze / chat) を `asyncio.to_thread` で
threadpool に逃がしているため、軽量リクエストが並行して進めることを
検証する.

複数人で使う際に「誰かが分析中だと他の人の画面が固まる」問題の再発防止.
"""

from __future__ import annotations

import asyncio
import time
from pathlib import Path
from typing import Any

import httpx
import pytest
from openpyxl import Workbook as OpyWorkbook

from backend.dependencies import get_llm_client, get_storage
from backend.llm_client import LLMResponse, ModelTier, Usage
from backend.main import create_app
from backend.storage import Storage
from core.models import CellFormula, ReferenceIndex, SheetInfo, Workbook


def _make_xlsx(path: Path) -> bytes:
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Calc"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)
    return path.read_bytes()


class _SlowLLM:
    """`time.sleep` で重い同期処理をシミュレートする LLM クライアント.

    これが event loop で直接呼ばれていると他のリクエストもブロックされる.
    `asyncio.to_thread` 経由なら他リクエストは即応する.
    """

    def __init__(self, delay_sec: float = 0.5) -> None:
        self.delay_sec = delay_sec
        self.pro_model = "slow-pro"
        self.fast_model = "slow-fast"

    def chat_completion(
        self,
        messages: list[dict[str, Any]],
        model: str | None = None,
        tier: ModelTier = "pro",
        cache_prefix: int = 0,
    ) -> str:
        time.sleep(self.delay_sec)
        return "[slow] done"

    def annotate_text(self, prompt: str, content: str, tier: ModelTier = "fast") -> str:
        time.sleep(self.delay_sec)
        return "[slow] annotation"

    def chat_completion_with_tools(
        self,
        messages: list[dict[str, Any]],
        tools: list[dict[str, Any]],
        model: str | None = None,
        tier: ModelTier = "pro",
        cache_prefix: int = 0,
    ) -> LLMResponse:
        time.sleep(self.delay_sec)
        return LLMResponse(
            content="[slow] reply",
            tool_calls=[],
            usage=Usage(prompt_tokens=1, completion_tokens=1, total_tokens=2),
            model="slow-pro",
        )


@pytest.fixture
def slow_llm() -> _SlowLLM:
    return _SlowLLM(delay_sec=0.5)


@pytest.fixture
def app_with_slow_llm(tmp_path: Path, slow_llm: _SlowLLM) -> Any:
    """`SlowLLM` と tmp_path Storage を注入した FastAPI アプリ."""
    storage = Storage(tmp_path / "jobs")
    app = create_app()
    app.dependency_overrides[get_storage] = lambda: storage
    app.dependency_overrides[get_llm_client] = lambda: slow_llm
    yield app, storage
    app.dependency_overrides.clear()


def _seed_job_with_workbook(storage: Storage) -> str:
    """analyze がすぐ走れる状態のジョブを作る (extracted 済)."""
    meta = storage.create_job("demo.xlsx", b"dummy")
    wb = Workbook(
        filename="demo.xlsx",
        sheets=[
            SheetInfo(
                name="Calc",
                rows=2,
                cols=1,
                formulas=[CellFormula(coord="A1", formula="=1+1", refs=[])],
            ),
        ],
    )
    storage.save_workbook(meta.job_id, wb)
    storage.save_references(meta.job_id, ReferenceIndex(refs={}))
    storage.update_status(meta.job_id, "extracted")
    return meta.job_id


@pytest.mark.asyncio
async def test_health_responds_while_analyze_is_slow(
    app_with_slow_llm: tuple[Any, Storage],
) -> None:
    """重い /analyze と並行に /health を叩いて、即応することを確認する.

    もし /analyze が event loop をブロックしていたら、/health は
    /analyze の完了 (= ~0.5 秒以上) まで返ってこない.
    `asyncio.to_thread` で逃がしているなら /health は即座に 200 を返す.
    """
    app, storage = app_with_slow_llm
    job_id = _seed_job_with_workbook(storage)

    async with httpx.AsyncClient(
        transport=httpx.ASGITransport(app=app),
        base_url="http://test",
    ) as client:
        # /analyze を投げて、その間に /health を測る
        analyze_task = asyncio.create_task(client.post(f"/analyze/{job_id}"))
        # analyze がスレッドプール側で sleep に入る瞬間を待つ
        await asyncio.sleep(0.05)

        t0 = time.perf_counter()
        health = await client.get("/health")
        elapsed = time.perf_counter() - t0

        analyze_resp = await analyze_task

    assert health.status_code == 200
    assert analyze_resp.status_code == 200
    # /health が /analyze の完了 (delay 0.5s) を待たされていないこと.
    # event loop が解放されていれば 0.1 秒未満で返るはず.
    # CI 環境のばらつきを考慮して 0.25 秒で線引き (delay 0.5s の半分).
    assert elapsed < 0.25, (
        f"/health was blocked while /analyze ran ({elapsed:.3f}s); "
        "event loop appears to be blocked by the LLM call"
    )


@pytest.mark.asyncio
async def test_jobs_responds_while_chat_is_slow(
    app_with_slow_llm: tuple[Any, Storage],
) -> None:
    """重い /chat 中に /jobs が即応することを確認する."""
    app, storage = app_with_slow_llm
    job_id = _seed_job_with_workbook(storage)

    async with httpx.AsyncClient(
        transport=httpx.ASGITransport(app=app),
        base_url="http://test",
    ) as client:
        chat_task = asyncio.create_task(
            client.post(f"/chat/{job_id}", json={"message": "hello"}),
        )
        await asyncio.sleep(0.05)

        t0 = time.perf_counter()
        jobs = await client.get("/jobs")
        elapsed = time.perf_counter() - t0

        chat_resp = await chat_task

    assert jobs.status_code == 200
    assert chat_resp.status_code == 200
    assert elapsed < 0.25, (
        f"/jobs was blocked while /chat ran ({elapsed:.3f}s); "
        "event loop appears to be blocked by the chat tool loop"
    )


@pytest.mark.asyncio
async def test_two_analyzes_run_in_parallel(
    app_with_slow_llm: tuple[Any, Storage],
) -> None:
    """2 つの /analyze が並行進行し、合計時間が直列より明確に短いことを確認.

    直列なら 0.5s × 2 = 1.0s, 並行なら ~0.5s. CI ばらつき考慮で < 0.85s 線.
    """
    app, storage = app_with_slow_llm
    job_a = _seed_job_with_workbook(storage)
    job_b = _seed_job_with_workbook(storage)

    async with httpx.AsyncClient(
        transport=httpx.ASGITransport(app=app),
        base_url="http://test",
    ) as client:
        t0 = time.perf_counter()
        r_a, r_b = await asyncio.gather(
            client.post(f"/analyze/{job_a}"),
            client.post(f"/analyze/{job_b}"),
        )
        elapsed = time.perf_counter() - t0

    assert r_a.status_code == 200
    assert r_b.status_code == 200
    # 直列実行 (~1.0s) より明確に短ければ to_thread が効いている.
    assert elapsed < 0.85, (
        f"two /analyze calls took {elapsed:.3f}s; "
        "they appear to be serialized rather than running in parallel"
    )


@pytest.mark.asyncio
async def test_health_responds_while_extract_is_slow(
    tmp_path: Path,
) -> None:
    """extract は実際の xlsx 処理を含むので、本物の同期コードでも
    event loop が解放されることを確認する.

    SlowLLM は使わず、サンプル xlsx を実際に extract させる.
    """
    storage = Storage(tmp_path / "jobs")
    app = create_app()
    app.dependency_overrides[get_storage] = lambda: storage
    # extract は LLM を使わないので依存はそのまま (Mock がフォールバック)

    sample = tmp_path / "s.xlsx"
    data = _make_xlsx(sample)

    try:
        async with httpx.AsyncClient(
            transport=httpx.ASGITransport(app=app),
            base_url="http://test",
        ) as client:
            extract_task = asyncio.create_task(
                client.post(
                    "/extract",
                    files={"file": ("s.xlsx", data, "application/octet-stream")},
                ),
            )
            await asyncio.sleep(0.01)
            t0 = time.perf_counter()
            health = await client.get("/health")
            elapsed = time.perf_counter() - t0
            extract_resp = await extract_task
    finally:
        app.dependency_overrides.clear()

    assert health.status_code == 200
    assert extract_resp.status_code == 200
    # 実 xlsx の extract は 50〜200ms 程度. /health はそれより十分速く返るはず.
    assert elapsed < 0.1, (
        f"/health blocked during /extract ({elapsed:.3f}s); "
        "the extract pipeline is running on the event loop"
    )
