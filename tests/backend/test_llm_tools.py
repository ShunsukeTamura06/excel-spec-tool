"""backend.llm_tools のテスト."""

import json
from pathlib import Path

import pytest
from openpyxl import Workbook as OpyWorkbook

from backend.llm_tools import (
    TOOL_DEFINITIONS,
    build_tool_definitions,
    execute_tool_call,
)
from backend.storage import Storage
from core.extractors.cells import extract_cells_to_sqlite
from core.extractors.workbook import extract_workbook
from core.models import (
    AnalysisRisk,
    CellFormula,
    ChartObject,
    ChartSeries,
    NamedRange,
    PivotTableInfo,
    PowerQueryInfo,
    Reference,
    ReferenceIndex,
    SheetInfo,
    VbaModule,
    VbaProcedure,
    Workbook,
)


@pytest.fixture
def job_with_cells(tmp_path: Path) -> tuple[Storage, str]:
    """cells.db を埋めたジョブを用意する."""
    storage = Storage(tmp_path / "jobs")

    # 小さな xlsx を生成
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Portfolio"
    headers = ["銘柄コード", "銘柄名", "保有口数", "現在値", "評価損益"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="ABC")
    ws.cell(row=2, column=2, value="株式会社A")
    ws.cell(row=2, column=3, value=100)
    src = tmp_path / "p.xlsx"
    wb.save(src)
    data = src.read_bytes()

    meta = storage.create_job("p.xlsx", data)
    extract_cells_to_sqlite(
        storage.get_original_path(meta.job_id),
        storage.cells_db_path(meta.job_id),
    )
    extracted = extract_workbook(storage.get_original_path(meta.job_id))
    extracted.filename = "p.xlsx"
    storage.save_workbook(meta.job_id, extracted)
    storage.save_references(meta.job_id, ReferenceIndex())
    return storage, meta.job_id


# ---------- 定義 ----------


class TestToolDefinitions:
    def test_all_tools_defined(self) -> None:
        names = {t["function"]["name"] for t in TOOL_DEFINITIONS}
        assert names == {
            "get_cells_range",
            "find_cells",
            "lookup_references",
            "list_vba_modules",
            "get_vba_procedure",
            "list_sheet_formulas",
            "list_workbook_objects",
            "list_analysis_risks",
            "lookup_external_function",
            "list_external_functions_used",
            "propose_named_range_fix",
            "propose_fixed_ref_replace",
            "propose_range_expansion",
            "propose_cell_text_edits",
            "propose_vba_procedure_replace",
        }

    def test_build_returns_list(self) -> None:
        tools = build_tool_definitions()
        assert isinstance(tools, list)
        assert tools == TOOL_DEFINITIONS

    def test_each_tool_has_schema(self) -> None:
        for t in TOOL_DEFINITIONS:
            assert t["type"] == "function"
            fn = t["function"]
            assert "name" in fn
            assert "description" in fn
            assert "parameters" in fn
            assert fn["parameters"]["type"] == "object"


# ---------- execute_tool_call ----------


class TestExecuteGetCellsRange:
    def test_returns_grid(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result_str = execute_tool_call(
            storage, job_id, "get_cells_range", {"sheet": "Portfolio", "range": "A1:E1"}
        )
        result = json.loads(result_str)
        assert result["sheet"] == "Portfolio"
        assert result["rows"][0][0]["value"] == "銘柄コード"

    def test_missing_args_returns_error_json(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result = json.loads(execute_tool_call(storage, job_id, "get_cells_range", {}))
        assert "error" in result

    def test_invalid_range_returns_error_json(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result = json.loads(
            execute_tool_call(
                storage, job_id, "get_cells_range", {"sheet": "P", "range": "garbage"}
            )
        )
        assert "error" in result


class TestExecuteProposeCellTextEdits:
    def test_returns_officecli_safe_plan(self, job_with_cells: tuple[Storage, str]) -> None:
        """空セルへの説明追加をOfficeCLI向け未適用計画として返す."""

        storage, job_id = job_with_cells
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_cell_text_edits",
                {
                    "edits": [
                        {
                            "sheet": "Portfolio",
                            "coord": "F1",
                            "value": "評価損益の見方",
                        }
                    ]
                },
            )
        )

        safe_plan = result["safe_plan"]
        assert safe_plan["plan"]["requested_provider"] == "officecli"
        assert safe_plan["plan"]["operation"]["kind"] == "cell_text_batch"
        assert safe_plan["expected_diff"]["cells"][0]["coord"] == "F1"
        assert safe_plan["expected_diff"]["cells"][0]["after_value"] == "評価損益の見方"

    def test_rejects_overwriting_existing_cell(
        self,
        job_with_cells: tuple[Storage, str],
    ) -> None:
        """既存値の上書きは提案段階で拒否する."""

        storage, job_id = job_with_cells
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_cell_text_edits",
                {
                    "edits": [
                        {
                            "sheet": "Portfolio",
                            "coord": "A1",
                            "value": "上書き",
                        }
                    ]
                },
            )
        )

        assert "error" in result
        assert "not empty" in result["error"]


class TestExecuteFindCells:
    def test_basic_find(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result = json.loads(
            execute_tool_call(storage, job_id, "find_cells", {"query": "銘柄コード"})
        )
        assert result["count"] >= 1

    def test_no_query_returns_error(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result = json.loads(execute_tool_call(storage, job_id, "find_cells", {}))
        assert "error" in result

    def test_sheet_filter(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        # シート指定なしと指定の両方で動作
        with_sheet = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "find_cells",
                {"query": "ABC", "sheet": "Portfolio"},
            )
        )
        assert with_sheet["count"] >= 1


class TestExecuteLookupReferences:
    def test_basic_lookup(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        # 参照インデックスを差し込んでおく
        idx = ReferenceIndex(
            refs={"Calc!H2": [Reference(kind="formula", from_="Out!K3", to="Calc!H2")]}
        )
        storage.save_references(job_id, idx)

        result = json.loads(
            execute_tool_call(storage, job_id, "lookup_references", {"target": "Calc!H2"})
        )
        assert result["count"] == 1
        assert result["refs"][0]["from"] == "Out!K3"
        assert "analysis_scope" in result

    def test_unknown_target_returns_empty(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        # references.json が無くてもエラーにはせず empty を返したいが、
        # 現実装は ToolExecutionError -> error JSON。どちらでもテスト的に許容。
        # ここではまず references を保存してから空ターゲットを引く。
        storage.save_references(job_id, ReferenceIndex())
        result = json.loads(
            execute_tool_call(storage, job_id, "lookup_references", {"target": "Nowhere!A1"})
        )
        assert result["count"] == 0
        assert "動的" in result["analysis_scope"]

    def test_missing_target_returns_error(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result = json.loads(execute_tool_call(storage, job_id, "lookup_references", {}))
        assert "error" in result


class TestExecuteUnknownTool:
    def test_returns_error(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        result = json.loads(execute_tool_call(storage, job_id, "no_such", {}))
        assert "error" in result


# ---------- VBA / formula 系ツール ----------


_VBA_CODE = (
    "Sub UpdateDaily()\n"
    '    Worksheets("Calc").Range("H2") = 1\n'
    "End Sub\n"
    "\n"
    "Function CalcTotal(x As Double) As Double\n"
    "    CalcTotal = x * 2\n"
    "End Function\n"
)


@pytest.fixture
def job_with_workbook(tmp_path: Path) -> tuple[Storage, str]:
    """extracted.json (Workbook) を埋めたジョブを用意する."""
    storage = Storage(tmp_path / "jobs")
    meta = storage.create_job("dummy.xlsm", b"x")

    wb = Workbook(
        filename="dummy.xlsm",
        sheets=[
            SheetInfo(
                name="Calc",
                rows=10,
                cols=10,
                formulas=[
                    CellFormula(coord="Calc!A1", formula="=SUM(B1:B10)", refs=["B1:B10"]),
                    CellFormula(
                        coord="Calc!A2",
                        formula="=SUMIF(Input!A:A, B2, Input!E:E)",
                        refs=["Input!A:A", "B2", "Input!E:E"],
                    ),
                    CellFormula(coord="Calc!A3", formula="=B1+B2", refs=["B1", "B2"]),
                ],
                charts=[
                    ChartObject(
                        name="Chart 1",
                        chart_type="barChart",
                        title="売上推移",
                        anchor="E2",
                        series=[
                            ChartSeries(
                                name="売上",
                                values_ref="Input!B2:B5",
                                categories_ref="Input!A2:A5",
                            )
                        ],
                    )
                ],
            ),
            SheetInfo(
                name="Input",
                rows=5,
                cols=5,
                pivot_tables=[
                    PivotTableInfo(
                        name="PivotTable1",
                        anchor="A10:C20",
                        source_sheet="Input",
                        source_ref="A1:E100",
                        value_fields=["合計 / 金額"],
                    )
                ],
            ),
        ],
        vba_modules=[
            VbaModule(
                name="Module1",
                type="Module",
                code=_VBA_CODE,
                procedures=[
                    VbaProcedure(
                        name="UpdateDaily",
                        kind="Sub",
                        start_line=1,
                        end_line=3,
                        code="",
                    ),
                    VbaProcedure(
                        name="CalcTotal",
                        kind="Function",
                        start_line=5,
                        end_line=7,
                        code="",
                    ),
                ],
            )
        ],
        power_queries=[
            PowerQueryInfo(
                name="Query - Sales",
                kind="power_query",
                connection_id="3",
                target_sheet="Input",
                target_name="SalesTable",
                source="Provider=Microsoft.Mashup.OleDb.1",
            )
        ],
        analysis_risks=[
            AnalysisRisk(
                category="dynamic_vba",
                severity="high",
                location="Module1:L2",
                evidence="Range(addr)",
                description="動的な Range 参照です。",
                recommendation="変数 addr の値を確認してください。",
            ),
            AnalysisRisk(
                category="external_dependency",
                severity="medium",
                location="Connection:Query - Sales",
                evidence="Provider=Microsoft.Mashup.OleDb.1",
                description="Power Query 接続があります。",
                recommendation="更新結果を確認してください。",
            ),
        ],
    )
    storage.save_workbook(meta.job_id, wb)
    storage.save_references(meta.job_id, ReferenceIndex())
    return storage, meta.job_id


class TestExecuteListVbaModules:
    def test_returns_modules(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(execute_tool_call(storage, job_id, "list_vba_modules", {}))
        assert result["count"] == 1
        m = result["modules"][0]
        assert m["name"] == "Module1"
        assert m["type"] == "Module"
        names = {p["name"] for p in m["procedures"]}
        assert names == {"UpdateDaily", "CalcTotal"}
        # コード本体は含まない
        assert "code" not in m
        assert "code" not in m["procedures"][0]

    def test_empty_when_no_vba(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        meta = storage.create_job("empty.xlsm", b"x")
        storage.save_workbook(meta.job_id, Workbook(filename="empty.xlsm"))
        result = json.loads(execute_tool_call(storage, meta.job_id, "list_vba_modules", {}))
        assert result == {"modules": [], "count": 0}

    def test_missing_workbook_returns_error(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        meta = storage.create_job("nx.xlsm", b"x")
        # extracted.json を保存していない状態
        result = json.loads(execute_tool_call(storage, meta.job_id, "list_vba_modules", {}))
        assert "error" in result


class TestExecuteGetVbaProcedure:
    def test_returns_code_from_module_lines(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "get_vba_procedure",
                {"module": "Module1", "name": "CalcTotal"},
            )
        )
        assert result["module"] == "Module1"
        assert result["name"] == "CalcTotal"
        assert result["kind"] == "Function"
        # 行範囲 5-7 のコードが切り出されている
        assert "Function CalcTotal" in result["code"]
        assert "End Function" in result["code"]
        # 別プロシージャは含まれない
        assert "Sub UpdateDaily" not in result["code"]

    def test_unknown_module(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "get_vba_procedure", {"module": "Nope", "name": "X"})
        )
        assert "error" in result
        assert "module not found" in result["error"]

    def test_unknown_procedure(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "get_vba_procedure",
                {"module": "Module1", "name": "Missing"},
            )
        )
        assert "error" in result
        assert "procedure not found" in result["error"]

    def test_missing_args(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(execute_tool_call(storage, job_id, "get_vba_procedure", {}))
        assert "error" in result


class TestExecuteProposeVbaProcedureReplace:
    def test_returns_windows_package_plan(
        self,
        job_with_workbook: tuple[Storage, str],
    ) -> None:
        """既存Subの完全置換をwindows_vbide計画として返す."""

        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_vba_procedure_replace",
                {
                    "module_name": "Module1",
                    "procedure_name": "UpdateDaily",
                    "new_code": (
                        'Sub UpdateDaily()\n    Worksheets("Calc").Range("H2") = 2\nEnd Sub'
                    ),
                },
            )
        )

        safe_plan = result["safe_plan"]
        assert safe_plan["plan"]["requested_provider"] == "windows_vbide"
        assert safe_plan["plan"]["operation"]["kind"] == "vba_procedure_replace"
        assert safe_plan["automation"] == "needs_review"
        assert "expected_diff" not in safe_plan

    def test_rejects_partial_vba_code(
        self,
        job_with_workbook: tuple[Storage, str],
    ) -> None:
        """完全なプロシージャでないコード断片を拒否する."""

        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_vba_procedure_replace",
                {
                    "module_name": "Module1",
                    "procedure_name": "UpdateDaily",
                    "new_code": 'Worksheets("Calc").Range("H2") = 2',
                },
            )
        )

        assert "error" in result


class TestExecuteListSheetFormulas:
    def test_all_formulas(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_sheet_formulas", {"sheet": "Calc"})
        )
        assert result["sheet"] == "Calc"
        assert result["total"] == 3
        assert result["returned"] == 3
        assert result["truncated"] is False
        coords = [f["coord"] for f in result["formulas"]]
        assert coords == ["Calc!A1", "Calc!A2", "Calc!A3"]

    def test_pattern_filter_case_insensitive(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "list_sheet_formulas",
                {"sheet": "Calc", "pattern": "sumif"},
            )
        )
        assert result["total"] == 1
        assert result["formulas"][0]["coord"] == "Calc!A2"

    def test_limit_truncates(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_sheet_formulas", {"sheet": "Calc", "limit": 2})
        )
        assert result["total"] == 3
        assert result["returned"] == 2
        assert result["truncated"] is True

    def test_unknown_sheet(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_sheet_formulas", {"sheet": "NoSuch"})
        )
        assert "error" in result

    def test_missing_sheet_arg(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(execute_tool_call(storage, job_id, "list_sheet_formulas", {}))
        assert "error" in result


class TestExecuteListWorkbookObjects:
    def test_returns_all_workbook_objects(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(execute_tool_call(storage, job_id, "list_workbook_objects", {}))

        assert result["counts"] == {"charts": 1, "pivot_tables": 1, "power_queries": 1}
        assert result["charts"][0]["title"] == "売上推移"
        assert result["pivot_tables"][0]["name"] == "PivotTable1"
        assert result["power_queries"][0]["name"] == "Query - Sales"
        assert "M コード本文は未解析" in result["analysis_scope"]

    def test_sheet_filter_excludes_workbook_level_connections(
        self,
        job_with_workbook: tuple[Storage, str],
    ) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_workbook_objects", {"sheet": "Calc"})
        )

        assert result["counts"] == {"charts": 1, "pivot_tables": 0, "power_queries": 0}
        assert result["charts"][0]["sheet"] == "Calc"

    def test_kind_filter(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_workbook_objects", {"kind": "pivot"})
        )

        assert result["counts"]["charts"] == 0
        assert result["counts"]["pivot_tables"] == 1
        assert result["counts"]["power_queries"] == 0

    def test_invalid_kind_returns_error(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_workbook_objects", {"kind": "bad"})
        )
        assert "error" in result


class TestExecuteListAnalysisRisks:
    def test_returns_risks(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(execute_tool_call(storage, job_id, "list_analysis_risks", {}))

        assert result["counts"] == {"high": 1, "medium": 1, "low": 0}
        assert result["total"] == 2
        assert result["risks"][0]["category"] == "dynamic_vba"
        assert "手動確認" in result["analysis_scope"]

    def test_filters_by_severity(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_analysis_risks", {"severity": "high"})
        )

        assert result["total"] == 1
        assert result["risks"][0]["severity"] == "high"

    def test_invalid_severity_returns_error(self, job_with_workbook: tuple[Storage, str]) -> None:
        storage, job_id = job_with_workbook
        result = json.loads(
            execute_tool_call(storage, job_id, "list_analysis_risks", {"severity": "bad"})
        )
        assert "error" in result


# ---------- 結果サイズ上限 (TOOL_RESULT_MAX_CHARS) ----------


class TestResultTruncation:
    """巨大な結果がコンテキストを食い潰さないよう、切り詰めが効くことを確認."""

    def test_oversize_result_is_capped(
        self,
        job_with_cells: tuple[Storage, str],
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        import backend.llm_tools as lt

        storage, job_id = job_with_cells
        # 関数自体を差し替えて、本番下限 (1000) より小さい値も試せるようにする。
        # マーカー (~95 字) を載せるためある程度のサイズが必要なので 200 に設定
        monkeypatch.setattr(lt, "_tool_result_max_chars", lambda: 200)
        result_str = execute_tool_call(
            storage, job_id, "get_cells_range", {"sheet": "Portfolio", "range": "A1:E1"}
        )
        assert len(result_str) <= 200
        assert "TRUNCATED" in result_str

    def test_marker_omitted_when_limit_below_marker_size(
        self,
        job_with_cells: tuple[Storage, str],
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """limit がマーカーより小さい場合でも、limit は必ず守られる."""
        import backend.llm_tools as lt

        storage, job_id = job_with_cells
        monkeypatch.setattr(lt, "_tool_result_max_chars", lambda: 50)
        result_str = execute_tool_call(
            storage, job_id, "get_cells_range", {"sheet": "Portfolio", "range": "A1:E1"}
        )
        assert len(result_str) <= 50

    def test_under_limit_passes_through(self, job_with_cells: tuple[Storage, str]) -> None:
        storage, job_id = job_with_cells
        # デフォルト上限 (20000) で十分に収まる小さな結果
        result_str = execute_tool_call(
            storage, job_id, "get_cells_range", {"sheet": "Portfolio", "range": "A1:E1"}
        )
        assert "TRUNCATED" not in result_str

    def test_env_var_overrides_default(self, monkeypatch: pytest.MonkeyPatch) -> None:
        import backend.llm_tools as lt

        monkeypatch.setenv("TOOL_RESULT_MAX_CHARS", "5000")
        assert lt._tool_result_max_chars() == 5000

    def test_env_var_below_floor_is_clamped(self, monkeypatch: pytest.MonkeyPatch) -> None:
        import backend.llm_tools as lt

        # 1000 未満は本番安全値として 1000 にクランプ
        monkeypatch.setenv("TOOL_RESULT_MAX_CHARS", "10")
        assert lt._tool_result_max_chars() == 1000

    def test_env_var_invalid_falls_back_to_default(self, monkeypatch: pytest.MonkeyPatch) -> None:
        import backend.llm_tools as lt

        monkeypatch.setenv("TOOL_RESULT_MAX_CHARS", "not-a-number")
        assert lt._tool_result_max_chars() == 20_000


# ---------- 外部関数 (Bloomberg) tool ----------


class TestLookupExternalFunction:
    def test_known_function(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        # storage を渡すが arguments のみで判定する
        out = execute_tool_call(storage, "any", "lookup_external_function", {"name": "BDH"})
        payload = json.loads(out)
        assert payload["name"] == "BDH"
        assert payload["vendor"] == "Bloomberg"
        assert payload["signature"].startswith("=BDH(")
        assert any("PX_LAST" in ex for ex in payload["examples"])

    def test_case_insensitive(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        out = execute_tool_call(storage, "any", "lookup_external_function", {"name": "bdp"})
        payload = json.loads(out)
        assert payload["name"] == "BDP"

    def test_unknown_returns_error_with_known_list(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        out = execute_tool_call(storage, "any", "lookup_external_function", {"name": "UNKNOWN"})
        payload = json.loads(out)
        assert "error" in payload
        assert "BDH" in payload["known"]

    def test_missing_name_returns_error(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        out = execute_tool_call(storage, "any", "lookup_external_function", {})
        payload = json.loads(out)
        assert "error" in payload


class TestListExternalFunctionsUsed:
    def test_returns_used_functions_with_counts(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        # ジョブを作って Workbook を保存
        meta = storage.create_job("demo.xlsm", b"d")
        wb = Workbook(
            filename="demo.xlsm",
            sheets=[
                SheetInfo(
                    name="Port",
                    rows=10,
                    cols=5,
                    formulas=[
                        CellFormula(
                            coord="A2",
                            formula='=BDP("AAPL US Equity", "PX_LAST")',
                            external_functions=["BDP"],
                        ),
                        CellFormula(
                            coord="A3",
                            formula='=BDP("MSFT US Equity", "PX_LAST")',
                            external_functions=["BDP"],
                        ),
                        CellFormula(
                            coord="B2",
                            formula='=BDH("AAPL US Equity", "PX_LAST", "-1Y")',
                            external_functions=["BDH"],
                        ),
                    ],
                )
            ],
        )
        storage.save_workbook(meta.job_id, wb)

        out = execute_tool_call(storage, meta.job_id, "list_external_functions_used", {})
        payload = json.loads(out)
        # 種類数 2, 合計 3 件
        assert payload["total_kinds"] == 2
        assert payload["total_uses"] == 3
        # BDP が先 (使用回数多い)
        names = [it["name"] for it in payload["items"]]
        assert names == ["BDP", "BDH"]
        # 主要箇所が拾えている
        bdp_item = next(it for it in payload["items"] if it["name"] == "BDP")
        assert "Port!A2" in bdp_item["top_locations"]
        assert bdp_item["count"] == 2

    def test_empty_when_no_external_functions(self, tmp_path: Path) -> None:
        storage = Storage(tmp_path / "jobs")
        meta = storage.create_job("demo.xlsm", b"d")
        wb = Workbook(filename="demo.xlsm", sheets=[SheetInfo(name="S", rows=1, cols=1)])
        storage.save_workbook(meta.job_id, wb)

        out = execute_tool_call(storage, meta.job_id, "list_external_functions_used", {})
        payload = json.loads(out)
        assert payload["items"] == []
        assert payload["total_kinds"] == 0
        assert payload["total_uses"] == 0


def test_tool_definitions_include_external_function_tools() -> None:
    """LLM に渡す tools 配列に新ツールが含まれる."""
    names = {t["function"]["name"] for t in build_tool_definitions()}
    assert "lookup_external_function" in names
    assert "list_external_functions_used" in names


@pytest.fixture
def job_with_named_range(tmp_path: Path) -> tuple[Storage, str]:
    """名前付き範囲を持つジョブ (workbook + references 済み) を用意する."""
    storage = Storage(tmp_path / "jobs")
    meta = storage.create_job("dummy.xlsx", b"x")

    wb = Workbook(
        filename="dummy.xlsx",
        sheets=[
            SheetInfo(
                name="Data",
                rows=5,
                cols=5,
                named_ranges=[NamedRange(name="TaxRate", refers_to="Data!$A$1")],
            ),
        ],
    )
    storage.save_workbook(meta.job_id, wb)
    storage.save_references(
        meta.job_id,
        ReferenceIndex(
            refs={"Data!A1": [Reference(kind="formula", from_="Calc!B2", to="Data!A1")]}
        ),
    )
    return storage, meta.job_id


class TestExecuteProposeNamedRangeFix:
    def test_returns_proposal_with_blast_radius(
        self, job_with_named_range: tuple[Storage, str]
    ) -> None:
        storage, job_id = job_with_named_range
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_named_range_fix",
                {"name": "TaxRate", "new_refers_to": "Data!$B$1"},
            )
        )
        proposal = result["proposal"]
        assert len(proposal["named_ranges"]) == 1
        nr = proposal["named_ranges"][0]
        assert nr["name"] == "TaxRate"
        assert nr["change_type"] == "modified"
        assert nr["before_refers_to"] == "Data!$A$1"
        assert nr["after_refers_to"] == "Data!$B$1"
        assert len(proposal["blast_radius"]) == 1
        assert "note" in result

    def test_unknown_name_returns_error(self, job_with_named_range: tuple[Storage, str]) -> None:
        storage, job_id = job_with_named_range
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_named_range_fix",
                {"name": "NoSuchName", "new_refers_to": "Data!$B$1"},
            )
        )
        assert "error" in result

    def test_missing_args_returns_error(self, job_with_named_range: tuple[Storage, str]) -> None:
        storage, job_id = job_with_named_range
        result = json.loads(
            execute_tool_call(storage, job_id, "propose_named_range_fix", {"name": "TaxRate"})
        )
        assert "error" in result


@pytest.fixture
def job_with_formulas(tmp_path: Path) -> tuple[Storage, str]:
    """数式を持つジョブ (workbook + references 済み) を用意する."""
    storage = Storage(tmp_path / "jobs")
    meta = storage.create_job("dummy.xlsx", b"x")

    wb = Workbook(
        filename="dummy.xlsx",
        sheets=[
            SheetInfo(
                name="Calc",
                rows=5,
                cols=5,
                formulas=[
                    CellFormula(coord="Calc!C1", formula="=Data!$B$5*2"),
                    CellFormula(coord="Calc!C2", formula="=SUM(Data!$A$1:$A$100)"),
                ],
            ),
        ],
    )
    storage.save_workbook(meta.job_id, wb)
    storage.save_references(
        meta.job_id,
        ReferenceIndex(
            refs={"Calc!C1": [Reference(kind="formula", from_="Report!D1", to="Calc!C1")]}
        ),
    )
    return storage, meta.job_id


class TestExecuteProposeFixedRefReplace:
    def test_returns_proposal_with_changed_cells(
        self, job_with_formulas: tuple[Storage, str]
    ) -> None:
        storage, job_id = job_with_formulas
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_fixed_ref_replace",
                {"old_ref": "Data!$B$5", "new_ref": "Data!$B$6"},
            )
        )
        proposal = result["proposal"]
        assert len(proposal["cells"]) == 1
        cell = proposal["cells"][0]
        assert cell["sheet"] == "Calc"
        assert cell["coord"] == "C1"
        assert cell["before_formula"] == "=Data!$B$5*2"
        assert cell["after_formula"] == "=Data!$B$6*2"
        assert result["changed_formula_count"] == 1
        assert len(proposal["blast_radius"]) == 1
        assert "note" in result

    def test_no_match_returns_error(self, job_with_formulas: tuple[Storage, str]) -> None:
        storage, job_id = job_with_formulas
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_fixed_ref_replace",
                {"old_ref": "Data!Z99", "new_ref": "Data!Z100"},
            )
        )
        assert "error" in result

    def test_missing_args_returns_error(self, job_with_formulas: tuple[Storage, str]) -> None:
        storage, job_id = job_with_formulas
        result = json.loads(
            execute_tool_call(
                storage, job_id, "propose_fixed_ref_replace", {"old_ref": "Data!$B$5"}
            )
        )
        assert "error" in result


class TestExecuteProposeRangeExpansion:
    def test_returns_proposal(self, job_with_formulas: tuple[Storage, str]) -> None:
        storage, job_id = job_with_formulas
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_range_expansion",
                {"old_range": "Data!$A$1:$A$100", "new_range": "Data!$A$1:$A$200"},
            )
        )
        proposal = result["proposal"]
        assert len(proposal["cells"]) == 1
        assert proposal["cells"][0]["after_formula"] == "=SUM(Data!$A$1:$A$200)"

    def test_shrinking_returns_error(self, job_with_formulas: tuple[Storage, str]) -> None:
        storage, job_id = job_with_formulas
        result = json.loads(
            execute_tool_call(
                storage,
                job_id,
                "propose_range_expansion",
                {"old_range": "Data!$A$1:$A$100", "new_range": "Data!$A$1:$A$50"},
            )
        )
        assert "error" in result
