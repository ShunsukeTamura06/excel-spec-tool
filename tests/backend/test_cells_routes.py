"""backend.routes.cells のテスト + Storage の cells クエリ機能."""

import uuid
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook as OpyWorkbook

from backend.storage import Storage


@pytest.fixture
def cells_xlsx_bytes(tmp_path: Path) -> bytes:
    """Portfolio 風の小さな表入り xlsx をバイト列で返す."""
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Portfolio"
    headers = ["銘柄コード", "銘柄名", "保有口数", "現在値", "評価損益"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    rows = [
        ("ABC", "株式会社A", 100, 1500, 50000),
        ("DEF", "株式会社D", 200, 800, -20000),
        ("GHI", "株式会社G", 50, 2200, 80000),
    ]
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    out = tmp_path / "p.xlsx"
    wb.save(out)
    return out.read_bytes()


# ---------- Storage の cells クエリ機能 ----------


class TestStorageCellsQueries:
    def test_get_cells_range_basic(self, backend_storage: Storage, cells_xlsx_bytes: bytes) -> None:
        from core.extractors.cells import extract_cells_to_sqlite

        # ジョブ作成 + cells.db 生成
        meta = backend_storage.create_job("p.xlsx", cells_xlsx_bytes)
        extract_cells_to_sqlite(
            backend_storage.get_original_path(meta.job_id),
            backend_storage.cells_db_path(meta.job_id),
        )
        # ヘッダ行を取得
        result = backend_storage.get_cells_range(meta.job_id, "Portfolio", "A1:E1")
        assert result["sheet"] == "Portfolio"
        assert result["origin_row"] == 1
        assert result["origin_col"] == 1
        rows = result["rows"]
        assert len(rows) == 1
        assert rows[0][0]["value"] == "銘柄コード"
        assert rows[0][4]["value"] == "評価損益"

    def test_get_cells_range_invalid_range(
        self, backend_storage: Storage, cells_xlsx_bytes: bytes
    ) -> None:
        from core.extractors.cells import extract_cells_to_sqlite

        meta = backend_storage.create_job("p.xlsx", cells_xlsx_bytes)
        extract_cells_to_sqlite(
            backend_storage.get_original_path(meta.job_id),
            backend_storage.cells_db_path(meta.job_id),
        )
        with pytest.raises(ValueError):
            backend_storage.get_cells_range(meta.job_id, "Portfolio", "garbage")

    def test_get_cells_range_no_db(self, backend_storage: Storage) -> None:
        meta = backend_storage.create_job("p.xlsx", b"x")  # cells.db 未生成
        with pytest.raises(FileNotFoundError):
            backend_storage.get_cells_range(meta.job_id, "S", "A1:B2")

    def test_find_cells_basic(self, backend_storage: Storage, cells_xlsx_bytes: bytes) -> None:
        from core.extractors.cells import extract_cells_to_sqlite

        meta = backend_storage.create_job("p.xlsx", cells_xlsx_bytes)
        extract_cells_to_sqlite(
            backend_storage.get_original_path(meta.job_id),
            backend_storage.cells_db_path(meta.job_id),
        )
        # "評価損益" 文字列で検索 (ヘッダにヒット)
        hits = backend_storage.find_cells(meta.job_id, "評価損益")
        assert len(hits) >= 1
        assert hits[0]["coord"] == "E1"

    def test_find_cells_sheet_filter(
        self, backend_storage: Storage, cells_xlsx_bytes: bytes
    ) -> None:
        from core.extractors.cells import extract_cells_to_sqlite

        meta = backend_storage.create_job("p.xlsx", cells_xlsx_bytes)
        extract_cells_to_sqlite(
            backend_storage.get_original_path(meta.job_id),
            backend_storage.cells_db_path(meta.job_id),
        )
        # 存在しないシート名で絞ると 0 件
        assert backend_storage.find_cells(meta.job_id, "ABC", sheet="NoSuch") == []
        # 正しいシート名で絞ると1件以上
        assert len(backend_storage.find_cells(meta.job_id, "ABC", sheet="Portfolio")) >= 1

    def test_find_cells_empty_query(
        self, backend_storage: Storage, cells_xlsx_bytes: bytes
    ) -> None:
        from core.extractors.cells import extract_cells_to_sqlite

        meta = backend_storage.create_job("p.xlsx", cells_xlsx_bytes)
        extract_cells_to_sqlite(
            backend_storage.get_original_path(meta.job_id),
            backend_storage.cells_db_path(meta.job_id),
        )
        assert backend_storage.find_cells(meta.job_id, "") == []

    def test_find_cells_limit(self, backend_storage: Storage, cells_xlsx_bytes: bytes) -> None:
        from core.extractors.cells import extract_cells_to_sqlite

        meta = backend_storage.create_job("p.xlsx", cells_xlsx_bytes)
        extract_cells_to_sqlite(
            backend_storage.get_original_path(meta.job_id),
            backend_storage.cells_db_path(meta.job_id),
        )
        # 株式会社 で複数ヒットするが limit=1 なら1件
        hits = backend_storage.find_cells(meta.job_id, "株式会社", limit=1)
        assert len(hits) == 1


# ---------- /cells/* API ----------


class TestCellsRoutes:
    def _setup_job(self, client: TestClient, body: bytes) -> str:
        r = client.post(
            "/extract",
            files={"file": ("p.xlsx", body, "application/octet-stream")},
        )
        assert r.status_code == 200, r.text
        return r.json()["job_id"]

    def test_get_range(self, client: TestClient, cells_xlsx_bytes: bytes) -> None:
        job_id = self._setup_job(client, cells_xlsx_bytes)
        r = client.get(
            f"/cells/{job_id}/range",
            params={"sheet": "Portfolio", "range": "A1:E1"},
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["sheet"] == "Portfolio"
        assert body["origin_row"] == 1
        assert len(body["rows"]) == 1
        assert body["rows"][0][0]["value"] == "銘柄コード"

    def test_get_range_invalid(self, client: TestClient, cells_xlsx_bytes: bytes) -> None:
        job_id = self._setup_job(client, cells_xlsx_bytes)
        r = client.get(
            f"/cells/{job_id}/range",
            params={"sheet": "Portfolio", "range": "not-a-range"},
        )
        assert r.status_code == 400

    def test_get_range_missing_job(self, client: TestClient) -> None:
        r = client.get(
            f"/cells/{uuid.uuid4()}/range",
            params={"sheet": "S", "range": "A1:B2"},
        )
        # 409 (cells.db 不在) または 404 のどちらか
        assert r.status_code in (404, 409)

    def test_get_range_invalid_job_id(self, client: TestClient) -> None:
        r = client.get(
            "/cells/not-uuid/range",
            params={"sheet": "S", "range": "A1:B2"},
        )
        assert r.status_code == 400

    def test_find(self, client: TestClient, cells_xlsx_bytes: bytes) -> None:
        job_id = self._setup_job(client, cells_xlsx_bytes)
        r = client.get(f"/cells/{job_id}/find", params={"q": "評価損益"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["count"] >= 1
        assert any(m["coord"] == "E1" for m in body["matches"])

    def test_find_sheet_filter(self, client: TestClient, cells_xlsx_bytes: bytes) -> None:
        job_id = self._setup_job(client, cells_xlsx_bytes)
        r = client.get(
            f"/cells/{job_id}/find",
            params={"q": "ABC", "sheet": "Portfolio"},
        )
        assert r.status_code == 200
        assert r.json()["count"] >= 1

    def test_find_limit_validation(self, client: TestClient, cells_xlsx_bytes: bytes) -> None:
        job_id = self._setup_job(client, cells_xlsx_bytes)
        # limit=0 は invalid
        assert client.get(f"/cells/{job_id}/find", params={"q": "x", "limit": 0}).status_code == 422
        # limit が大きすぎる場合も invalid
        assert (
            client.get(f"/cells/{job_id}/find", params={"q": "x", "limit": 9999}).status_code == 422
        )


class TestExtractBuildsCellsDb:
    def test_xlsx_creates_cells_db(
        self, client: TestClient, cells_xlsx_bytes: bytes, backend_storage: Storage
    ) -> None:
        job_id = client.post(
            "/extract",
            files={"file": ("p.xlsx", cells_xlsx_bytes, "application/octet-stream")},
        ).json()["job_id"]
        assert backend_storage.has_cells_db(job_id)

    def test_xls_is_rejected_before_cells_db(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        # .xls は解析できないためアップロード時点で 415 で弾かれ、ジョブも cells.db も作られない
        r = client.post(
            "/extract",
            files={"file": ("a.xls", b"not real", "application/octet-stream")},
        )
        assert r.status_code == 415, r.text
        assert backend_storage.list_jobs() == []
