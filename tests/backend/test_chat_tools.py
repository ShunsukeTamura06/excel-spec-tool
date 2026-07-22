"""backend.routes.chat の function calling ループ統合テスト.

MockLLMClient.queue_response() でツール呼び出しをスクリプト化し、
チャットルートが想定通りに tool ループを回すことを検証する。
"""

import json
from collections.abc import Iterator
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook as OpyWorkbook

from backend.dependencies import get_llm_client, get_storage
from backend.llm_client import LLMResponse, LLMToolCall, MockLLMClient
from backend.main import create_app
from backend.storage import Storage


@pytest.fixture
def storage(tmp_path: Path) -> Storage:
    return Storage(tmp_path / "jobs")


@pytest.fixture
def scripted_llm() -> MockLLMClient:
    """テストごとに queue_response でスクリプトを組む."""
    return MockLLMClient()


@pytest.fixture
def app_client(storage: Storage, scripted_llm: MockLLMClient) -> Iterator[TestClient]:
    app = create_app()
    app.dependency_overrides[get_storage] = lambda: storage
    app.dependency_overrides[get_llm_client] = lambda: scripted_llm
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture
def xlsx_bytes(tmp_path: Path) -> bytes:
    wb = OpyWorkbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Portfolio"
    headers = ["銘柄コード", "銘柄名", "保有口数", "現在値", "評価損益", "実現損益"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value="ABC")
    ws.cell(row=2, column=2, value="株式会社A")
    out = tmp_path / "p.xlsx"
    wb.save(out)
    return out.read_bytes()


def _setup_job(client: TestClient, body: bytes) -> str:
    r = client.post(
        "/extract",
        files={"file": ("p.xlsx", body, "application/octet-stream")},
    )
    assert r.status_code == 200, r.text
    return r.json()["job_id"]


class TestToolLoop:
    def test_vba_package_trace_uses_short_windows_guidance(self) -> None:
        """VBA提案カードがある場合は長文手順でなくZIP導線を返す."""

        from backend.routes.chat import _final_reply_for_tool_trace

        reply = _final_reply_for_tool_trace(
            "長い手順",
            [{"name": "propose_vba_procedure_replace", "result": {}}],
        )

        assert "Windows用ZIP" in reply
        assert "revised.xlsm" in reply
        assert "長い手順" not in reply

    def test_no_tool_calls_returns_simple_reply(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        # 1ターンだけで終わるシナリオ
        scripted_llm.queue_response(LLMResponse(content="ok"))

        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "hi"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["reply"] == "ok"
        assert body["tool_trace"] == []

    def test_does_not_claim_nonexistent_change_card(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
    ) -> None:
        """propose未実行の回答から、存在しないボタン案内を除去する."""

        scripted_llm.queue_response(
            LLMResponse(
                content=(
                    "入力規則とデータ検証の修正カードを作りました。"
                    "画面の「修正版を作る」ボタンを押してください。"
                )
            )
        )

        job_id = _setup_job(app_client, xlsx_bytes)
        response = app_client.post(
            f"/chat/{job_id}",
            json={"message": "入力規則とデータ検証を追加したい"},
        )

        assert response.status_code == 200, response.text
        body = response.json()
        assert body["tool_trace"] == []
        assert body["reply"] == (
            "この依頼では変更カードを作成できませんでした。"
            "そのため「修正版を作る」ボタンは表示されません。\n\n"
            "入力規則やデータ検証の設定は、現在の自動変更範囲外です。"
        )

    def test_cell_text_request_retries_until_change_card_is_created(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
    ) -> None:
        """安価モデルが手順だけ返しても、対応可能な依頼は変更カードへ再誘導する."""

        scripted_llm.queue_response(LLMResponse(content="G列へ手で説明を入力してください"))
        scripted_llm.queue_response(
            LLMResponse(
                tool_calls=[
                    LLMToolCall(
                        id="range",
                        name="get_cells_range",
                        arguments={"sheet": "Portfolio", "range": "G1:G2"},
                    )
                ]
            )
        )
        scripted_llm.queue_response(LLMResponse(content="確認できました"))
        scripted_llm.queue_response(
            LLMResponse(
                tool_calls=[
                    LLMToolCall(
                        id="proposal",
                        name="propose_cell_text_edits",
                        arguments={
                            "edits": [
                                {
                                    "sheet": "Portfolio",
                                    "coord": "G1",
                                    "value": "各列の説明",
                                }
                            ]
                        },
                    )
                ]
            )
        )
        scripted_llm.queue_response(LLMResponse(content="変更内容を確認してください"))

        job_id = _setup_job(app_client, xlsx_bytes)
        response = app_client.post(
            f"/chat/{job_id}",
            json={"message": "Portfolioシートに説明を追加したい"},
        )

        assert response.status_code == 200, response.text
        body = response.json()
        assert body["reply"] == (
            "1個の空セルへ説明テキストを追加します。\n\n"
            "下の「変更内容を確認」で内容を見て、「修正版を作る」を押してください。"
            "原本は変更されません。"
        )
        assert [item["name"] for item in body["tool_trace"]] == [
            "get_cells_range",
            "propose_cell_text_edits",
        ]
        proposal = body["tool_trace"][1]["result"]["safe_plan"]
        assert proposal["plan"]["requested_provider"] == "officecli"

    def test_single_tool_call_then_final_reply(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        # 1) LLM が find_cells を呼ぶ
        # 2) tool 結果を受け取って最終応答
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[
                    LLMToolCall(
                        id="t1",
                        name="find_cells",
                        arguments={"query": "実現損益"},
                    )
                ],
            )
        )
        scripted_llm.queue_response(LLMResponse(content="実現損益は F1 です"))

        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "実現損益はどこ？"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["reply"] == "実現損益は F1 です"
        assert len(body["tool_trace"]) == 1
        assert body["tool_trace"][0]["name"] == "find_cells"
        assert body["tool_trace"][0]["arguments"] == {"query": "実現損益"}
        assert body["tool_trace"][0]["result"]["count"] >= 1
        assistant_messages = [m for m in body["history"] if m["role"] == "assistant"]
        assert assistant_messages[-1]["tool_trace"][0]["name"] == "find_cells"
        assert assistant_messages[-1]["tool_trace"][0]["result"]["count"] >= 1

    def test_stream_returns_progress_events_and_final_reply(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[
                    LLMToolCall(
                        id="t1",
                        name="find_cells",
                        arguments={"query": "実現損益"},
                    )
                ],
            )
        )
        scripted_llm.queue_response(LLMResponse(content="実現損益は F1 です"))

        job_id = _setup_job(app_client, xlsx_bytes)
        with app_client.stream(
            "POST",
            f"/chat/{job_id}/stream",
            json={"message": "実現損益はどこ？"},
        ) as r:
            assert r.status_code == 200, r.text
            body = r.read().decode()

        assert "event: status" in body
        assert "event: tool_start" in body
        assert "event: tool_result" in body
        assert "event: final" in body
        final_payload = body.split("event: final\n", maxsplit=1)[1].split("\n\n", maxsplit=1)[0]
        final_json = json.loads(final_payload.removeprefix("data: "))
        assert final_json["reply"] == "実現損益は F1 です"
        assert final_json["tool_trace"][0]["name"] == "find_cells"
        assert "result" in final_json["tool_trace"][0]
        assistant_messages = [m for m in final_json["history"] if m["role"] == "assistant"]
        assert assistant_messages[-1]["tool_trace"][0]["name"] == "find_cells"
        assert "result" in assistant_messages[-1]["tool_trace"][0]

    def test_multiple_tool_calls(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        # find -> get_range -> 最終応答
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[
                    LLMToolCall(id="t1", name="find_cells", arguments={"query": "実現損益"})
                ],
            )
        )
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[
                    LLMToolCall(
                        id="t2",
                        name="get_cells_range",
                        arguments={"sheet": "Portfolio", "range": "A1:F1"},
                    )
                ],
            )
        )
        scripted_llm.queue_response(LLMResponse(content="G 列に追加してください"))

        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "MTM 列追加"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["reply"] == "G 列に追加してください"
        assert len(body["tool_trace"]) == 2
        assert [t["name"] for t in body["tool_trace"]] == [
            "find_cells",
            "get_cells_range",
        ]

    def test_tool_loop_caps_iterations(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        # 無限ループ風: 上限回数までは常に tool_call を返す
        from backend.routes.chat import MAX_TOOL_ITERATIONS

        for i in range(MAX_TOOL_ITERATIONS):
            scripted_llm.queue_response(
                LLMResponse(
                    content=None,
                    tool_calls=[
                        LLMToolCall(
                            id=f"t{i}",
                            name="find_cells",
                            arguments={"query": "x"},
                        )
                    ],
                )
            )
        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "loop"})
        assert r.status_code == 200
        # MAX_TOOL_ITERATIONS で打ち切られていること
        body = r.json()

        assert len(body["tool_trace"]) == MAX_TOOL_ITERATIONS
        assert body["reply"] == "[mock:mock-pro] received: loop"
        assert "tool loop max iterations" not in body["reply"]
        assert "ツール確認が上限に達し" not in body["reply"]

    def test_tool_loop_fallback_history_is_not_sent_to_llm(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
        storage: Storage,
    ) -> None:
        """過去の上限到達メッセージで、同一セッションの後続会話を汚染しない."""
        from core.models import ChatMessage

        scripted_llm.queue_response(LLMResponse(content="fresh answer"))
        job_id = _setup_job(app_client, xlsx_bytes)
        storage.append_chat_message(
            job_id,
            ChatMessage(role="user", content="old question", timestamp="2026-01-01T00:00:00Z"),
        )
        storage.append_chat_message(
            job_id,
            ChatMessage(
                role="assistant",
                content="ツール確認が上限に達し、最終回答を生成できませんでした。質問範囲を少し絞ってもう一度聞いてください。",
                timestamp="2026-01-01T00:00:01Z",
            ),
        )

        r = app_client.post(f"/chat/{job_id}", json={"message": "new question"})

        assert r.status_code == 200
        assert r.json()["reply"] == "fresh answer"
        sent_contents = [str(m.get("content", "")) for m in scripted_llm.calls[-1]["messages"]]
        assert not any("ツール確認が上限に達し" in c for c in sent_contents)

    def test_history_contains_only_user_and_assistant(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        # tool 呼び出しを挟んでも、保存される履歴は user/assistant ペアのみ
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[LLMToolCall(id="t1", name="find_cells", arguments={"query": "x"})],
            )
        )
        scripted_llm.queue_response(LLMResponse(content="final"))

        job_id = _setup_job(app_client, xlsx_bytes)
        app_client.post(f"/chat/{job_id}", json={"message": "q"})

        r = app_client.get(f"/chat/{job_id}/history")
        history = r.json()["history"]
        assert len(history) == 2
        assert {h["role"] for h in history} == {"user", "assistant"}

    def test_unknown_tool_does_not_crash(
        self, app_client: TestClient, xlsx_bytes: bytes, scripted_llm: MockLLMClient
    ) -> None:
        # 存在しない tool 名を LLM が要求してもクラッシュしない
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[LLMToolCall(id="t1", name="no_such_tool", arguments={})],
            )
        )
        scripted_llm.queue_response(LLMResponse(content="recovered"))

        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "x"})
        assert r.status_code == 200
        assert r.json()["reply"] == "recovered"

    def test_chat_always_pro(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
    ) -> None:
        """チャットツールループは prompt caching を活かすため tier=pro 固定."""
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[LLMToolCall(id="t1", name="find_cells", arguments={"query": "x"})],
            )
        )
        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[
                    LLMToolCall(
                        id="t2",
                        name="get_cells_range",
                        arguments={"sheet": "Portfolio", "range": "A1:B1"},
                    )
                ],
            )
        )
        scripted_llm.queue_response(LLMResponse(content="done"))

        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "q"})
        assert r.status_code == 200

        tiers = [c["tier"] for c in scripted_llm.calls]
        assert tiers == ["pro", "pro", "pro"]
        models = [c["model"] for c in scripted_llm.calls]
        assert all(m == "mock-pro" for m in models)

    def test_simple_question_also_pro(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
    ) -> None:
        """単純な質問 (1 反復) も pro で完結する."""
        scripted_llm.queue_response(LLMResponse(content="simple answer"))

        job_id = _setup_job(app_client, xlsx_bytes)
        r = app_client.post(f"/chat/{job_id}", json={"message": "q"})
        assert r.status_code == 200
        tiers = [c["tier"] for c in scripted_llm.calls]
        assert tiers == ["pro"]

    def test_history_trimmed_when_over_limit(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
        storage: Storage,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """Phase B-2: 履歴が制限を超えると古い分は summary system message に集約される."""
        from core.models import ChatMessage

        # 上限を 2 ペアに絞ってテストしやすくする
        monkeypatch.setenv("CHAT_HISTORY_LIMIT_PAIRS", "2")

        scripted_llm.queue_response(LLMResponse(content="latest answer"))

        job_id = _setup_job(app_client, xlsx_bytes)
        # 5 ペア分の古い履歴を仕込む (10 メッセージ)
        ts = "2026-01-01T00:00:00Z"
        for i in range(5):
            storage.append_chat_message(
                job_id,
                ChatMessage(role="user", content=f"old question {i}", timestamp=ts),
            )
            storage.append_chat_message(
                job_id,
                ChatMessage(role="assistant", content=f"old answer {i}", timestamp=ts),
            )

        r = app_client.post(f"/chat/{job_id}", json={"message": "new question"})
        assert r.status_code == 200
        assert r.json()["reply"] == "latest answer"

        # MockLLMClient に渡された messages を検証
        sent_messages = scripted_llm.calls[-1]["messages"]
        # 構成: [system(spec), system(summary), recent 2 ペア = 4 msgs, user(new) ] = 7
        assert len(sent_messages) == 7
        # 最初の system は spec を含む大きいやつ
        assert sent_messages[0]["role"] == "system"
        # 2 つ目の system が summary
        assert sent_messages[1]["role"] == "system"
        assert "過去のやりとり概要" in sent_messages[1]["content"]
        # 古い質問は summary に入っている (recent には入っていない)
        assert "old question 0" in sent_messages[1]["content"]
        # 直近 2 ペアのみが messages に並ぶ
        kept = [m for m in sent_messages if m["role"] in ("user", "assistant")]
        assert [m["content"] for m in kept] == [
            "old question 3",
            "old answer 3",
            "old question 4",
            "old answer 4",
            "new question",
        ]

    def test_short_history_no_summary(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
        storage: Storage,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """履歴が短ければ summary は付与されない."""
        from core.models import ChatMessage

        monkeypatch.setenv("CHAT_HISTORY_LIMIT_PAIRS", "5")
        scripted_llm.queue_response(LLMResponse(content="ok"))

        job_id = _setup_job(app_client, xlsx_bytes)
        storage.append_chat_message(job_id, ChatMessage(role="user", content="q1", timestamp="t1"))
        storage.append_chat_message(
            job_id, ChatMessage(role="assistant", content="a1", timestamp="t1")
        )

        r = app_client.post(f"/chat/{job_id}", json={"message": "q2"})
        assert r.status_code == 200
        sent_messages = scripted_llm.calls[-1]["messages"]
        # system 1 個のみ (summary は付かない)
        system_msgs = [m for m in sent_messages if m["role"] == "system"]
        assert len(system_msgs) == 1
        assert "過去のやりとり概要" not in system_msgs[0]["content"]

    def test_second_chat_call_hits_cache(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Phase B-3: 同じジョブで 2 回チャットすると 2 回目は cache 命中."""
        scripted_llm.queue_response(LLMResponse(content="first"))
        scripted_llm.queue_response(LLMResponse(content="second"))

        job_id = _setup_job(app_client, xlsx_bytes)
        with caplog.at_level("INFO", logger="backend.routes.chat"):
            r1 = app_client.post(f"/chat/{job_id}", json={"message": "q1"})
            r2 = app_client.post(f"/chat/{job_id}", json={"message": "q2"})
        assert r1.status_code == 200
        assert r2.status_code == 200

        # 両方とも cache_prefix=1 で呼ばれている
        assert scripted_llm.calls[0]["cache_prefix"] == 1
        assert scripted_llm.calls[1]["cache_prefix"] == 1

        # ログを解析: llm usage の cached=N を順に取り出す
        usage_logs = [r.message for r in caplog.records if "llm usage iter=" in r.message]
        assert len(usage_logs) == 2

        def _cached(log: str) -> int:
            # "...cached=N" を抽出
            for tok in log.split():
                if tok.startswith("cached="):
                    return int(tok.split("=", 1)[1])
            return -1

        assert _cached(usage_logs[0]) == 0  # 初回は miss
        assert _cached(usage_logs[1]) > 0  # 2 回目は hit

    def test_cumulative_usage_logged(
        self,
        app_client: TestClient,
        xlsx_bytes: bytes,
        scripted_llm: MockLLMClient,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        """tool ループ完了時に累計トークンがサマリログに出力される."""
        from backend.llm_client import Usage

        scripted_llm.queue_response(
            LLMResponse(
                content=None,
                tool_calls=[LLMToolCall(id="t1", name="find_cells", arguments={"query": "x"})],
                usage=Usage(prompt_tokens=100, completion_tokens=20, total_tokens=120),
                model="fast",
            )
        )
        scripted_llm.queue_response(
            LLMResponse(
                content="done",
                usage=Usage(prompt_tokens=180, completion_tokens=10, total_tokens=190),
                model="pro",
            )
        )

        job_id = _setup_job(app_client, xlsx_bytes)
        with caplog.at_level("INFO", logger="backend.routes.chat"):
            r = app_client.post(f"/chat/{job_id}", json={"message": "q"})
        assert r.status_code == 200

        # 各イテレーションの usage ログが出ている
        usage_logs = [rec.message for rec in caplog.records if "llm usage iter=" in rec.message]
        assert len(usage_logs) == 2
        assert "model=fast" in usage_logs[0]
        assert "model=pro" in usage_logs[1]

        # 最終サマリに累計が乗っている
        summary = [rec.message for rec in caplog.records if "llm final response:" in rec.message]
        assert len(summary) == 1
        # 100 + 180 = 280, 20 + 10 = 30, 120 + 190 = 310
        assert "cumulative_prompt=280" in summary[0]
        assert "completion=30" in summary[0]
        assert "total=310" in summary[0]
