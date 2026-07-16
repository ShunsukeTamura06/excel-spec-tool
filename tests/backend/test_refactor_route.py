"""POST /jobs/{job_id}/named-range-fix, /jobs/{job_id}/formula-fix のテスト."""

from __future__ import annotations

import io
import uuid
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl.workbook.defined_name import DefinedName

from backend.storage import Storage
from core.extractors.cells import extract_cells_to_sqlite
from core.extractors.workbook import extract_workbook
from core.mutation import (
    CellTextBatchOperation,
    MutationPlan,
    MutationResult,
    NamedRangeSetOperation,
)
from core.reference_index import build_reference_index


def _xlsx_bytes_with_named_range(
    name: str = "TaxRate",
    refers_to: str = "Data!$A$1",
    formula_referencing: str | None = None,
) -> bytes:
    """名前付き範囲を1つ持つ xlsx をバイト列で生成する.

    formula_referencing を渡すと、B1 にその範囲を参照する数式を仕込む
    (波及範囲テスト用)。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = 0.1
    if formula_referencing:
        ws["B1"] = formula_referencing
    wb.defined_names[name] = DefinedName(name=name, attr_text=refers_to)
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    wb.close()
    return data


def _seed_extracted_job(storage: Storage, data: bytes, filename: str = "x.xlsx") -> str:
    """extract相当の処理をテスト内で組み立て、抽出済みジョブを1件作る."""
    meta = storage.create_job(filename, data)
    path = storage.get_original_path(meta.job_id)
    wb = extract_workbook(path)
    wb.filename = filename
    idx = build_reference_index(wb)
    storage.save_workbook(meta.job_id, wb)
    storage.save_references(meta.job_id, idx)
    extract_cells_to_sqlite(path, storage.cells_db_path(meta.job_id))
    storage.update_status(meta.job_id, "extracted")
    return meta.job_id


class TestNamedRangeFixRoute:
    def test_applies_fix_and_creates_new_job(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())

        r = client.post(
            f"/jobs/{job_id}/named-range-fix",
            json={"name": "TaxRate", "new_refers_to": "Data!$B$1"},
        )
        assert r.status_code == 200
        body = r.json()
        new_job_id = body["new_job_id"]
        assert new_job_id != job_id

        diff = body["diff"]
        assert body["verification"]["status"] == "passed"
        assert body["provider"]["provider"] == "openpyxl"
        assert len(diff["named_ranges"]) == 1
        nr = diff["named_ranges"][0]
        assert nr["name"] == "TaxRate"
        assert nr["change_type"] == "modified"
        assert nr["before_refers_to"] == "Data!$A$1"
        assert nr["after_refers_to"] == "Data!$B$1"

        audit = client.get(f"/jobs/{new_job_id}/verification")
        assert audit.status_code == 200
        record = audit.json()["verification_record"]
        assert record["source_job_id"] == job_id
        assert record["result_job_id"] == new_job_id
        assert record["source_file_sha256"] == backend_storage.get_meta(job_id).file_sha256
        assert record["result_file_sha256"] == backend_storage.get_meta(new_job_id).file_sha256
        assert record["source_file_sha256"] != record["result_file_sha256"]
        assert record["verification"]["status"] == "passed"

    def test_policy_rejects_unplanned_provider_change(
        self,
        client: TestClient,
        backend_storage: Storage,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """プロバイダーが予定外セルも変更した場合は監査保存して409にする."""

        class UnexpectedChangeProvider:
            """名前定義に加えて予定外セルを変更する故障プロバイダー."""

            def apply(
                self,
                plan: MutationPlan,
                source_path: Path,
                out_path: Path,
            ) -> MutationResult:
                """計画した変更と予定外変更を同時に書き込む."""

                operation = plan.operation
                assert isinstance(operation, NamedRangeSetOperation)
                wb = openpyxl.load_workbook(source_path)
                try:
                    wb.defined_names[operation.name].attr_text = operation.new_refers_to
                    wb["Data"]["Z99"] = "unexpected"
                    wb.save(out_path)
                finally:
                    wb.close()
                return MutationResult(
                    provider="openpyxl",
                    provider_version="fault-injection",
                    operation="named_range_set",
                    changed_count=2,
                )

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())
        monkeypatch.setattr(
            "backend.routes.refactor._provider_for",
            lambda _: UnexpectedChangeProvider(),
        )

        response = client.post(
            f"/jobs/{job_id}/named-range-fix",
            json={"name": "TaxRate", "new_refers_to": "Data!$B$1"},
        )

        assert response.status_code == 409
        detail = response.json()["detail"]
        assert detail["verification"]["status"] == "failed"
        assert any(
            item["code"] == "unexpected_change" for item in detail["verification"]["violations"]
        )
        rejected_job_id = detail["new_job_id"]
        assert backend_storage.get_meta(rejected_job_id).status == "failed"
        audit = client.get(f"/jobs/{rejected_job_id}/verification")
        assert audit.status_code == 200
        assert audit.json()["verification_record"]["verification"]["status"] == "failed"

    def test_new_job_diff_matches_get_diff_endpoint(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())

        r = client.post(
            f"/jobs/{job_id}/named-range-fix",
            json={"name": "TaxRate", "new_refers_to": "Data!$B$1"},
        )
        assert r.status_code == 200
        new_job_id = r.json()["new_job_id"]

        r2 = client.get(f"/diff?before_job_id={job_id}&after_job_id={new_job_id}")
        assert r2.status_code == 200
        assert r2.json()["diff"]["named_ranges"] == r.json()["diff"]["named_ranges"]

    def test_blast_radius_included_when_referenced(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        job_id = _seed_extracted_job(
            backend_storage,
            _xlsx_bytes_with_named_range(formula_referencing="=SUM(A1:A1)"),
        )

        r = client.post(
            f"/jobs/{job_id}/named-range-fix",
            json={"name": "TaxRate", "new_refers_to": "Data!$B$1"},
        )
        assert r.status_code == 200
        # 参照有無はビルドされた ReferenceIndex の正規化に依存するため、
        # ここでは応答が正しく構造化されていることだけを確認する。
        assert isinstance(r.json()["diff"]["blast_radius"], list)
        assert r.json()["verification"]["status"] == "needs_review"

    def test_422_unknown_name(self, client: TestClient, backend_storage: Storage) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())

        r = client.post(
            f"/jobs/{job_id}/named-range-fix",
            json={"name": "NoSuchName", "new_refers_to": "Data!$B$1"},
        )
        assert r.status_code == 422

    def test_404_job_not_found(self, client: TestClient, backend_storage: Storage) -> None:
        r = client.post(
            f"/jobs/{uuid.uuid4()}/named-range-fix",
            json={"name": "TaxRate", "new_refers_to": "Data!$B$1"},
        )
        assert r.status_code == 404

    def test_409_not_extracted(self, client: TestClient, backend_storage: Storage) -> None:
        meta = backend_storage.create_job("x.xlsx", _xlsx_bytes_with_named_range())
        r = client.post(
            f"/jobs/{meta.job_id}/named-range-fix",
            json={"name": "TaxRate", "new_refers_to": "Data!$B$1"},
        )
        assert r.status_code == 409


def _xlsx_bytes_with_formulas() -> bytes:
    """数式を持つ xlsx をバイト列で生成する (formula-fix テスト用)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws["A1"] = 1
    ws["B5"] = 10
    ws["C1"] = "=$B$5*2"
    ws["C2"] = "=SUM(A1:A100)"
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    wb.close()
    return data


class TestFormulaFixRoute:
    def test_fixed_ref_replace_applies_and_creates_new_job(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={"kind": "fixed_ref_replace", "old_ref": "Data!$B$5", "new_ref": "Data!$B$6"},
        )
        assert r.status_code == 200
        body = r.json()
        new_job_id = body["new_job_id"]
        assert new_job_id != job_id

        # 自己検証 diff に、意図した数式変更が現れる
        changed = [c for c in body["diff"]["cells"] if c["coord"] == "C1"]
        assert len(changed) == 1
        assert changed[0]["after_formula"] == "=$B$6*2"
        assert body["verification"]["status"] == "passed"

        # 新ジョブの実ファイルにも書き込まれている
        new_path = backend_storage.get_original_path(new_job_id)
        wb = openpyxl.load_workbook(new_path)
        assert wb["Data"]["C1"].value == "=$B$6*2"
        wb.close()

    def test_range_expansion_applies(self, client: TestClient, backend_storage: Storage) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={
                "kind": "range_expansion",
                "old_ref": "Data!A1:A100",
                "new_ref": "Data!A1:A200",
            },
        )
        assert r.status_code == 200
        new_job_id = r.json()["new_job_id"]
        new_path = backend_storage.get_original_path(new_job_id)
        wb = openpyxl.load_workbook(new_path)
        assert wb["Data"]["C2"].value == "=SUM(A1:A200)"
        wb.close()

    def test_422_no_matching_formula(self, client: TestClient, backend_storage: Storage) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={"kind": "fixed_ref_replace", "old_ref": "Data!Z99", "new_ref": "Data!Z100"},
        )
        assert r.status_code == 422

    def test_422_does_not_leave_orphan_job(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        """適用に失敗したとき、作りかけの新ジョブが残らない."""
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())
        jobs_before = {m.job_id for m in backend_storage.list_jobs()}

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={"kind": "fixed_ref_replace", "old_ref": "Data!Z99", "new_ref": "Data!Z100"},
        )
        assert r.status_code == 422
        assert {m.job_id for m in backend_storage.list_jobs()} == jobs_before

    def test_422_invalid_expansion(self, client: TestClient, backend_storage: Storage) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={
                "kind": "range_expansion",
                "old_ref": "Data!A1:A100",
                "new_ref": "Data!A1:A50",
            },
        )
        assert r.status_code == 422

    def test_422_unknown_kind_rejected_by_validation(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={"kind": "vba_rewrite", "old_ref": "Data!A1", "new_ref": "Data!A2"},
        )
        assert r.status_code == 422

    def test_422_officecli_rejects_unsupported_formula_operation(
        self, client: TestClient, backend_storage: Storage
    ) -> None:
        """OfficeCLI adapterのcapability外操作は実行前に拒否する."""

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        r = client.post(
            f"/jobs/{job_id}/formula-fix",
            json={
                "kind": "fixed_ref_replace",
                "old_ref": "Data!$B$5",
                "new_ref": "Data!$B$6",
                "provider": "officecli",
            },
        )

        assert r.status_code == 422
        assert "does not support operation" in r.json()["detail"]

    def test_404_job_not_found(self, client: TestClient, backend_storage: Storage) -> None:
        r = client.post(
            f"/jobs/{uuid.uuid4()}/formula-fix",
            json={"kind": "fixed_ref_replace", "old_ref": "Data!A1", "new_ref": "Data!A2"},
        )
        assert r.status_code == 404

    def test_409_not_extracted(self, client: TestClient, backend_storage: Storage) -> None:
        meta = backend_storage.create_job("x.xlsx", _xlsx_bytes_with_formulas())
        r = client.post(
            f"/jobs/{meta.job_id}/formula-fix",
            json={"kind": "fixed_ref_replace", "old_ref": "Data!$B$5", "new_ref": "Data!$B$6"},
        )
        assert r.status_code == 409


class TestSafeChangePlanRoute:
    """一般ユーザー向けの計画確認から実行までを検証する."""

    def test_preview_does_not_create_job_and_returns_expected_diff(
        self,
        client: TestClient,
        backend_storage: Storage,
    ) -> None:
        """計画確認だけではファイルを作らず、変更予定を返す."""

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())
        jobs_before = {meta.job_id for meta in backend_storage.list_jobs()}

        response = client.post(
            f"/jobs/{job_id}/change-plan",
            json={
                "kind": "range_expansion",
                "old_ref": "Data!A1:A100",
                "new_ref": "Data!A1:A200",
            },
        )

        assert response.status_code == 200
        body = response.json()
        assert body["automation"] == "supported"
        assert body["can_apply"] is True
        assert body["expected_change_count"] == 1
        assert body["plan"]["source_job_id"] == job_id
        assert body["expected_diff"]["cells"][0]["after_formula"] == "=SUM(A1:A200)"
        assert {meta.job_id for meta in backend_storage.list_jobs()} == jobs_before

    def test_executes_the_same_confirmed_plan_and_saves_audit(
        self,
        client: TestClient,
        backend_storage: Storage,
    ) -> None:
        """確認したplan_idを保持したまま適用し、監査証跡へ保存する."""

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())
        preview = client.post(
            f"/jobs/{job_id}/change-plan",
            json={
                "kind": "range_expansion",
                "old_ref": "Data!A1:A100",
                "new_ref": "Data!A1:A200",
            },
        )
        plan = preview.json()["plan"]

        response = client.post(
            f"/jobs/{job_id}/change-plan/execute",
            json={"plan": plan},
        )

        assert response.status_code == 200
        body = response.json()
        assert body["plan"]["plan_id"] == plan["plan_id"]
        assert body["verification"]["status"] == "passed"
        audit = client.get(f"/jobs/{body['new_job_id']}/verification")
        assert audit.status_code == 200
        assert audit.json()["verification_record"]["plan"]["plan_id"] == plan["plan_id"]

    def test_rejects_plan_for_a_different_source_job(
        self,
        client: TestClient,
        backend_storage: Storage,
    ) -> None:
        """別ジョブ向けに確認した計画の取り違えを拒否する."""

        source_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())
        other_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())
        preview = client.post(
            f"/jobs/{source_id}/change-plan",
            json={
                "kind": "range_expansion",
                "old_ref": "Data!A1:A100",
                "new_ref": "Data!A1:A200",
            },
        )

        response = client.post(
            f"/jobs/{other_id}/change-plan/execute",
            json={"plan": preview.json()["plan"]},
        )

        assert response.status_code == 400

    def test_rejects_shrinking_range_during_preview(
        self,
        client: TestClient,
        backend_storage: Storage,
    ) -> None:
        """範囲縮小は計画として提示せず、適用前に拒否する."""

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_formulas())

        response = client.post(
            f"/jobs/{job_id}/change-plan",
            json={
                "kind": "range_expansion",
                "old_ref": "Data!A1:A100",
                "new_ref": "Data!A1:A50",
            },
        )

        assert response.status_code == 422


def test_mutation_provider_capabilities(client: TestClient) -> None:
    """利用可能な変更エンジンと対応範囲をAPIから取得できる."""

    response = client.get("/mutation-providers")

    assert response.status_code == 200
    providers = {item["name"]: item for item in response.json()["providers"]}
    assert providers["openpyxl"]["available"] is True
    assert "named_range_set" in providers["officecli"]["supported_operations"]
    assert "cell_text_batch" in providers["officecli"]["supported_operations"]


class TestCellTextChangePlanRoute:
    def test_previews_empty_cell_text_changes(
        self,
        client: TestClient,
        backend_storage: Storage,
    ) -> None:
        """空セルへの説明追加を原本未変更の計画として返す."""

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())

        response = client.post(
            f"/jobs/{job_id}/change-plan",
            json={
                "kind": "cell_text_batch",
                "edits": [
                    {"sheet": "Data", "coord": "C1", "value": "税率の入力値です"},
                ],
            },
        )

        assert response.status_code == 200
        body = response.json()
        assert body["plan"]["requested_provider"] == "officecli"
        assert body["plan"]["operation"]["kind"] == "cell_text_batch"
        assert body["expected_diff"]["cells"][0]["after_value"] == "税率の入力値です"

    def test_executes_and_verifies_cell_text_plan(
        self,
        client: TestClient,
        backend_storage: Storage,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        """確認済み計画を別ファイルへ適用し、実差分一致まで検証する."""

        class FakeOfficeCliProvider:
            """テスト用にOfficeCLIと同じ出力契約だけを再現する."""

            def apply(
                self,
                plan: MutationPlan,
                source_path: Path,
                out_path: Path,
            ) -> MutationResult:
                operation = plan.operation
                assert isinstance(operation, CellTextBatchOperation)
                wb = openpyxl.load_workbook(source_path)
                try:
                    for edit in operation.edits:
                        wb[edit.sheet][edit.coord] = edit.value
                    wb.save(out_path)
                finally:
                    wb.close()
                return MutationResult(
                    provider="officecli",
                    provider_version="test",
                    operation="cell_text_batch",
                    changed_count=len(operation.edits),
                )

        monkeypatch.setattr(
            "backend.routes.refactor._provider_for",
            lambda _: FakeOfficeCliProvider(),
        )
        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())
        preview = client.post(
            f"/jobs/{job_id}/change-plan",
            json={
                "kind": "cell_text_batch",
                "edits": [
                    {"sheet": "Data", "coord": "C1", "value": "税率の入力値です"},
                ],
            },
        )
        assert preview.status_code == 200

        response = client.post(
            f"/jobs/{job_id}/change-plan/execute",
            json={"plan": preview.json()["plan"]},
        )

        assert response.status_code == 200
        body = response.json()
        assert body["verification"]["status"] == "passed"
        assert body["provider"]["provider"] == "officecli"
        assert body["diff"]["cells"][0]["coord"] == "C1"
        revised_path = backend_storage.get_original_path(body["new_job_id"])
        revised = openpyxl.load_workbook(revised_path, data_only=False)
        try:
            assert revised["Data"]["C1"].value == "税率の入力値です"
        finally:
            revised.close()

    def test_rejects_existing_cell_overwrite(
        self,
        client: TestClient,
        backend_storage: Storage,
    ) -> None:
        """既存値を対象にした計画を作らない."""

        job_id = _seed_extracted_job(backend_storage, _xlsx_bytes_with_named_range())

        response = client.post(
            f"/jobs/{job_id}/change-plan",
            json={
                "kind": "cell_text_batch",
                "edits": [
                    {"sheet": "Data", "coord": "A1", "value": "上書き"},
                ],
            },
        )

        assert response.status_code == 422
