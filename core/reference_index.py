"""参照インデックス構築モジュール.

Workbook (数式 + VBA) を走査し、「あるセル/範囲を参照しているもの」を
逆引きできる ReferenceIndex を作る。

参照キーは正規化済み (シート修飾を補い、`$` を除去し、列文字を大文字化) で
格納されるため、`Calc!H2` / `$H$2` / `h2` などの表記揺れは同一キーになる。
範囲交差での検索は `find_overlapping()` を使う。

SPEC.md §4.3 参照。
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass

from core.models import Reference, ReferenceIndex, VbaModule, Workbook

logger = logging.getLogger(__name__)


# Excel の論理上の最大行・列. 全列指定 (A:A) を「行=1..MAX」として表現する目的で使う.
_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COL = 16_384


@dataclass(frozen=True)
class ParsedRef:
    """参照を構造化した形で保持する.

    シートが None の場合、`overlaps()` は wildcard 扱いになる
    (VBA の bare `Range("A1")` のようにシート不明の参照を表す)。
    """

    sheet: str | None
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    def overlaps(self, other: ParsedRef) -> bool:
        """別の ParsedRef と範囲が重なるか.

        シートが両方とも判明していて異なる場合は不一致。どちらかが
        None ならシート不問 (wildcard)。
        """
        if self.sheet is not None and other.sheet is not None and self.sheet != other.sheet:
            return False
        return not (
            self.max_row < other.min_row
            or self.min_row > other.max_row
            or self.max_col < other.min_col
            or self.min_col > other.max_col
        )


# シート修飾を切り出す: `Sheet1!A1` / `'My Sheet'!A1` / `''Quoted''!A1` の3形態
_SHEET_QUALIFIED_RE = re.compile(r"^(?:'((?:[^']|'')+)'|([^'!]+))!(.+)$")
# 全列指定 (例 `A:A`, `AA:AB`)
_FULL_COL_RE = re.compile(r"^([A-Z]+):([A-Z]+)$")
# 全行指定 (例 `1:5`)
_FULL_ROW_RE = re.compile(r"^(\d+):(\d+)$")


def _parse_ref(ref: str, owner_sheet: str | None = None) -> ParsedRef | None:
    """参照文字列を `ParsedRef` に正規化する.

    Args:
        ref: 参照文字列. 例: `A1`, `Calc!H2:H100`, `'My Sheet'!A1`,
            `$A$1:$B$2`, `A:A`, `1:1`.
        owner_sheet: シート修飾が ref に含まれていない場合に補うシート名。
            数式から取り出した ref ではその数式が属するシートを渡すことで、
            同一シート内の相対参照 (`A2`) も `Calc!A2` として正規化される。

    Returns:
        パースできれば `ParsedRef`. ファンクション名・テーブル名など範囲として
        解釈できない場合は None.
    """
    s = ref.strip()
    if not s:
        return None

    sheet = owner_sheet
    range_part = s
    m = _SHEET_QUALIFIED_RE.match(s)
    if m:
        raw_sheet = m.group(1) if m.group(1) is not None else m.group(2)
        # Excel は引用符内で `''` を `'` のエスケープにする
        sheet = raw_sheet.replace("''", "'").strip()
        range_part = m.group(3).strip()

    range_part = range_part.replace("$", "").upper()
    if not range_part:
        return None

    m_col = _FULL_COL_RE.match(range_part)
    if m_col:
        try:
            from openpyxl.utils.cell import column_index_from_string

            a = column_index_from_string(m_col.group(1))
            b = column_index_from_string(m_col.group(2))
        except Exception:  # noqa: BLE001
            return None
        return ParsedRef(sheet, 1, min(a, b), _EXCEL_MAX_ROW, max(a, b))

    m_row = _FULL_ROW_RE.match(range_part)
    if m_row:
        a_row = int(m_row.group(1))
        b_row = int(m_row.group(2))
        return ParsedRef(sheet, min(a_row, b_row), 1, max(a_row, b_row), _EXCEL_MAX_COL)

    try:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(range_part)
    except Exception:  # noqa: BLE001
        return None
    if min_col is None or min_row is None or max_col is None or max_row is None:
        return None
    return ParsedRef(
        sheet,
        int(min_row),
        int(min_col),
        int(max_row),
        int(max_col),
    )


def _canonical_key(parsed: ParsedRef) -> str:
    """`ParsedRef` から逆引きインデックスのキー文字列を生成する.

    フォーマット:
      - シートあり: `Sheet!RANGE`
      - シートなし (VBA bare): `RANGE`

    RANGE は次のルールで再構築する:
      - 全列: `A:A` / `A:Z`
      - 全行: `1:1` / `1:5`
      - 単一セル: `A1`
      - 通常範囲: `A1:B2`
    """
    from openpyxl.utils import get_column_letter

    full_col = parsed.min_row == 1 and parsed.max_row == _EXCEL_MAX_ROW
    full_row = parsed.min_col == 1 and parsed.max_col == _EXCEL_MAX_COL

    if full_col and not full_row:
        a = get_column_letter(parsed.min_col)
        b = get_column_letter(parsed.max_col)
        range_str = f"{a}:{a}" if a == b else f"{a}:{b}"
    elif full_row and not full_col:
        a_row = parsed.min_row
        b_row = parsed.max_row
        range_str = f"{a_row}:{a_row}" if a_row == b_row else f"{a_row}:{b_row}"
    else:
        a_col = get_column_letter(parsed.min_col)
        b_col = get_column_letter(parsed.max_col)
        if parsed.min_row == parsed.max_row and parsed.min_col == parsed.max_col:
            range_str = f"{a_col}{parsed.min_row}"
        else:
            range_str = f"{a_col}{parsed.min_row}:{b_col}{parsed.max_row}"

    if parsed.sheet:
        return f"{parsed.sheet}!{range_str}"
    return range_str


def _normalize_target(raw: str, owner_sheet: str | None = None) -> tuple[str, ParsedRef | None]:
    """raw 文字列を (canonical_key, parsed) に正規化する.

    パース不能なら (raw, None) をそのまま返す (テーブル名・関数名等)。
    """
    parsed = _parse_ref(raw, owner_sheet=owner_sheet)
    if parsed is None:
        return raw, None
    return _canonical_key(parsed), parsed


# --- VBA 参照抽出 -------------------------------------------------------------
# VBA は実行時に参照先を組み立てられるため、静的解析だけで完全な依存関係は作れない。
# ここでは正規表現で文字列断片を拾うのではなく、コメント/文字列/括弧を理解する
# 軽いスキャナで「静的に確定できる参照」だけを登録する。


@dataclass(frozen=True)
class _VbaRefHit:
    """VBA コード内で見つかった静的参照."""

    sheet: str | None
    ref: str
    code: str


def _col_num_to_letters(col: int) -> str:
    """1始まりの列番号をExcel列文字に変換 (1 -> A, 27 -> AA)."""
    if col < 1:
        return ""
    s = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        s = chr(ord("A") + rem) + s
    return s


def _qualify(sheet: str | None, ref: str) -> str:
    """シート名と範囲を結合して `Sheet!Range` 形式に. シートが None なら範囲のみ."""
    if sheet:
        return f"{sheet}!{ref}"
    return ref


def _find_enclosing_procedure(module: VbaModule, line_num: int) -> str | None:
    """指定行を含むプロシージャ名を返す. どのプロシージャにも属さなければ None."""
    for p in module.procedures:
        if p.start_line <= line_num <= p.end_line:
            return p.name
    return None


def _from_label(module: VbaModule, line_num: int) -> str:
    """VBA参照の from_ ラベルを `Module1.UpdateDaily:L47` の形に整える."""
    proc = _find_enclosing_procedure(module, line_num)
    if proc:
        return f"{module.name}.{proc}:L{line_num}"
    return f"{module.name}:L{line_num}"


def _strip_vba_comment(line: str) -> str:
    """VBA のコメント部分を除いたコード行を返す.

    Args:
        line: VBA ソース 1 行.

    Returns:
        文字列リテラル内の `'` をコメントと誤認せずに、実コード部分だけを返す。
    """
    in_string = False
    i = 0
    while i < len(line):
        ch = line[i]
        if ch == '"':
            if in_string and i + 1 < len(line) and line[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
        elif ch == "'" and not in_string:
            return line[:i]
        i += 1
    return line


def _split_vba_statements(line: str) -> list[str]:
    """1 行を VBA の `:` 区切りステートメントに分割する.

    Args:
        line: コメント除去済みの VBA ソース 1 行.

    Returns:
        文字列リテラル内の `:` は区切りと見なさないステートメント一覧。
    """
    statements: list[str] = []
    in_string = False
    start = 0
    i = 0
    while i < len(line):
        ch = line[i]
        if ch == '"':
            if in_string and i + 1 < len(line) and line[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
        elif ch == ":" and not in_string:
            statements.append(line[start:i])
            start = i + 1
        i += 1
    statements.append(line[start:])
    return statements


def _find_top_level_equals(statement: str) -> int | None:
    """トップレベルの `=` 位置を返す.

    Args:
        statement: VBA ステートメント.

    Returns:
        文字列・括弧の内側を除いた `=` の位置。なければ None。
    """
    in_string = False
    depth = 0
    i = 0
    while i < len(statement):
        ch = statement[i]
        if ch == '"':
            if in_string and i + 1 < len(statement) and statement[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
        elif not in_string:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(0, depth - 1)
            elif ch == "=" and depth == 0:
                return i
        i += 1
    return None


def _iter_vba_identifiers(code: str) -> Iterable[tuple[str, int, int]]:
    """コード上の識別子トークンを文字列リテラル外から列挙する.

    Args:
        code: VBA コード片.

    Yields:
        `(識別子, 開始位置, 終了位置)`。
    """
    in_string = False
    i = 0
    while i < len(code):
        ch = code[i]
        if ch == '"':
            if in_string and i + 1 < len(code) and code[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
            i += 1
            continue
        if in_string:
            i += 1
            continue
        if ch.isalpha() or ch == "_":
            start = i
            i += 1
            while i < len(code) and (code[i].isalnum() or code[i] == "_"):
                i += 1
            yield code[start:i], start, i
            continue
        i += 1


def _read_call_arguments(code: str, ident_end: int) -> tuple[str, int] | None:
    """識別子直後の呼び出し括弧から引数文字列を読む.

    Args:
        code: VBA コード片.
        ident_end: 関数/メソッド名識別子の終了位置.

    Returns:
        `(括弧内の文字列, 呼び出し全体の終了位置)`。呼び出しでなければ None。
    """
    i = ident_end
    while i < len(code) and code[i].isspace():
        i += 1
    if i >= len(code) or code[i] != "(":
        return None

    start = i + 1
    depth = 1
    in_string = False
    i += 1
    while i < len(code):
        ch = code[i]
        if ch == '"':
            if in_string and i + 1 < len(code) and code[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
        elif not in_string:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    return code[start:i], i + 1
        i += 1
    return None


def _literal_string_arg(arg_text: str) -> str | None:
    """引数が単一の文字列リテラルなら値を返す.

    Args:
        arg_text: 呼び出し括弧内の引数文字列.

    Returns:
        `"A1"` のように静的に確定できる単一文字列。`"A" & row` は None。
    """
    s = arg_text.strip()
    if not s.startswith('"'):
        return None

    chars: list[str] = []
    i = 1
    while i < len(s):
        ch = s[i]
        if ch == '"':
            if i + 1 < len(s) and s[i + 1] == '"':
                chars.append('"')
                i += 2
                continue
            tail = s[i + 1 :].strip()
            return "".join(chars) if tail == "" else None
        chars.append(ch)
        i += 1
    return None


def _split_top_level_args(arg_text: str) -> list[str]:
    """呼び出し引数をトップレベルのカンマで分割する.

    Args:
        arg_text: 呼び出し括弧内の引数文字列.

    Returns:
        文字列・括弧の内側を分割しない引数一覧。
    """
    args: list[str] = []
    in_string = False
    depth = 0
    start = 0
    i = 0
    while i < len(arg_text):
        ch = arg_text[i]
        if ch == '"':
            if in_string and i + 1 < len(arg_text) and arg_text[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
        elif not in_string:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth = max(0, depth - 1)
            elif ch == "," and depth == 0:
                args.append(arg_text[start:i].strip())
                start = i + 1
        i += 1
    args.append(arg_text[start:].strip())
    return args


def _sheet_from_expression(
    expr: str,
    sheet_vars: dict[str, str],
    *,
    require_terminal: bool = False,
) -> str | None:
    """VBA 式から静的に分かるシート名を解決する.

    Args:
        expr: `Worksheets("Calc")` や `ws` などの式.
        sheet_vars: `Set ws = Worksheets("Calc")` で分かった変数表.
        require_terminal: True の場合、式全体がシートオブジェクトで終わる時だけ
            解決する。`Worksheets("Calc").Range("A1")` は除外する。

    Returns:
        シート名。インデックス指定や動的式など確定できない場合は None。
    """
    s = expr.strip()
    if not s:
        return None
    var_name = s.lower()
    if var_name in sheet_vars:
        return sheet_vars[var_name]

    found: str | None = None
    for ident, _start, end in _iter_vba_identifiers(s):
        if ident.lower() not in {"worksheets", "sheets"}:
            continue
        call = _read_call_arguments(s, end)
        if call is None:
            continue
        if require_terminal and s[call[1] :].strip():
            continue
        value = _literal_string_arg(call[0])
        if value:
            found = value
    return found


def _update_sheet_variable(statement: str, sheet_vars: dict[str, str]) -> None:
    """`Set ws = Worksheets("X")` 形式の単純なシート変数を追跡する.

    Args:
        statement: VBA ステートメント.
        sheet_vars: 更新対象の変数表。
    """
    stripped = statement.strip()
    lower = stripped.lower()
    if not lower.startswith("set "):
        return

    eq_pos = _find_top_level_equals(stripped)
    if eq_pos is None:
        return

    lhs = stripped[4:eq_pos].strip()
    rhs = stripped[eq_pos + 1 :].strip()
    if not lhs or not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", lhs):
        return
    sheet = _sheet_from_expression(rhs, sheet_vars, require_terminal=True)
    key = lhs.lower()
    if sheet is None:
        sheet_vars.pop(key, None)
    else:
        sheet_vars[key] = sheet


def _call_owner_sheet(
    statement: str,
    ident_start: int,
    sheet_vars: dict[str, str],
    with_stack: list[str | None],
    owner_sheet: str | None,
) -> str | None:
    """`Range` / `Cells` 呼び出しの所有シートを解決する.

    Args:
        statement: VBA ステートメント.
        ident_start: `Range` / `Cells` 識別子の開始位置.
        sheet_vars: 単純なシート変数表.
        with_stack: `With Worksheets("X")` のネスト.
        owner_sheet: シートモジュールなどから推定できる所有シート.

    Returns:
        静的に分かるシート名。不明なら None。
    """
    i = ident_start - 1
    while i >= 0 and statement[i].isspace():
        i -= 1
    if i < 0 or statement[i] != ".":
        return owner_sheet

    expr = statement[:i].strip()
    if not expr:
        return with_stack[-1] if with_stack else owner_sheet

    sheet = _sheet_from_expression(expr, sheet_vars)
    if sheet is not None:
        return sheet

    # `foo(ws.Range("A1"))` のように前置式が長い場合は末尾の識別子も見る。
    matches = list(_iter_vba_identifiers(expr))
    if matches:
        last_ident = matches[-1][0].lower()
        if last_ident in sheet_vars:
            return sheet_vars[last_ident]
    return None


def _iter_bracket_refs(statement: str, owner_sheet: str | None) -> Iterable[_VbaRefHit]:
    """`[Sheet!A1]` / `[A1]` の短縮参照を列挙する.

    Args:
        statement: VBA ステートメント.
        owner_sheet: シートモジュールなどから推定できる所有シート.

    Yields:
        静的参照ヒット。
    """
    in_string = False
    i = 0
    while i < len(statement):
        ch = statement[i]
        if ch == '"':
            if in_string and i + 1 < len(statement) and statement[i + 1] == '"':
                i += 2
                continue
            in_string = not in_string
        elif ch == "[" and not in_string:
            end = statement.find("]", i + 1)
            if end == -1:
                return
            content = statement[i + 1 : end].strip()
            if content:
                parsed = _parse_ref(content, owner_sheet=owner_sheet)
                if parsed is not None:
                    yield _VbaRefHit(
                        parsed.sheet,
                        _canonical_key(parsed).split("!", 1)[-1],
                        statement[i : end + 1],
                    )
            i = end
        i += 1


def _extract_refs_from_statement(
    statement: str,
    sheet_vars: dict[str, str],
    with_stack: list[str | None],
    owner_sheet: str | None,
) -> list[_VbaRefHit]:
    """1 ステートメントから静的な Range/Cells 参照を抽出する.

    Args:
        statement: VBA ステートメント.
        sheet_vars: 単純なシート変数表.
        with_stack: `With` 文で解決済みのシート名スタック.
        owner_sheet: シートモジュールなどから推定できる所有シート.

    Returns:
        見つかった静的参照。動的に組み立てる参照は返さない。
    """
    refs = list(_iter_bracket_refs(statement, owner_sheet))

    for ident, start, end in _iter_vba_identifiers(statement):
        ident_lower = ident.lower()
        if ident_lower not in {"range", "cells"}:
            continue
        call = _read_call_arguments(statement, end)
        if call is None:
            continue
        args_text, call_end = call
        sheet = _call_owner_sheet(statement, start, sheet_vars, with_stack, owner_sheet)
        code = statement[start:call_end].strip()

        if ident_lower == "range":
            ref = _literal_string_arg(args_text)
            if ref:
                refs.append(_VbaRefHit(sheet, ref, code))
            continue

        args = _split_top_level_args(args_text)
        if len(args) < 2:
            continue
        if not args[0].isdigit() or not args[1].isdigit():
            continue
        row = int(args[0])
        col = int(args[1])
        cell_ref = f"{_col_num_to_letters(col)}{row}"
        if cell_ref:
            refs.append(_VbaRefHit(sheet, cell_ref, code))

    return refs


def _module_owner_sheet(module: VbaModule, sheet_names: set[str]) -> str | None:
    """モジュール名からシート所有者を推定する.

    Args:
        module: VBA モジュール.
        sheet_names: ブック内のワークシート名集合.

    Returns:
        Document モジュール名がシート名と一致する場合のみ、そのシート名。
    """
    if module.type == "Document" and module.name in sheet_names:
        return module.name
    return None


def _scan_vba_module(module: VbaModule, sheet_names: set[str] | None = None) -> list[Reference]:
    """1モジュールの code を走査して Reference を抽出する.

    Args:
        module: 走査対象の VBA モジュール.
        sheet_names: ブック内シート名。Document モジュールの bare 参照を
            シートに補正できる場合に使う。

    Returns:
        静的に確定できるセル/範囲参照。
    """
    refs: list[Reference] = []
    if not module.code:
        return refs

    owner_sheet = _module_owner_sheet(module, sheet_names or set())
    sheet_vars: dict[str, str] = {}
    with_stack: list[str | None] = []

    lines = module.code.splitlines()
    for idx, raw_line in enumerate(lines, start=1):
        for statement in _split_vba_statements(_strip_vba_comment(raw_line)):
            stripped = statement.strip()
            lower = stripped.lower()
            if not stripped:
                continue

            if lower == "end with":
                if with_stack:
                    with_stack.pop()
                continue

            if lower.startswith("with "):
                with_stack.append(
                    _sheet_from_expression(stripped[5:], sheet_vars, require_terminal=True)
                )
                continue

            _update_sheet_variable(stripped, sheet_vars)
            for hit in _extract_refs_from_statement(stripped, sheet_vars, with_stack, owner_sheet):
                refs.append(
                    Reference(
                        kind="vba",
                        from_=_from_label(module, idx),
                        to=_qualify(hit.sheet, hit.ref),
                        code=hit.code,
                    )
                )

    return refs


def build_reference_index(wb: Workbook) -> ReferenceIndex:
    """Workbook 全体から参照の逆引きインデックスを構築する.

    キーは canonical 形式 (シート修飾付き / `$` 除去 / 列大文字) に正規化される。
    パース不能な参照 (関数名・テーブル名等) は raw 文字列のままキーになる。

    Args:
        wb: Workbook モデル. sheets/formulas と vba_modules/code が埋まっている前提.

    Returns:
        ReferenceIndex: refs[正規化キー] -> [Reference, ...]
    """
    bucket: dict[str, list[Reference]] = defaultdict(list)

    # 数式側: 各 CellFormula.refs を逆引きする. 同一シート内の相対参照 (例 `A2`)
    # にはオーナーシート名を補って `Calc!A2` に揃える.
    for sheet in wb.sheets:
        for f in sheet.formulas:
            for raw_target in f.refs:
                key, _ = _normalize_target(raw_target, owner_sheet=sheet.name)
                bucket[key].append(
                    Reference(
                        kind="formula",
                        from_=f.coord,
                        to=key,
                        code=f.formula,
                    )
                )
        for chart in sheet.charts:
            chart_name = chart.title or chart.name or chart.chart_type or "chart"
            for idx, series in enumerate(chart.series, start=1):
                refs = [series.values_ref, series.categories_ref]
                for raw_target in refs:
                    if not raw_target:
                        continue
                    key, _ = _normalize_target(raw_target, owner_sheet=sheet.name)
                    bucket[key].append(
                        Reference(
                            kind="chart",
                            from_=f"{sheet.name}!Chart:{chart_name}:S{idx}",
                            to=key,
                            code=series.name or chart.chart_type,
                        )
                    )
        for pivot in sheet.pivot_tables:
            raw_source = ""
            if pivot.source_sheet and pivot.source_ref:
                raw_source = f"{pivot.source_sheet}!{pivot.source_ref}"
            elif pivot.source_ref:
                raw_source = pivot.source_ref
            elif pivot.source_name:
                raw_source = pivot.source_name
            if not raw_source:
                continue
            key, _ = _normalize_target(raw_source, owner_sheet=sheet.name)
            bucket[key].append(
                Reference(
                    kind="pivot",
                    from_=f"{sheet.name}!Pivot:{pivot.name}",
                    to=key,
                    code=", ".join(pivot.value_fields) or pivot.name,
                )
            )

    # VBA側: code を軽量スキャナで走査. _scan_vba_module が返した raw `to` を
    # canonical キーに置換して登録する。静的に確定できない動的参照は登録しない。
    sheet_names = {sheet.name for sheet in wb.sheets}
    for module in wb.vba_modules:
        for ref in _scan_vba_module(module, sheet_names=sheet_names):
            key, _ = _normalize_target(ref.to, owner_sheet=None)
            bucket[key].append(
                Reference(
                    kind=ref.kind,
                    from_=ref.from_,
                    to=key,
                    code=ref.code,
                )
            )

    return ReferenceIndex(refs=dict(bucket))


def find_overlapping(
    index: ReferenceIndex,
    target: str,
    owner_sheet: str | None = None,
) -> list[Reference]:
    """target と範囲が重なる参照を全件返す.

    完全一致だけでなく、範囲交差で検出する。例:
      - index に `Input!A:A` があり target=`Input!A5` → ヒット
      - index に `Calc!A1:J100` があり target=`Calc!H2` → ヒット
      - index に sheet 不明の VBA bare 参照 (`A1:J100`) があれば、
        target のシートに関係なく範囲が重なればヒット (wildcard 扱い)

    パース不能なキーは厳密文字列一致でのみマッチする。

    Args:
        index: `build_reference_index` の結果.
        target: 検索したい参照先 (例: `Calc!H2`, `Input!A5`).
        owner_sheet: target がシート修飾なしの場合に補うシート名 (任意).

    Returns:
        重なる Reference を index 登録順で返す. 重複は除去しない
        (1 つの数式が同じセルを 2 回参照していれば 2 件返る)。
    """
    target_key, target_parsed = _normalize_target(target, owner_sheet=owner_sheet)

    # target がパース不能: 完全一致のみ
    if target_parsed is None:
        return list(index.refs.get(target_key, []))

    results: list[Reference] = []
    seen_keys: set[str] = set()

    # 1) 完全一致 (高速パス)
    exact = index.refs.get(target_key)
    if exact:
        results.extend(exact)
        seen_keys.add(target_key)

    # 2) 範囲交差
    for key, refs in index.refs.items():
        if key in seen_keys:
            continue
        parsed = _parse_ref(key, owner_sheet=None)
        if parsed is None:
            continue
        if parsed.overlaps(target_parsed):
            results.extend(refs)

    return results
