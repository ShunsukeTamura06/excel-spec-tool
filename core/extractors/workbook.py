"""ワークブック構造抽出モジュール.

openpyxl で .xlsx / .xlsm を開き、数式セル・名前付き範囲・条件付き書式・外部リンクを
Pydantic モデルに詰める。VBA は本モジュールでは扱わない (core.extractors.vba を別途呼ぶ).

docs/SPEC.ja.md §4.2 参照。
"""

from __future__ import annotations

import logging
import posixpath
import re
import zipfile
from collections.abc import Callable
from pathlib import Path
from typing import Literal
from xml.etree import ElementTree as ET

from core.exceptions import ExtractionError
from core.external_functions import detect_in_formula
from core.models import (
    CellFormula,
    ChartObject,
    ChartSeries,
    ConditionalFormat,
    DataValidation,
    ExcelTable,
    FormControl,
    NamedRange,
    PivotTableInfo,
    PowerQueryInfo,
    SheetInfo,
    Workbook,
)

# プレビュー範囲（先頭 N 行 × M 列）. 解釈は加えず literal に出力するだけ.
PREVIEW_MAX_ROWS = 20
PREVIEW_MAX_COLS = 20

logger = logging.getLogger(__name__)


# `Sheet1!$A$1` や `'My Sheet'!A1:B10` の先頭シート名部分を取り出す
_SHEET_PREFIX_RE = re.compile(r"^(?:'([^']+)'|([^!'\s]+))!")


def _extract_sheet_name(refers_to: str) -> str | None:
    """`Sheet1!$A$1` 形式の参照から先頭のシート名を抜き出す.

    ヒットしない (シート名なし) 場合は None.
    """
    m = _SHEET_PREFIX_RE.match(refers_to.strip())
    if not m:
        return None
    return m.group(1) or m.group(2)


def _extract_formula_refs(formula: str) -> list[str]:
    """数式から RANGE トークンを抽出する.

    openpyxl の Tokenizer に頼る。`=SUMIF(Input!A:A, A2, Input!E:E)` から
    `["Input!A:A", "A2", "Input!E:E"]` を得る想定。
    """
    from openpyxl.formula.tokenizer import Tokenizer

    if not formula:
        return []
    if not formula.startswith("="):
        formula = "=" + formula

    try:
        tok = Tokenizer(formula)
    except Exception:  # noqa: BLE001 - 構文エラーは握りつぶしてログ
        logger.debug("Failed to tokenize formula: %r", formula)
        return []

    refs: list[str] = []
    for t in tok.items:
        if getattr(t, "subtype", None) == "RANGE":
            refs.append(t.value)
    return refs


def _extract_formulas(ws: object) -> list[CellFormula]:
    """1シートから数式セルを抽出する."""
    formulas: list[CellFormula] = []
    sheet_title: str = ws.title  # type: ignore[attr-defined]

    for row in ws.iter_rows():  # type: ignore[attr-defined]
        for cell in row:
            if cell.data_type != "f":
                continue
            value = cell.value
            if value is None:
                continue
            formula_str = str(value)
            coord = f"{sheet_title}!{cell.coordinate}"
            refs = _extract_formula_refs(formula_str)
            external_fns = detect_in_formula(formula_str)
            formulas.append(
                CellFormula(
                    coord=coord,
                    formula=formula_str,
                    refs=refs,
                    external_functions=external_fns,
                )
            )
    return formulas


def _extract_conditional_formats(ws: object) -> list[ConditionalFormat]:
    """1シートから条件付き書式を抽出する."""
    cfs: list[ConditionalFormat] = []
    cf_list = getattr(ws, "conditional_formatting", None)
    if cf_list is None:
        return cfs
    try:
        for rng, rules in cf_list._cf_rules.items():
            range_str = str(rng.sqref) if hasattr(rng, "sqref") else str(rng)
            for rule in rules:
                rule_str = _summarize_cf_rule(rule)
                cfs.append(ConditionalFormat(range=range_str, rule=rule_str))
    except Exception:  # noqa: BLE001
        logger.debug(
            "Failed to extract conditional formats from %s",
            sheet_title := getattr(ws, "title", "?"),
        )
        _ = sheet_title
    return cfs


def _summarize_cf_rule(rule: object) -> str:
    """ConditionalFormatRule を人間が読める短い文字列に."""
    rtype = getattr(rule, "type", None) or "rule"
    operator = getattr(rule, "operator", None)
    formula = getattr(rule, "formula", None)
    parts = [str(rtype)]
    if operator:
        parts.append(str(operator))
    if formula:
        formulas_list = list(formula) if not isinstance(formula, str) else [formula]
        parts.append(",".join(str(f) for f in formulas_list))
    return " ".join(parts)


def _extract_external_links(wb: object) -> list[str]:
    """外部リンクを抽出する.

    `wb._external_links` は private API なので失敗を許容する。
    """
    links: list[str] = []
    raw = getattr(wb, "_external_links", None)
    if not raw:
        return links
    try:
        for link in raw:
            file_link = getattr(link, "file_link", None)
            target = getattr(file_link, "Target", None) if file_link else None
            if target:
                links.append(str(target))
    except Exception:  # noqa: BLE001
        logger.debug("Failed to enumerate external links")
    return links


def _attach_named_ranges(wb: object, sheets_by_name: dict[str, SheetInfo]) -> None:
    """ワークブック定義名を、参照先シートの SheetInfo.named_ranges に振り分ける.

    シートが特定できない (refers_to が解析不能 or 該当シート無し) 場合は
    最初のシートに付ける。
    """
    defined_names = getattr(wb, "defined_names", None)
    if not defined_names:
        return

    try:
        items = list(defined_names.items())
    except Exception:  # noqa: BLE001
        logger.debug("Failed to enumerate defined_names")
        return

    for name, defn in items:
        refers_to = getattr(defn, "value", None) or ""
        sheet_name = _extract_sheet_name(refers_to)
        target_sheet: SheetInfo | None = None
        if sheet_name and sheet_name in sheets_by_name:
            target_sheet = sheets_by_name[sheet_name]
        elif sheets_by_name:
            target_sheet = next(iter(sheets_by_name.values()))
        if target_sheet is not None:
            target_sheet.named_ranges.append(NamedRange(name=name, refers_to=refers_to))


def _extract_excel_tables(ws: object) -> list[ExcelTable]:
    """ws.tables から Excel テーブル (ListObject) を抽出する.

    `ws.tables` は openpyxl が解釈した確定情報なのでヒューリスティック不要。
    """
    result: list[ExcelTable] = []
    tables_attr = getattr(ws, "tables", None)
    if tables_attr is None:
        return result
    try:
        # openpyxl 3.1 の TableList:
        # - keys() / __iter__ で table 名一覧
        # - items() は (name, ref_str) を返す (Table オブジェクト本体ではない)
        # - get(name) で Table オブジェクト本体を取れる
        names: list[str]
        if hasattr(tables_attr, "keys"):
            names = list(tables_attr.keys())
        else:
            names = [getattr(t, "name", "") for t in tables_attr]
        for name in names:
            table = tables_attr.get(name) if hasattr(tables_attr, "get") else None
            if table is None:
                continue
            display_name = (
                getattr(table, "displayName", None) or getattr(table, "name", None) or name
            )
            ref = getattr(table, "ref", "") or ""
            header_row_count = getattr(table, "headerRowCount", 1) or 1
            if not display_name or not ref:
                continue
            result.append(
                ExcelTable(
                    name=str(display_name),
                    ref=str(ref),
                    header_row_count=int(header_row_count),
                )
            )
    except Exception:  # noqa: BLE001
        logger.debug("Failed to enumerate tables on sheet")
    return result


def _extract_data_validations(ws: object) -> list[DataValidation]:
    """openpyxl の data_validations を抽出する.

    各 DataValidation は適用範囲 (sqref) ごとに 1 行 (sqref に複数範囲が
    含まれていれば、それを space 区切りでまとめて格納). type / formula1 /
    prompt / error message も拾う.
    """
    out: list[DataValidation] = []
    dv_attr = getattr(ws, "data_validations", None)
    if dv_attr is None:
        return out
    try:
        items = list(getattr(dv_attr, "dataValidation", []) or [])
    except Exception:  # noqa: BLE001
        logger.debug("Failed to enumerate data_validations on sheet")
        return out

    for dv in items:
        try:
            sqref = getattr(dv, "sqref", None)
            range_str = str(sqref) if sqref is not None else ""
            dv_type = str(getattr(dv, "type", "") or "")
            formula1 = getattr(dv, "formula1", "") or ""
            operator = str(getattr(dv, "operator", "") or "")
            prompt = str(getattr(dv, "prompt", "") or "")
            error = str(getattr(dv, "error", "") or "")
            allow_blank = bool(getattr(dv, "allowBlank", True))
            if not range_str or not dv_type:
                continue
            out.append(
                DataValidation(
                    range=range_str,
                    type=dv_type,
                    formula=str(formula1),
                    operator=operator,
                    prompt=prompt,
                    error=error,
                    allow_blank=allow_blank,
                )
            )
        except Exception:  # noqa: BLE001
            logger.debug("Skipping malformed data_validation entry")
            continue
    return out


# ----- フォームコントロール (VML-based) ----------------------------------

# VML 名前空間
_VML_NS = {
    "v": "urn:schemas-microsoft-com:vml",
    "x": "urn:schemas-microsoft-com:office:excel",
    "o": "urn:schemas-microsoft-com:office:office",
}
# rels 名前空間
_REL_NS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}


def _resolve_package_part(base_dir: str, target: str) -> str:
    """Open XML パッケージ内の Relationship Target を正規化する.

    Args:
        base_dir: relationship 所有パーツのディレクトリ。例: ``xl/worksheets``。
        target: rels に書かれた Target。相対パスまたは ``/xl/...`` 形式。

    Returns:
        zip 内で読める、先頭スラッシュなしの正規化パス。
    """
    if not target:
        return ""
    if target.startswith("/"):
        return posixpath.normpath(target.lstrip("/"))
    return posixpath.normpath(posixpath.join(base_dir, target))


def _local_name(tag: str) -> str:
    """XMLタグ名から名前空間を除いたローカル名を返す.

    Args:
        tag: ElementTree のタグ名。

    Returns:
        ローカル名。
    """
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _relationship_id(el: ET.Element) -> str:
    """officeDocument relationship id を返す.

    Args:
        el: XML 要素。

    Returns:
        ``r:id`` 相当の属性値。なければ空文字。
    """
    return el.attrib.get(
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", ""
    )


def _xlsm_sheet_paths(zf: zipfile.ZipFile) -> dict[str, str]:
    """xlsm zip から「シート名 → worksheet パーツパス」を返す.

    Args:
        zf: xlsx/xlsm/xltm/xlsb を開いた ZipFile。

    Returns:
        シート名をキー、``xl/worksheets/sheet*.xml`` を値にしたマップ。
    """
    name_to_sheet_path: dict[str, str] = {}
    try:
        wb_xml = zf.read("xl/workbook.xml")
        wb_rels = zf.read("xl/_rels/workbook.xml.rels")
    except KeyError:
        return name_to_sheet_path

    # r:id → sheet file
    rid_to_target: dict[str, str] = {}
    try:
        root = ET.fromstring(wb_rels)
        for rel in root.iter(
            "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
        ):
            rid_to_target[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")
    except ET.ParseError:
        return name_to_sheet_path

    # sheet name → r:id → sheet file path
    try:
        root = ET.fromstring(wb_xml)
        for sheet in root.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
            name = sheet.attrib.get("name", "")
            rid = _relationship_id(sheet)
            target = rid_to_target.get(rid, "")
            if not name or not target:
                continue
            # workbook.xml.rels の Target は通常 "worksheets/sheet1.xml"。
            sheet_path = _resolve_package_part("xl", target)
            name_to_sheet_path[name] = sheet_path
    except ET.ParseError:
        return name_to_sheet_path
    return name_to_sheet_path


def _sheet_related_parts(
    zf: zipfile.ZipFile,
    predicate: Callable[[ET.Element], bool],
) -> dict[str, list[str]]:
    """各シートの relationship から条件に合うパーツ一覧を返す.

    Args:
        zf: xlsx/xlsm/xltm/xlsb を開いた ZipFile。
        predicate: Relationship 要素を受け取り、対象なら True を返す関数。

    Returns:
        シート名をキー、関連パーツ一覧を値にしたマップ。
    """
    sheet_to_parts: dict[str, list[str]] = {}
    name_to_sheet_path = _xlsm_sheet_paths(zf)

    for name, sheet_path in name_to_sheet_path.items():
        # 例: xl/worksheets/sheet1.xml → xl/worksheets/_rels/sheet1.xml.rels
        if "/" in sheet_path:
            dir_part, file_part = sheet_path.rsplit("/", 1)
        else:
            dir_part, file_part = "", sheet_path
        rels_path = f"{dir_part}/_rels/{file_part}.rels"
        try:
            rels_xml = zf.read(rels_path)
        except KeyError:
            continue
        try:
            rels_root = ET.fromstring(rels_xml)
        except ET.ParseError:
            continue
        for rel in rels_root.iter(
            "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
        ):
            if not predicate(rel):
                continue
            target = rel.attrib.get("Target", "")
            part_path = _resolve_package_part(dir_part, target)
            sheet_to_parts.setdefault(name, []).append(part_path)

    return sheet_to_parts


def _xlsm_sheet_to_vml_map(zf: zipfile.ZipFile) -> dict[str, list[str]]:
    """xlsm zip から「シート名 → vmlDrawing*.vml のフルパス一覧」を返す.

    Args:
        zf: xlsm/xltm/xlsb を開いた ZipFile。

    Returns:
        シート名をキー、関連する VML パーツ一覧を値にしたマップ。
        rels の解決にコケた場合は、取れた範囲だけ返す。
    """

    def _is_vml(rel: ET.Element) -> bool:
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        return target.lower().endswith(".vml") and (
            "vmldrawing" in target.lower() or rel_type.endswith("/vmlDrawing")
        )

    return _sheet_related_parts(zf, _is_vml)


def _xlsm_sheet_to_drawing_map(zf: zipfile.ZipFile) -> dict[str, list[str]]:
    """xlsx/xlsm zip から「シート名 → drawing*.xml のフルパス一覧」を返す.

    Args:
        zf: xlsx/xlsm/xltm/xlsb を開いた ZipFile。

    Returns:
        シート名をキー、関連する DrawingML パーツ一覧を値にしたマップ。
    """

    def _is_drawing(rel: ET.Element) -> bool:
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        return target.lower().endswith(".xml") and (
            "drawings/drawing" in target.lower() or rel_type.endswith("/drawing")
        )

    return _sheet_related_parts(zf, _is_drawing)


def _normalize_vml_xml(vml_bytes: bytes) -> bytes:
    """Excel が出す HTML 風の VML を XML として読める形へ補正する.

    Args:
        vml_bytes: ``xl/drawings/vmlDrawing*.vml`` の内容。

    Returns:
        ElementTree で読めるように軽く補正した VML バイト列。
    """

    def _close_br(match: re.Match[bytes]) -> bytes:
        attrs = match.group("attrs") or b""
        if attrs.strip().endswith(b"/"):
            return match.group(0)
        return b"<br" + attrs + b" />"

    return re.sub(rb"<br(?P<attrs>(?=[\s/>])[^<>]*)>", _close_br, vml_bytes, flags=re.I)


def _element_text_with_breaks(el: ET.Element) -> str:
    """XML 要素配下のテキストを、br を空白として抽出する.

    Args:
        el: テキスト抽出対象の XML 要素。

    Returns:
        結合したテキスト。
    """
    parts: list[str] = []
    if el.text:
        parts.append(el.text)
    for child in el:
        if _local_name(child.tag).lower() == "br":
            parts.append(" ")
        parts.append(_element_text_with_breaks(child))
        if child.tail:
            parts.append(child.tail)
    return "".join(parts)


def _parse_vml_form_controls(vml_bytes: bytes) -> list[FormControl]:
    """VML 1 ファイル分から FormControl を抽出する.

    各 v:shape の中の x:ClientData (type 属性) と x:FmlaMacro を見る:
      - <x:ClientData ObjectType="Button|Checkbox|Drop|Spin|...">
      - <x:FmlaMacro>マクロ名</x:FmlaMacro>
      - <v:textbox> 配下のテキスト = ボタン表面文字
      - <x:Anchor>FromCol,FromColOff,FromRow,FromRowOff,ToCol,...</x:Anchor>
    """
    out: list[FormControl] = []
    try:
        root = ET.fromstring(_normalize_vml_xml(vml_bytes))
    except ET.ParseError:
        return out

    # shape 要素は v: 名前空間配下. ET は名前空間付きで {ns}local の形でアクセス.
    v_shape_tag = f"{{{_VML_NS['v']}}}shape"
    x_cd_tag = f"{{{_VML_NS['x']}}}ClientData"
    x_fmla_tag = f"{{{_VML_NS['x']}}}FmlaMacro"
    x_anchor_tag = f"{{{_VML_NS['x']}}}Anchor"
    v_textbox_tag = f"{{{_VML_NS['v']}}}textbox"

    object_type_map = {
        "Button": "button",
        "Checkbox": "checkbox",
        "Radio": "radio",
        "Drop": "dropdown",
        "List": "listbox",
        "Spin": "spinner",
        "Scroll": "scrollbar",
        "GBox": "groupbox",
        "Label": "label",
    }

    for shape in root.iter(v_shape_tag):
        name = shape.attrib.get(f"{{{_VML_NS['o']}}}spid", "") or shape.attrib.get("id", "")
        cd = shape.find(x_cd_tag)
        if cd is None:
            continue
        obj_type = cd.attrib.get("ObjectType", "")
        kind = object_type_map.get(obj_type, obj_type.lower() if obj_type else "control")

        fmla = cd.find(x_fmla_tag)
        macro = (fmla.text or "").strip() if fmla is not None and fmla.text else ""

        anchor_el = cd.find(x_anchor_tag)
        anchor_text = ""
        if anchor_el is not None and anchor_el.text:
            anchor_text = _anchor_to_cell(anchor_el.text)

        # ボタンの表面テキストを v:textbox から拾う (best-effort)
        text = ""
        textbox = shape.find(v_textbox_tag)
        if textbox is not None:
            # textbox 配下の全テキストを連結
            text = _element_text_with_breaks(textbox).strip()
            # 改行・余分な空白を 1 つに圧縮
            text = re.sub(r"\s+", " ", text)[:120]

        # マクロも表示テキストも空のコントロールは情報量が薄いのでスキップ
        if not macro and not text:
            continue

        out.append(
            FormControl(
                kind=kind or "control",
                name=name,
                text=text,
                macro=macro,
                anchor=anchor_text,
            )
        )
    return out


def _anchor_to_cell(anchor_text: str) -> str:
    """`x:Anchor` の `FromCol,FromColOff,FromRow,FromRowOff,...` を "A1" 形式へ.

    パース失敗時は空文字を返す.
    """
    try:
        from openpyxl.utils import get_column_letter

        parts = [p.strip() for p in anchor_text.split(",")]
        if len(parts) < 4:
            return ""
        from_col = int(parts[0])
        from_row = int(parts[2])
        return f"{get_column_letter(from_col + 1)}{from_row + 1}"
    except (ValueError, IndexError):
        return ""


def _drawing_anchor_to_cell(anchor: ET.Element) -> str:
    """DrawingML anchor 要素から左上セルを A1 形式で返す.

    Args:
        anchor: ``xdr:twoCellAnchor`` または ``xdr:oneCellAnchor`` 要素。

    Returns:
        左上セル。取得できない場合は空文字。
    """
    try:
        from openpyxl.utils import get_column_letter

        from_el = next((c for c in anchor if _local_name(c.tag) == "from"), None)
        if from_el is None:
            return ""
        col_text = ""
        row_text = ""
        for child in from_el:
            if _local_name(child.tag) == "col":
                col_text = child.text or ""
            elif _local_name(child.tag) == "row":
                row_text = child.text or ""
        if not col_text or not row_text:
            return ""
        return f"{get_column_letter(int(col_text) + 1)}{int(row_text) + 1}"
    except (ValueError, IndexError):
        return ""


def _first_descendant(el: ET.Element, local_name: str) -> ET.Element | None:
    """指定ローカル名の子孫要素を1件返す.

    Args:
        el: 探索開始要素。
        local_name: 名前空間を除いたタグ名。

    Returns:
        見つかった要素。なければ None。
    """
    return next((d for d in el.iter() if _local_name(d.tag) == local_name), None)


def _drawing_text(el: ET.Element) -> str:
    """DrawingML 要素配下のテキストを抽出する.

    Args:
        el: ``xdr:sp`` などの DrawingML 要素。

    Returns:
        結合・空白正規化したテキスト。
    """
    parts = [d.text or "" for d in el.iter() if _local_name(d.tag) == "t" and d.text]
    return re.sub(r"\s+", " ", "".join(parts)).strip()[:120]


def _control_kind(name: str, text: str, fallback: str = "control") -> str:
    """コントロール名や表示文字から種別を推定する.

    Args:
        name: コントロール名。
        text: 表示テキスト。
        fallback: 推定できない場合の種別。

    Returns:
        種別文字列。
    """
    value = f"{name} {text}".lower()
    if "button" in value or "ボタン" in value:
        return "button"
    if "check" in value or "チェック" in value:
        return "checkbox"
    if "option" in value or "radio" in value:
        return "radio"
    return fallback


def _extract_macro_from_control_props(props_bytes: bytes) -> str:
    """ctrlProp XML からマクロ名を抽出する.

    Args:
        props_bytes: ``xl/ctrlProps/ctrlProp*.xml`` の内容。

    Returns:
        マクロ名。取得できない場合は空文字。
    """
    try:
        root = ET.fromstring(props_bytes)
    except ET.ParseError:
        return ""
    for key, value in root.attrib.items():
        if _local_name(key).lower() == "macro" and value:
            return value.strip()
    for el in root.iter():
        if _local_name(el.tag).lower() in {"macro", "fmlamacro"} and el.text:
            return el.text.strip()
    return ""


def _drawing_control_macros(zf: zipfile.ZipFile, drawing_path: str) -> dict[str, str]:
    """DrawingML の relationship から control property のマクロを返す.

    Args:
        zf: xlsx/xlsm/xltm/xlsb を開いた ZipFile。
        drawing_path: ``xl/drawings/drawing*.xml`` のパス。

    Returns:
        relationship id をキー、マクロ名を値にしたマップ。
    """
    if "/" in drawing_path:
        dir_part, file_part = drawing_path.rsplit("/", 1)
    else:
        dir_part, file_part = "", drawing_path
    rels_path = f"{dir_part}/_rels/{file_part}.rels"
    try:
        rels_xml = zf.read(rels_path)
        rels_root = ET.fromstring(rels_xml)
    except (KeyError, ET.ParseError):
        return {}

    out: dict[str, str] = {}
    for rel in rels_root.iter(
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    ):
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        if "ctrlprop" not in target.lower() and not rel_type.endswith("/ctrlProp"):
            continue
        props_path = _resolve_package_part(dir_part, target)
        try:
            macro = _extract_macro_from_control_props(zf.read(props_path))
        except KeyError:
            continue
        if macro:
            out[rel.attrib.get("Id", "")] = macro
    return out


def _part_relationship_targets(
    zf: zipfile.ZipFile,
    part_path: str,
    predicate: Callable[[ET.Element], bool],
) -> dict[str, str]:
    """任意パーツの relationship から条件に合う `r:id -> target path` を返す.

    Args:
        zf: xlsx/xlsm を開いた ZipFile。
        part_path: relationship 所有パーツのパス。
        predicate: Relationship 要素を受け取り対象判定する関数。

    Returns:
        relationship id をキー、zip 内パーツパスを値にしたマップ。
    """
    if "/" in part_path:
        dir_part, file_part = part_path.rsplit("/", 1)
    else:
        dir_part, file_part = "", part_path
    rels_path = f"{dir_part}/_rels/{file_part}.rels"
    try:
        rels_xml = zf.read(rels_path)
        rels_root = ET.fromstring(rels_xml)
    except (KeyError, ET.ParseError):
        return {}

    out: dict[str, str] = {}
    for rel in rels_root.iter(
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    ):
        if not predicate(rel):
            continue
        rid = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if rid and target:
            out[rid] = _resolve_package_part(dir_part, target)
    return out


def _drawing_chart_targets(zf: zipfile.ZipFile, drawing_path: str) -> dict[str, str]:
    """DrawingML relationship から chart パーツを返す."""

    def _is_chart(rel: ET.Element) -> bool:
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        return target.lower().endswith(".xml") and (
            "charts/chart" in target.lower() or rel_type.endswith("/chart")
        )

    return _part_relationship_targets(zf, drawing_path, _is_chart)


def _direct_child(el: ET.Element, local_name: str) -> ET.Element | None:
    """指定ローカル名の直接子要素を1件返す."""
    return next((c for c in el if _local_name(c.tag) == local_name), None)


def _direct_children(el: ET.Element, local_name: str) -> list[ET.Element]:
    """指定ローカル名の直接子要素を全件返す."""
    return [c for c in el if _local_name(c.tag) == local_name]


def _first_formula(el: ET.Element) -> str:
    """要素配下の最初の `f` テキストを返す."""
    f = _first_descendant(el, "f")
    return (f.text or "").strip() if f is not None and f.text else ""


def _first_value(el: ET.Element) -> str:
    """要素配下の最初の `v` テキストを返す."""
    v = _first_descendant(el, "v")
    return (v.text or "").strip() if v is not None and v.text else ""


def _chart_text(el: ET.Element) -> str:
    """chart パーツ内のリッチテキストを結合して返す."""
    parts = [d.text or "" for d in el.iter() if _local_name(d.tag) == "t" and d.text]
    return re.sub(r"\s+", " ", "".join(parts)).strip()[:120]


def _parse_chart_xml(chart_bytes: bytes) -> tuple[str, str, list[ChartSeries]]:
    """chart*.xml からタイトル・種類・系列参照を抽出する.

    Returns:
        (title, chart_type, series)。
    """
    try:
        root = ET.fromstring(chart_bytes)
    except ET.ParseError:
        return "", "", []

    title = ""
    title_el = _first_descendant(root, "title")
    if title_el is not None:
        title = _chart_text(title_el)

    chart_type = ""
    for el in root.iter():
        local = _local_name(el.tag)
        if local.endswith("Chart") and local not in {"chart", "chartSpace"}:
            chart_type = local
            break

    series: list[ChartSeries] = []
    for ser in root.iter():
        if _local_name(ser.tag) != "ser":
            continue
        tx = _direct_child(ser, "tx")
        name = ""
        if tx is not None:
            name = _first_formula(tx) or _first_value(tx) or _chart_text(tx)

        cat = _direct_child(ser, "cat")
        val = _direct_child(ser, "val")
        series.append(
            ChartSeries(
                name=name,
                values_ref=_first_formula(val) if val is not None else "",
                categories_ref=_first_formula(cat) if cat is not None else "",
            )
        )
    return title, chart_type, series


def _extract_charts(file_path: Path, sheet_names: list[str]) -> dict[str, list[ChartObject]]:
    """Excel ファイルからシート別のグラフ情報を抽出する.

    DrawingML の chart relationship と chart*.xml を直接読む。取れるのは
    チャート種別・タイトル・配置セル・系列セル参照に限定する。
    """
    out: dict[str, list[ChartObject]] = {sn: [] for sn in sheet_names}
    if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return out
    try:
        with zipfile.ZipFile(file_path) as zf:
            sheet_to_drawings = _xlsm_sheet_to_drawing_map(zf)
            for sheet_name, drawing_paths in sheet_to_drawings.items():
                if sheet_name not in out:
                    continue
                for drawing_path in drawing_paths:
                    chart_targets = _drawing_chart_targets(zf, drawing_path)
                    if not chart_targets:
                        continue
                    try:
                        drawing_root = ET.fromstring(zf.read(drawing_path))
                    except (KeyError, ET.ParseError):
                        continue
                    for anchor in drawing_root.iter():
                        if _local_name(anchor.tag) not in {"twoCellAnchor", "oneCellAnchor"}:
                            continue
                        chart_el = _first_descendant(anchor, "chart")
                        if chart_el is None:
                            continue
                        rid = _relationship_id(chart_el)
                        chart_path = chart_targets.get(rid)
                        if not chart_path:
                            continue
                        c_nv_pr = _first_descendant(anchor, "cNvPr")
                        name = ""
                        if c_nv_pr is not None:
                            name = c_nv_pr.attrib.get("name", "") or c_nv_pr.attrib.get("descr", "")
                        try:
                            title, chart_type, series = _parse_chart_xml(zf.read(chart_path))
                        except KeyError:
                            continue
                        out[sheet_name].append(
                            ChartObject(
                                name=name,
                                chart_type=chart_type,
                                title=title,
                                anchor=_drawing_anchor_to_cell(anchor),
                                series=series,
                            )
                        )
    except (zipfile.BadZipFile, OSError) as e:
        logger.warning("failed to read charts from %s: %s", file_path.name, e)
    return out


def _xlsm_sheet_to_pivot_table_map(zf: zipfile.ZipFile) -> dict[str, list[str]]:
    """xlsx/xlsm zip から「シート名 → pivotTableDefinition パーツ一覧」を返す."""

    def _is_pivot(rel: ET.Element) -> bool:
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        return target.lower().endswith(".xml") and (
            "pivottables/pivottable" in target.lower() or rel_type.endswith("/pivotTable")
        )

    return _sheet_related_parts(zf, _is_pivot)


def _parse_pivot_cache(cache_bytes: bytes) -> tuple[str, str, str, str, list[str]]:
    """pivotCacheDefinition から元データとフィールド名を抽出する."""
    try:
        root = ET.fromstring(cache_bytes)
    except ET.ParseError:
        return "", "", "", "", []

    source_type = ""
    source_sheet = ""
    source_ref = ""
    source_name = ""
    cache_source = _first_descendant(root, "cacheSource")
    if cache_source is not None:
        source_type = cache_source.attrib.get("type", "")
        worksheet_source = _first_descendant(cache_source, "worksheetSource")
        if worksheet_source is not None:
            source_sheet = worksheet_source.attrib.get("sheet", "")
            source_ref = worksheet_source.attrib.get("ref", "")
            source_name = worksheet_source.attrib.get("name", "")

    fields = [
        field.attrib.get("name", "")
        for field in root.iter()
        if _local_name(field.tag) == "cacheField"
    ]
    return source_type, source_sheet, source_ref, source_name, fields


def _workbook_pivot_caches(
    zf: zipfile.ZipFile,
) -> dict[str, tuple[str, str, str, str, list[str]]]:
    """workbook.xml の pivotCaches から `cacheId -> cache info` を返す."""
    try:
        wb_xml = zf.read("xl/workbook.xml")
        wb_rels_xml = zf.read("xl/_rels/workbook.xml.rels")
        wb_root = ET.fromstring(wb_xml)
        rels_root = ET.fromstring(wb_rels_xml)
    except (KeyError, ET.ParseError):
        return {}

    rid_to_target: dict[str, str] = {}
    for rel in rels_root.iter(
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    ):
        rel_type = rel.attrib.get("Type", "")
        target = rel.attrib.get("Target", "")
        if rel_type.endswith("/pivotCacheDefinition") or "pivotcachedefinition" in target.lower():
            rid_to_target[rel.attrib.get("Id", "")] = _resolve_package_part("xl", target)

    out: dict[str, tuple[str, str, str, str, list[str]]] = {}
    for pivot_cache in wb_root.iter(
        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}pivotCache"
    ):
        cache_id = pivot_cache.attrib.get("cacheId", "")
        rid = _relationship_id(pivot_cache)
        cache_target = rid_to_target.get(rid)
        if not cache_id or not cache_target:
            continue
        try:
            out[cache_id] = _parse_pivot_cache(zf.read(cache_target))
        except KeyError:
            continue
    return out


def _field_name(fields: list[str], raw_index: str) -> str:
    """pivot field index をフィールド名へ変換する."""
    try:
        idx = int(raw_index)
    except ValueError:
        return raw_index
    if 0 <= idx < len(fields) and fields[idx]:
        return fields[idx]
    return raw_index


def _pivot_axis_fields(root: ET.Element, axis_name: str, fields: list[str]) -> list[str]:
    """rowFields / colFields / pageFields からフィールド名一覧を返す."""
    axis = _direct_child(root, axis_name)
    if axis is None:
        return []
    values: list[str] = []
    for field in _direct_children(axis, "field"):
        raw = field.attrib.get("x", "")
        if raw:
            values.append(_field_name(fields, raw))
    return values


def _pivot_value_fields(root: ET.Element, fields: list[str]) -> list[str]:
    """dataFields から値フィールド名一覧を返す."""
    data_fields = _direct_child(root, "dataFields")
    if data_fields is None:
        return []
    values: list[str] = []
    for field in _direct_children(data_fields, "dataField"):
        raw = field.attrib.get("fld", "")
        field_name = _field_name(fields, raw) if raw else ""
        display_name = field.attrib.get("name", "")
        subtotal = field.attrib.get("subtotal", "")
        label = display_name or field_name
        if field_name and display_name and field_name not in display_name:
            label = f"{display_name} ({field_name})"
        if subtotal and subtotal != "sum":
            label = f"{label} [{subtotal}]"
        if label:
            values.append(label)
    return values


def _parse_pivot_table_xml(
    pivot_bytes: bytes,
    cache_info: tuple[str, str, str, str, list[str]] | None,
) -> PivotTableInfo | None:
    """pivotTableDefinition を PivotTableInfo に変換する."""
    try:
        root = ET.fromstring(pivot_bytes)
    except ET.ParseError:
        return None

    name = root.attrib.get("name", "")
    cache_id = root.attrib.get("cacheId", "")
    if not name:
        return None

    source_type = source_sheet = source_ref = source_name = ""
    fields: list[str] = []
    if cache_info is not None:
        source_type, source_sheet, source_ref, source_name, fields = cache_info

    location = _direct_child(root, "location")
    anchor = location.attrib.get("ref", "") if location is not None else ""
    return PivotTableInfo(
        name=name,
        anchor=anchor,
        cache_id=cache_id,
        source_type=source_type,
        source_sheet=source_sheet,
        source_ref=source_ref,
        source_name=source_name,
        row_fields=_pivot_axis_fields(root, "rowFields", fields),
        column_fields=_pivot_axis_fields(root, "colFields", fields),
        value_fields=_pivot_value_fields(root, fields),
        filter_fields=_pivot_axis_fields(root, "pageFields", fields),
    )


def _extract_pivot_tables(
    file_path: Path,
    sheet_names: list[str],
) -> dict[str, list[PivotTableInfo]]:
    """Excel ファイルからシート別のピボットテーブル情報を抽出する."""
    out: dict[str, list[PivotTableInfo]] = {sn: [] for sn in sheet_names}
    if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return out
    try:
        with zipfile.ZipFile(file_path) as zf:
            cache_map = _workbook_pivot_caches(zf)
            sheet_to_pivots = _xlsm_sheet_to_pivot_table_map(zf)
            for sheet_name, pivot_paths in sheet_to_pivots.items():
                if sheet_name not in out:
                    continue
                for pivot_path in pivot_paths:
                    try:
                        pivot_root = ET.fromstring(zf.read(pivot_path))
                    except (KeyError, ET.ParseError):
                        continue
                    cache_id = pivot_root.attrib.get("cacheId", "")
                    info = _parse_pivot_table_xml(
                        ET.tostring(pivot_root),
                        cache_map.get(cache_id),
                    )
                    if info is not None:
                        out[sheet_name].append(info)
    except (zipfile.BadZipFile, OSError) as e:
        logger.warning("failed to read pivot tables from %s: %s", file_path.name, e)
    return out


_SENSITIVE_CONNECTION_RE = re.compile(
    r"(?i)\b(password|pwd|token|access\s*token|secret|api[_ -]?key|client[_ -]?secret)"
    r"\s*=\s*[^;,\s]+"
)


def _redact_connection_text(text: str) -> str:
    """接続文字列から秘匿値をマスクする."""
    if not text:
        return ""
    redacted = _SENSITIVE_CONNECTION_RE.sub(lambda m: f"{m.group(1)}=***", text)
    redacted = re.sub(r"(?i)([?&](?:password|pwd|token|key|secret)=)[^&#]+", r"\1***", redacted)
    return redacted[:1000]


def _connection_kind(name: str, source: str, command: str) -> Literal["power_query", "connection"]:
    """接続名・接続文字列から Power Query らしさを判定する."""
    haystack = f"{name}\n{source}\n{command}".lower()
    if (
        name.lower().startswith("query - ")
        or "microsoft.mashup" in haystack
        or "mashup" in haystack
    ):
        return "power_query"
    return "connection"


def _parse_connections_xml(connections_bytes: bytes) -> dict[str, PowerQueryInfo]:
    """connections.xml から Power Query / 外部接続を抽出する."""
    try:
        root = ET.fromstring(connections_bytes)
    except ET.ParseError:
        return {}

    out: dict[str, PowerQueryInfo] = {}
    for conn in root.iter():
        if _local_name(conn.tag) != "connection":
            continue
        cid = conn.attrib.get("id", "")
        name = conn.attrib.get("name", "") or cid
        connection_type = conn.attrib.get("type", "")
        description = conn.attrib.get("description", "")
        refresh_on_load = conn.attrib.get("refreshOnLoad", "0") in {"1", "true", "True"}
        source = ""
        command = ""
        for child in conn:
            local = _local_name(child.tag)
            if local in {"dbPr", "olapPr"}:
                source = child.attrib.get("connection", "") or source
                command = child.attrib.get("command", "") or command
            elif local == "webPr":
                source = child.attrib.get("url", "") or source
            elif local == "textPr":
                source = child.attrib.get("sourceFile", "") or source
        source = _redact_connection_text(source)
        command = _redact_connection_text(command)
        out[cid] = PowerQueryInfo(
            name=name,
            kind=_connection_kind(name, source, command),
            connection_id=cid,
            connection_type=connection_type,
            description=description,
            refresh_on_load=refresh_on_load,
            source=source,
            command=command,
            confidence="explicit",
        )
    return out


def _xlsm_sheet_to_query_table_map(zf: zipfile.ZipFile) -> dict[str, list[str]]:
    """xlsx/xlsm zip から「シート名 → queryTable パーツ一覧」を返す."""

    def _is_query_table(rel: ET.Element) -> bool:
        target = rel.attrib.get("Target", "")
        rel_type = rel.attrib.get("Type", "")
        return target.lower().endswith(".xml") and (
            "querytables/querytable" in target.lower() or rel_type.endswith("/queryTable")
        )

    return _sheet_related_parts(zf, _is_query_table)


def _attach_query_table_targets(
    zf: zipfile.ZipFile,
    connections: dict[str, PowerQueryInfo],
) -> None:
    """queryTable パーツから接続の出力先シート/名前を補完する."""
    sheet_to_query_tables = _xlsm_sheet_to_query_table_map(zf)
    for sheet_name, query_paths in sheet_to_query_tables.items():
        for query_path in query_paths:
            try:
                root = ET.fromstring(zf.read(query_path))
            except (KeyError, ET.ParseError):
                continue
            cid = root.attrib.get("connectionId", "")
            query_name = root.attrib.get("name", "")
            refresh_on_load = root.attrib.get("refreshOnLoad", "0") in {"1", "true", "True"}
            if cid in connections:
                current = connections[cid]
                connections[cid] = current.model_copy(
                    update={
                        "target_sheet": sheet_name,
                        "target_name": query_name,
                        "refresh_on_load": current.refresh_on_load or refresh_on_load,
                    }
                )
            elif query_name:
                connections[f"queryTable:{query_path}"] = PowerQueryInfo(
                    name=query_name,
                    kind="connection",
                    connection_id=cid,
                    target_sheet=sheet_name,
                    target_name=query_name,
                    refresh_on_load=refresh_on_load,
                    confidence="explicit" if cid else "unknown",
                )


def _extract_power_queries(file_path: Path) -> list[PowerQueryInfo]:
    """Power Query / 外部接続の棚卸し情報を抽出する.

    初期対応では `xl/connections.xml` と queryTable の出力先だけを読む。
    M コード解析は保存形式差分が大きいためここでは行わない。
    """
    if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return []
    try:
        with zipfile.ZipFile(file_path) as zf:
            try:
                connections = _parse_connections_xml(zf.read("xl/connections.xml"))
            except KeyError:
                connections = {}
            _attach_query_table_targets(zf, connections)
            return list(connections.values())
    except (zipfile.BadZipFile, OSError) as e:
        logger.warning("failed to read data connections from %s: %s", file_path.name, e)
        return []


def _parse_drawing_form_controls(
    drawing_bytes: bytes,
    control_macros: dict[str, str] | None = None,
) -> list[FormControl]:
    """DrawingML からフォームコントロール相当のボタンを抽出する.

    Args:
        drawing_bytes: ``xl/drawings/drawing*.xml`` の内容。
        control_macros: drawing rels から得た ``r:id -> macro`` マップ。

    Returns:
        抽出したフォームコントロール一覧。
    """
    out: list[FormControl] = []
    control_macros = control_macros or {}
    try:
        root = ET.fromstring(drawing_bytes)
    except ET.ParseError:
        return out

    for anchor in root.iter():
        if _local_name(anchor.tag) not in {"twoCellAnchor", "oneCellAnchor"}:
            continue
        control_el = _first_descendant(anchor, "control")
        control_name = control_el.attrib.get("name", "") if control_el is not None else ""
        control_rid = _relationship_id(control_el) if control_el is not None else ""
        control_macro = control_macros.get(control_rid, "")

        for shape in (d for d in anchor.iter() if _local_name(d.tag) == "sp"):
            macro = (shape.attrib.get("macro", "") or control_macro).strip()
            c_nv_pr = _first_descendant(shape, "cNvPr")
            name = ""
            if c_nv_pr is not None:
                name = c_nv_pr.attrib.get("name", "") or c_nv_pr.attrib.get("descr", "")
            if not name:
                name = control_name
            text = _drawing_text(shape)

            # 通常の図形をフォームとして誤検出しないよう、macro または xdr:control が
            # あるものだけフォームコントロール候補にする。
            if not macro and control_el is None:
                continue
            if not macro and not text and not name:
                continue

            out.append(
                FormControl(
                    kind=_control_kind(name, text, "control"),
                    name=name,
                    text=text,
                    macro=macro,
                    anchor=_drawing_anchor_to_cell(anchor),
                )
            )
    return out


def _dedupe_form_controls(controls: list[FormControl]) -> list[FormControl]:
    """同一コントロールの重複を取り除く.

    Args:
        controls: フォームコントロール一覧。

    Returns:
        入力順を保った重複除去後の一覧。
    """
    seen: set[tuple[str, str, str, str, str]] = set()
    out: list[FormControl] = []
    for control in controls:
        key = (control.kind, control.name, control.text, control.macro, control.anchor)
        if key in seen:
            continue
        seen.add(key)
        out.append(control)
    return out


def _extract_form_controls(file_path: Path, sheet_names: list[str]) -> dict[str, list[FormControl]]:
    """Excel ファイルから (シート名 → フォームコントロール) を抽出する.

    VML 形式と DrawingML + ctrlProps 形式の両方を best-effort で読む。

    Args:
        file_path: 対象 Excel ファイル。
        sheet_names: 抽出対象シート名。

    Returns:
        シート名をキー、フォームコントロール一覧を値にしたマップ。
    """
    out: dict[str, list[FormControl]] = {sn: [] for sn in sheet_names}
    if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xlsb"}:
        return out
    try:
        with zipfile.ZipFile(file_path) as zf:
            sheet_to_vml = _xlsm_sheet_to_vml_map(zf)
            for sheet_name, vml_paths in sheet_to_vml.items():
                if sheet_name not in out:
                    continue
                for vml_path in vml_paths:
                    try:
                        vml_bytes = zf.read(vml_path)
                    except KeyError:
                        continue
                    out[sheet_name].extend(_parse_vml_form_controls(vml_bytes))
            sheet_to_drawings = _xlsm_sheet_to_drawing_map(zf)
            for sheet_name, drawing_paths in sheet_to_drawings.items():
                if sheet_name not in out:
                    continue
                for drawing_path in drawing_paths:
                    try:
                        drawing_bytes = zf.read(drawing_path)
                    except KeyError:
                        continue
                    out[sheet_name].extend(
                        _parse_drawing_form_controls(
                            drawing_bytes,
                            _drawing_control_macros(zf, drawing_path),
                        )
                    )
            for sheet_name, controls in out.items():
                out[sheet_name] = _dedupe_form_controls(controls)
    except (zipfile.BadZipFile, OSError) as e:
        logger.warning("failed to read form controls from %s: %s", file_path.name, e)
    return out


def _extract_merged_ranges(ws: object) -> list[str]:
    """シートのマージ範囲を文字列リストで返す."""
    result: list[str] = []
    merged_attr = getattr(ws, "merged_cells", None)
    if merged_attr is None:
        return result
    try:
        for rng in merged_attr.ranges:
            result.append(str(rng))
    except Exception:  # noqa: BLE001
        logger.debug("Failed to enumerate merged_cells")
    return result


def _extract_preview(ws: object) -> tuple[list[list[str | None]], str]:
    """シート冒頭の N 行 × M 列を literal に取得する.

    Returns:
        (preview_rows, origin) のタプル. preview_rows は等長の2次元リスト、
        各要素は文字列化したセル値か None (空セル). origin は "A1" 等の起点座標.
    """
    from openpyxl.utils import get_column_letter

    max_row = min(getattr(ws, "max_row", 0) or 0, PREVIEW_MAX_ROWS)
    max_col = min(getattr(ws, "max_column", 0) or 0, PREVIEW_MAX_COLS)
    if max_row == 0 or max_col == 0:
        return [], ""

    rows: list[list[str | None]] = []
    try:
        for raw_row in ws.iter_rows(  # type: ignore[attr-defined]
            min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
        ):
            row_vals: list[str | None] = []
            for v in raw_row:
                row_vals.append(None if v is None else str(v))
            rows.append(row_vals)
    except Exception:  # noqa: BLE001
        logger.debug("Failed to extract preview rows")
        return [], ""

    origin = f"A1:{get_column_letter(max_col)}{max_row}"
    return rows, origin


def extract_workbook(file_path: Path) -> Workbook:
    """Excelファイルからシート構造を抽出する.

    Args:
        file_path: 対象ファイルのパス. .xlsx / .xlsm を想定.

    Returns:
        Workbook モデル. VBA モジュールは含めない (vba_modules は空).
        `.xls` (旧バイナリ形式) が渡された場合は sheets=[] で返し警告ログを出す.

    Raises:
        ExtractionError: ファイルが存在しない、または openpyxl が開けない場合.
    """
    if not file_path.exists():
        raise ExtractionError(f"File not found: {file_path}")

    if file_path.suffix.lower() == ".xls":
        logger.warning(
            ".xls (legacy binary) is not supported by openpyxl; "
            "returning empty workbook structure for %s",
            file_path.name,
        )
        return Workbook(filename=file_path.name)

    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=str(file_path), keep_vba=True, data_only=False)
    except Exception as e:  # noqa: BLE001
        raise ExtractionError(f"openpyxl failed to open {file_path}: {e}") from e

    sheets: list[SheetInfo] = []
    sheets_by_name: dict[str, SheetInfo] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        preview_rows, preview_origin = _extract_preview(ws)
        info = SheetInfo(
            name=sn,
            rows=ws.max_row or 0,
            cols=ws.max_column or 0,
            formulas=_extract_formulas(ws),
            conditional_formats=_extract_conditional_formats(ws),
            tables=_extract_excel_tables(ws),
            merged_ranges=_extract_merged_ranges(ws),
            data_validations=_extract_data_validations(ws),
            preview_rows=preview_rows,
            preview_origin=preview_origin,
        )
        sheets.append(info)
        sheets_by_name[sn] = info

    _attach_named_ranges(wb, sheets_by_name)

    chart_map = _extract_charts(file_path, list(sheets_by_name.keys()))
    for sn, charts in chart_map.items():
        if sn in sheets_by_name and charts:
            sheets_by_name[sn].charts = charts

    pivot_map = _extract_pivot_tables(file_path, list(sheets_by_name.keys()))
    for sn, pivot_tables in pivot_map.items():
        if sn in sheets_by_name and pivot_tables:
            sheets_by_name[sn].pivot_tables = pivot_tables

    # フォームコントロール (ボタン → マクロ紐付け) は xlsm の VML を直接読む.
    # openpyxl は VML を解釈しないので zipfile + XML パースで自前抽出.
    fc_map = _extract_form_controls(file_path, list(sheets_by_name.keys()))
    for sn, controls in fc_map.items():
        if sn in sheets_by_name and controls:
            sheets_by_name[sn].form_controls = controls

    return Workbook(
        filename=file_path.name,
        sheets=sheets,
        external_links=_extract_external_links(wb),
        power_queries=_extract_power_queries(file_path),
    )
