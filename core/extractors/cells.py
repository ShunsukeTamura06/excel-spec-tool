"""全セルを SQLite に書き出すモジュール.

設計書本体には載せない「生データ」を、LLM がツール経由で参照できる形で
永続化する。詳細は本ファイル先頭の docstring を参照。

スキーマ:
    cells(sheet, row, col, coord, value, data_type, formula, number_format)
    merged_ranges(sheet, range)

- 計算値とフォーミュラを両方持つために openpyxl を 2 度ロードする
  (`data_only=True` でキャッシュ計算値, `data_only=False` で数式)。
- 値は全て TEXT (文字列化) で格納する。元の型は `data_type` カラムに残す。
- `.xls` は対象外 (openpyxl で読めないためスキップ)。
"""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from core.exceptions import ExtractionError

logger = logging.getLogger(__name__)


SCHEMA = """
CREATE TABLE IF NOT EXISTS cells (
    sheet TEXT NOT NULL,
    row INTEGER NOT NULL,
    col INTEGER NOT NULL,
    coord TEXT NOT NULL,
    value TEXT,
    data_type TEXT,
    formula TEXT,
    number_format TEXT,
    PRIMARY KEY (sheet, row, col)
);
CREATE INDEX IF NOT EXISTS idx_cells_sheet_row ON cells(sheet, row);
CREATE INDEX IF NOT EXISTS idx_cells_value ON cells(value);

CREATE TABLE IF NOT EXISTS merged_ranges (
    sheet TEXT NOT NULL,
    range TEXT NOT NULL,
    PRIMARY KEY (sheet, range)
);
"""


def _stringify(v: Any) -> str | None:
    if v is None:
        return None
    return str(v)


def extract_cells_to_sqlite(file_path: Path, db_path: Path) -> int:
    """xlsx/xlsm の全セルを SQLite に書き出す.

    Args:
        file_path: 入力ファイル. .xls は ExtractionError。
        db_path: 出力 SQLite ファイルパス. 既存ならテーブルだけクリアして再作成。

    Returns:
        書き込んだセル数.

    Raises:
        ExtractionError: ファイル不在 / .xls / openpyxl オープン失敗.
    """
    if not file_path.exists():
        raise ExtractionError(f"File not found: {file_path}")
    if file_path.suffix.lower() == ".xls":
        raise ExtractionError(".xls (legacy binary) is not supported; cells.db cannot be built")

    from openpyxl import load_workbook

    # 数式テキストを取るための pass (data_only=False)
    # read_only=True は merged_cells を読まないので、ここでは非 read_only で開く
    try:
        wb_f = load_workbook(filename=str(file_path), data_only=False)
    except Exception as e:  # noqa: BLE001
        raise ExtractionError(f"openpyxl failed to open (formula pass) {file_path}: {e}") from e

    # 計算結果を取るための pass (data_only=True)
    try:
        wb_v = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        wb_f.close()
        raise ExtractionError(f"openpyxl failed to open (value pass) {file_path}: {e}") from e

    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(str(db_path))
    try:
        conn.executescript(SCHEMA)
        total = 0
        for sheet_name in wb_f.sheetnames:
            if sheet_name not in wb_v.sheetnames:
                continue
            ws_f = wb_f[sheet_name]
            ws_v = wb_v[sheet_name]

            # マージ範囲
            try:
                merged = list(ws_f.merged_cells.ranges)
                if merged:
                    conn.executemany(
                        "INSERT OR IGNORE INTO merged_ranges(sheet, range) VALUES (?, ?)",
                        [(sheet_name, str(r)) for r in merged],
                    )
            except Exception:  # noqa: BLE001
                logger.debug("Failed to read merged_cells on %s", sheet_name)

            # セル本体: 数式 pass と 値 pass を平行に走査
            rows_v_iter = ws_v.iter_rows()
            for row_f in ws_f.iter_rows():
                try:
                    row_v = next(rows_v_iter)
                except StopIteration:
                    row_v = ()
                # 列方向に zip。長さがズレることがあるので max を取る
                pairs = list(_zip_longest_cells(row_f, row_v))
                buffer = []
                for cell_f, cell_v in pairs:
                    if cell_f is None and cell_v is None:
                        continue
                    # read_only モードでは EmptyCell が混じるため、value で判定する
                    value_f = getattr(cell_f, "value", None) if cell_f is not None else None
                    value_v = getattr(cell_v, "value", None) if cell_v is not None else None
                    if value_f is None and value_v is None:
                        continue

                    # row/col を頑健に取得 (EmptyCell でも row/column 属性は持つ)
                    base = cell_f if cell_f is not None else cell_v
                    row = getattr(base, "row", None)
                    col = getattr(base, "column", None)
                    if row is None or col is None:
                        continue
                    coord = f"{get_column_letter(int(col))}{int(row)}"

                    data_type_f = getattr(cell_f, "data_type", None) if cell_f else None
                    raw_formula = value_f if data_type_f == "f" else None
                    # 値: 計算結果 (data_only) を優先、無ければ数式 pass の値
                    value: Any = (
                        value_v
                        if value_v is not None
                        else (value_f if data_type_f != "f" else None)
                    )
                    number_format = getattr(cell_f, "number_format", None) if cell_f else None
                    buffer.append(
                        (
                            sheet_name,
                            int(row),
                            int(col),
                            coord,
                            _stringify(value),
                            data_type_f,
                            _stringify(raw_formula),
                            number_format if number_format and number_format != "General" else None,
                        )
                    )
                if buffer:
                    conn.executemany(
                        "INSERT OR REPLACE INTO cells"
                        "(sheet, row, col, coord, value, data_type, formula, number_format) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                        buffer,
                    )
                    total += len(buffer)
        conn.commit()
        return total
    finally:
        conn.close()
        wb_f.close()
        wb_v.close()


def _zip_longest_cells(row_f: Any, row_v: Any) -> Any:
    """formula pass と value pass のセル列を列番号で揃えて zip する.

    iter_rows は普通同じ長さの列を返すが、念のため。
    """
    list_f = list(row_f)
    list_v = list(row_v)
    n = max(len(list_f), len(list_v))
    for i in range(n):
        yield (
            list_f[i] if i < len(list_f) else None,
            list_v[i] if i < len(list_v) else None,
        )
