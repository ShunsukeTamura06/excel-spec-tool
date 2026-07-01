"""xlsx/xlsm 正規化 diff のノイズ除去 実現可能性スパイク.

docs/VISION.ja.md §6 の前提3「xlsm 正規化 diff のノイズ除去
(calcChain/sharedStrings/style/rel id/pivot cache) が実用精度に届くか」を検証する。

このスパイクは Windows / Excel を必要としない (openpyxl のみで完結する)。
このPC (Mac/Linux) 上でそのまま実行できる。

検証する仮説:
    xlsx の生 XML (zip 内パーツ) をそのまま diff すると、実際の意味的な変更が
    なくても保存のたびに calcChain.xml / sharedStrings.xml / styles.xml /
    .rels の id 等が変化し、ノイズだらけになる。
    一方、既存の `core.extractors.cells.extract_cells_to_sqlite` (セル単位の
    正規化抽出: シート/座標/値/数式/表示形式) を経由して比較すれば、
    このノイズを素通りできるのではないか。

実行例:
    uv run python spikes/xlsx_diff_noise/probe.py
"""

from __future__ import annotations

import argparse
import json
import logging
import shutil
import sqlite3
import sys
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from core.extractors.cells import extract_cells_to_sqlite  # noqa: E402

logger = logging.getLogger("xlsx_diff_noise")


@dataclass
class CaseResult:
    """1つの比較ケース (baseline vs variant) の結果."""

    name: str
    description: str
    raw_parts_changed: list[str] = field(default_factory=list)
    semantic_cells_changed: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "raw_parts_changed_count": len(self.raw_parts_changed),
            "raw_parts_changed": self.raw_parts_changed,
            "semantic_cells_changed_count": len(self.semantic_cells_changed),
            "semantic_cells_changed": self.semantic_cells_changed,
        }


def _diff_raw_zip_parts(before: Path, after: Path) -> list[str]:
    """2つの xlsx の zip 内パーツをバイト単位で比較し、内容が変わったパーツ名を返す."""

    with zipfile.ZipFile(before) as zf_before, zipfile.ZipFile(after) as zf_after:
        names_before = set(zf_before.namelist())
        names_after = set(zf_after.namelist())
        changed: list[str] = sorted(names_before ^ names_after)
        for name in sorted(names_before & names_after):
            if zf_before.read(name) != zf_after.read(name):
                changed.append(name)
    return sorted(set(changed))


def _dump_cells(xlsx_path: Path, db_path: Path) -> dict[str, tuple[Any, Any, Any]]:
    """セル正規化抽出 (extract_cells_to_sqlite) を実行し、座標をキーにした辞書を返す."""

    extract_cells_to_sqlite(xlsx_path, db_path)
    conn = sqlite3.connect(str(db_path))
    try:
        rows = conn.execute(
            "SELECT sheet, coord, value, formula, number_format FROM cells"
        ).fetchall()
    finally:
        conn.close()
    return {
        f"{sheet}!{coord}": (value, formula, number_format)
        for sheet, coord, value, formula, number_format in rows
    }


def _diff_semantic_cells(before: Path, after: Path, work_dir: Path) -> list[str]:
    """extract_cells_to_sqlite の結果を比較し、値/数式/表示形式が変わったセル座標を返す."""

    cells_before = _dump_cells(before, work_dir / "before.db")
    cells_after = _dump_cells(after, work_dir / "after.db")
    changed: list[str] = []
    for key in cells_before.keys() | cells_after.keys():
        if cells_before.get(key) != cells_after.get(key):
            changed.append(key)
    return sorted(changed)


_REPO_ROOT = Path(__file__).resolve().parents[2]
_SAMPLE_XLSX = _REPO_ROOT / "frontend" / "public" / "samples" / "retail_monthly_ops.xlsx"


def _make_baseline(work_dir: Path) -> Path:
    """既存のリポジトリ同梱サンプルをコピーしてベースラインにする.

    `scripts/make_sample.py` の `main()` は同じ固定パスに直接書き込むため、ここで
    再生成すると git 管理下のサンプルファイルを上書きしてしまう。既に生成済みの
    ファイルをコピーするだけに留める (再生成したい場合は事前に
    `uv run python scripts/make_sample.py` を手動実行する)。
    """

    if not _SAMPLE_XLSX.exists():
        raise FileNotFoundError(
            f"サンプルが見つかりません: {_SAMPLE_XLSX}。"
            "先に `uv run python scripts/make_sample.py` を実行してください。"
        )
    baseline = work_dir / "baseline.xlsx"
    shutil.copy2(_SAMPLE_XLSX, baseline)
    return baseline


def _make_noop_resave(baseline: Path, work_dir: Path) -> Path:
    """意味的な変更を一切加えず、openpyxl で読み込んで保存し直すだけのファイルを作る.

    これが「保存し直しただけで生 XML に何が起きるか (=ノイズ)」の観測対象。
    """

    wb = load_workbook(baseline)
    out = work_dir / "variant_resave.xlsx"
    wb.save(out)
    wb.close()
    return out


def _make_single_edit(
    baseline: Path, work_dir: Path, sheet: str, coord: str, new_formula: str
) -> Path:
    """1セルの数式だけを意図的に書き換えたファイルを作る (=検知したい本物の変更)."""

    wb = load_workbook(baseline)
    ws = wb[sheet]
    ws[coord] = new_formula
    out = work_dir / "variant_edit.xlsx"
    wb.save(out)
    wb.close()
    return out


def run(work_dir: Path, sheet: str, coord: str, new_formula: str) -> dict[str, Any]:
    """全ケースを実行しレポートを組み立てる."""

    baseline = _make_baseline(work_dir)
    logger.info("baseline: %s", baseline)

    cases: list[CaseResult] = []

    resave = _make_noop_resave(baseline, work_dir)
    case = CaseResult(
        name="noop_resave",
        description="意味的な変更なしで保存し直しただけ (=ノイズのみのはずのケース)",
    )
    case.raw_parts_changed = _diff_raw_zip_parts(baseline, resave)
    case.semantic_cells_changed = _diff_semantic_cells(baseline, resave, work_dir / "noop")
    cases.append(case)
    logger.info(
        "[noop_resave] raw_parts=%d semantic_cells=%d",
        len(case.raw_parts_changed),
        len(case.semantic_cells_changed),
    )

    edited = _make_single_edit(baseline, work_dir, sheet, coord, new_formula)
    case = CaseResult(
        name="single_edit",
        description=f"{sheet}!{coord} のみ '{new_formula}' に書き換えたケース (=本物の変更1件)",
    )
    case.raw_parts_changed = _diff_raw_zip_parts(baseline, edited)
    case.semantic_cells_changed = _diff_semantic_cells(baseline, edited, work_dir / "edit")
    cases.append(case)
    logger.info(
        "[single_edit] raw_parts=%d semantic_cells=%d changed=%s",
        len(case.raw_parts_changed),
        len(case.semantic_cells_changed),
        case.semantic_cells_changed,
    )

    return {
        "baseline": str(baseline),
        "cases": [c.as_dict() for c in cases],
        "verdict": _verdict(cases, expected_cell=f"{sheet}!{coord}"),
    }


def _verdict(cases: list[CaseResult], expected_cell: str) -> dict[str, Any]:
    """仮説 (正規化抽出経由の diff ならノイズを素通りできる) が成立したかを機械的に判定する."""

    noop = next(c for c in cases if c.name == "noop_resave")
    edit = next(c for c in cases if c.name == "single_edit")

    noop_has_raw_noise = len(noop.raw_parts_changed) > 0
    noop_semantic_clean = len(noop.semantic_cells_changed) == 0
    edit_semantic_isolated = edit.semantic_cells_changed == [expected_cell]

    return {
        "noop_resave_produces_raw_xml_noise": noop_has_raw_noise,
        "noop_resave_semantic_diff_is_clean": noop_semantic_clean,
        "single_edit_semantic_diff_isolates_exact_cell": edit_semantic_isolated,
        "hypothesis_confirmed": noop_has_raw_noise
        and noop_semantic_clean
        and edit_semantic_isolated,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="xlsx 正規化 diff ノイズ除去スパイク")
    parser.add_argument("--out-dir", default="spike_out", help="作業/出力先ディレクトリ")
    parser.add_argument("--sheet", default="設定", help="編集対象シート名")
    parser.add_argument("--coord", default="B4", help="編集対象セル座標")
    parser.add_argument(
        "--new-formula",
        default="=1-0.05",
        help="編集対象セルに書き込む数式 (意味は変えて値は変わらない例など)",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    work_dir = Path(args.out_dir)
    work_dir.mkdir(parents=True, exist_ok=True)

    report = run(work_dir, args.sheet, args.coord, args.new_formula)
    report_path = work_dir / "report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    logger.info("=" * 60)
    logger.info("report: %s", report_path.resolve())
    logger.info("verdict: %s", json.dumps(report["verdict"], ensure_ascii=False))
    logger.info("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
